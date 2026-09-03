"""匯出→匯入 round-trip：/api/export 匯出的 xlsx 必須能原封不動再匯入，且關鍵欄位一致。
（先前 export 用 field_meta label 當表頭、跟 import 期待的 field_mapping 標題對不上，round-trip 會失敗。）
"""
import io
import sys
import tempfile
from pathlib import Path

ROOT = Path(__file__).resolve().parents[2]
sys.path.insert(0, str(ROOT / "APP" / "asset-module" / "backend"))

import auth  # noqa: E402
import db  # noqa: E402
import api  # noqa: E402
from excel_import import import_excel  # noqa: E402
from fastapi.testclient import TestClient  # noqa: E402

_PW = "test-password-123"


def _client(tmp):
    db_path = Path(tmp) / "src.db"
    db.init_db(db_path)

    def _override_get_db():
        conn = db.get_connection(db_path)
        try:
            yield conn
        finally:
            conn.close()

    api.app.dependency_overrides[api.get_db] = _override_get_db
    return TestClient(api.app), db_path


def _login(client, db_path):
    conn = db.get_connection(db_path)
    try:
        db.create_user(conn, "tester", auth.hash_password(_PW))
    finally:
        conn.close()
    assert client.post("/api/auth/login", json={"username": "tester", "password": _PW}).status_code == 200


def test_export_can_be_reimported():
    with tempfile.TemporaryDirectory() as tmp:
        client, src_db = _client(tmp)
        try:
            _login(client, src_db)
            # 種一台硬體 + 綁一筆人員/軟體
            conn = db.get_connection(src_db)
            db.insert_hardware(
                conn, asset_serial="RT-001", hostname="rt-host01", ip="10.9.9.9",
                api_id="core_x", owner="王五", custodian="趙六", environment="正式",
                device_model="Dell R760", os="RHEL 8.9", integrity=3, confidentiality=2, availability=1,
            )
            conn.execute(
                "INSERT INTO personnel (asset_serial, person_name, phone) VALUES ('RT-001','王五','0900')"
            )
            conn.execute(
                "INSERT INTO software (asset_serial, asset_name, db_software) VALUES ('RT-001','核心AP','Oracle')"
            )
            conn.commit()
            conn.close()

            # 匯出
            resp = client.get("/api/export")
            assert resp.status_code == 200
            xlsx = Path(tmp) / "exp.xlsx"
            xlsx.write_bytes(resp.content)

            # 匯進一個全新的 DB
            dst_db = Path(tmp) / "dst.db"
            db.init_db(dst_db)
            conn2 = db.get_connection(dst_db)
            summary = import_excel(xlsx, conn2)

            # 硬體關鍵欄位要一致
            row = conn2.execute("SELECT * FROM hardware WHERE asset_serial='RT-001'").fetchone()
            assert row is not None, f"round-trip 後找不到硬體，summary={summary}"
            assert row["hostname"] == "rt-host01"
            assert row["ip"] == "10.9.9.9"
            assert row["api_id"] == "core_x"
            assert row["owner"] == "王五"
            assert row["custodian"] == "趙六"
            assert row["environment"] == "正式"
            # 人員/軟體也要進得來
            assert conn2.execute("SELECT COUNT(*) FROM personnel WHERE asset_serial='RT-001'").fetchone()[0] == 1
            assert conn2.execute("SELECT COUNT(*) FROM software WHERE asset_serial='RT-001'").fetchone()[0] == 1
            assert summary["sheets"]["硬體"]["inserted"] == 1
            conn2.close()
        finally:
            api.app.dependency_overrides.clear()


# ===== 帳外資產（DYN-/VC-/AUTO-）預設不匯出、勾選才含 =====
#
# 2026-08-25 使用者：「要交出新的 CIA 資產清單，也是要完全準確」，並提出一個
# 開放式設計問題——帳外資產（存活清單掃到／vCenter 收到、實際存在但沒登記的）
# 要不要一起交給資產清單維護單位。這是使用者要拍板的事，這裡不是替他決定答案，
# 是把決定權交給匯出當下的一個參數，並驗證兩種模式都做對。

def _seed_mixed(conn):
    db.insert_hardware(conn, asset_serial="HW-CIA-1", hostname="cia-host", ip="10.9.1.1")
    db.insert_hardware(conn, asset_serial="DYN-abc", hostname="dyn-host", ip="10.9.1.2")
    db.insert_hardware(conn, asset_serial="VC-xyz", hostname="vc-host", ip="10.9.1.3")
    conn.commit()


def _read_hardware_sheet(content: bytes):
    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(content))
    ws = wb["硬體"]
    rows = list(ws.iter_rows(values_only=True))
    header = list(rows[0])
    return header, rows[1:]


def test_預設不含帳外資產_只有CIA登記的():
    with tempfile.TemporaryDirectory() as tmp:
        client, src_db = _client(tmp)
        try:
            _login(client, src_db)
            conn = db.get_connection(src_db)
            _seed_mixed(conn)
            conn.close()

            resp = client.get("/api/export")
            assert resp.status_code == 200
            header, rows = _read_hardware_sheet(resp.content)
            serial_idx = header.index("資產序號") if "資產序號" in header else \
                next(i for i, h in enumerate(header) if h and "序號" in h)
            serials = {r[serial_idx] for r in rows}
            assert serials == {"HW-CIA-1"}, f"預設不該含帳外資產，實際：{serials}"
        finally:
            api.app.dependency_overrides.clear()


def test_勾選後含帳外資產_且來源欄標示正確():
    with tempfile.TemporaryDirectory() as tmp:
        client, src_db = _client(tmp)
        try:
            _login(client, src_db)
            conn = db.get_connection(src_db)
            _seed_mixed(conn)
            conn.close()

            resp = client.get("/api/export", params={"include_off_book": "true"})
            assert resp.status_code == 200
            header, rows = _read_hardware_sheet(resp.content)
            assert header[-1] == "來源"
            by_serial = {r[0] if "資產序號" not in header else r[header.index("資產序號")]: r[-1]
                        for r in rows}
            # 用主機名對照，比對序號欄位位置更穩定
            hostname_idx = header.index("主機名稱") if "主機名稱" in header else \
                next(i for i, h in enumerate(header) if h and "主機" in h)
            source_by_host = {r[hostname_idx]: r[-1] for r in rows}
            assert source_by_host == {"cia-host": "CIA", "dyn-host": "DYN", "vc-host": "VC"}
        finally:
            api.app.dependency_overrides.clear()


def test_人員軟體分頁的帳外資產也一併排除():
    """人員/軟體是靠 asset_serial 掛在硬體底下的附屬資料，沒有獨立的帳外判定——
    篩選要跟硬體分頁一致，不能硬體排除了、人員那筆卻漏網留下來。"""
    with tempfile.TemporaryDirectory() as tmp:
        client, src_db = _client(tmp)
        try:
            _login(client, src_db)
            conn = db.get_connection(src_db)
            _seed_mixed(conn)
            conn.execute(
                "INSERT INTO personnel (asset_serial, person_name) VALUES ('DYN-abc','某人')"
            )
            conn.commit()
            conn.close()

            resp = client.get("/api/export")
            import openpyxl

            wb = openpyxl.load_workbook(io.BytesIO(resp.content))
            ws = wb["人員"]
            rows = list(ws.iter_rows(values_only=True))[1:]
            assert len(rows) == 0, "帳外資產底下的人員資料預設也不該匯出"
        finally:
            api.app.dependency_overrides.clear()
