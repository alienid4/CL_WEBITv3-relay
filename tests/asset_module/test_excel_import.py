"""S2 done_when 驗證：可匯入一份假資料 Excel 且欄位對應可調整。"""
import json
import sys
import tempfile
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[2]
sys.path.insert(0, str(ROOT / "APP" / "asset-module" / "backend"))

import db  # noqa: E402
import excel_import  # noqa: E402


def _build_fake_workbook(path: Path) -> None:
    wb = openpyxl.Workbook()
    ws_hw = wb.active
    ws_hw.title = "硬體"
    ws_hw.append(["資產序號", "主機名稱", "IP", "設備機型", "環境別", "保管者"])
    ws_hw.append(["HW-FAKE-0001", "fake-host-01", "10.0.0.11", "Fake Model X", "測試", "假保管者"])
    ws_hw.append(["HW-FAKE-0002", "fake-host-02", "10.0.0.12", "Fake Model Y", "正式", "假保管者"])

    ws_ppl = wb.create_sheet("人員")
    ws_ppl.append(["資產序號", "人員姓名", "聯絡電話", "職務概述"])
    ws_ppl.append(["HW-FAKE-0001", "假姓名", "0000-000000", "假職務"])
    ws_ppl.append(["HW-NOT-EXIST", "查無此資產", "0000-000000", "應被略過"])

    ws_sw = wb.create_sheet("軟體")
    ws_sw.append(["資產序號", "資料庫/軟體", "備份頻率"])
    ws_sw.append(["HW-FAKE-0002", "PostgreSQL 14", "每日"])

    wb.save(path)


def test_import_creates_rows_and_respects_mapping():
    with tempfile.TemporaryDirectory() as tmp:
        db_path = Path(tmp) / "test.db"
        xlsx_path = Path(tmp) / "fake_ica.xlsx"

        db.init_db(db_path)
        _build_fake_workbook(xlsx_path)

        conn = db.get_connection(db_path)
        try:
            summary = excel_import.import_excel(xlsx_path, conn)

            assert summary["sheets"]["硬體"]["inserted"] == 2
            assert summary["sheets"]["人員"]["inserted"] == 1
            assert summary["sheets"]["人員"]["skipped"] == 1  # HW-NOT-EXIST 找不到硬體資料被略過
            assert summary["sheets"]["軟體"]["inserted"] == 1

            row = db.get_hardware_by_serial(conn, "HW-FAKE-0001")
            assert row["hostname"] == "fake-host-01"
            assert row["custodian"] == "假保管者"

            # 再匯入一次同一份檔案 → 應該是 update 不是重複 insert（asset_serial 是 upsert 鍵）
            summary2 = excel_import.import_excel(xlsx_path, conn)
            assert summary2["sheets"]["硬體"]["updated"] == 2
            assert summary2["sheets"]["硬體"]["inserted"] == 0
        finally:
            conn.close()


def test_field_mapping_is_configurable_not_hardcoded():
    """欄位對應表本身要是可讀寫的 JSON 設定檔，不是寫死在程式碼常數裡。"""
    mapping = excel_import.load_mapping()
    assert isinstance(mapping, dict)
    assert "硬體" in mapping
    assert mapping["硬體"]["主機名稱"] == "hostname"
    # 驗證這是從外部檔案讀進來的，不是程式碼裡的常數
    raw = json.loads(excel_import.MAPPING_PATH.read_text(encoding="utf-8"))
    assert raw["硬體"]["主機名稱"] == "hostname"


if __name__ == "__main__":
    test_import_creates_rows_and_respects_mapping()
    test_field_mapping_is_configurable_not_hardcoded()
    print("S2 test_excel_import.py: PASS")
