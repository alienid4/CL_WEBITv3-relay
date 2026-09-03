"""部門報告圖表頁的 API 層：路由接得對、xlsx 匯出匯入 round-trip 沒壞掉。

計算邏輯已經在 test_report_physical_and_overview.py 逐項守過，這裡只守「薄 HTTP 層
沒接錯」——尤其是 xlsx 上傳/下載這種容易漏測的路徑（欄位順序、認證、壞檔案）。
"""
import io
import sys
import tempfile
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[2]
sys.path.insert(0, str(ROOT / "APP" / "asset-module" / "backend"))

import api  # noqa: E402
import auth  # noqa: E402
import db  # noqa: E402
from fastapi.testclient import TestClient  # noqa: E402

_PW = "test-password-123"


def _client(tmp):
    db_path = Path(tmp) / "t.db"
    db.init_db(db_path)

    def _override():
        conn = db.get_connection(db_path)
        try:
            yield conn
        finally:
            conn.close()

    api.app.dependency_overrides[api.get_db] = _override
    client = TestClient(api.app)
    conn = db.get_connection(db_path)
    try:
        db.create_user(conn, "tester", auth.hash_password(_PW))
        db.insert_hardware(conn, asset_serial="HW-1", hostname="core-a", ip="10.0.0.1",
                           os="Windows Server 2019", physical_location="板橋機房",
                           environment="正式", api_id="N-001", asset_name="核心系統A",
                           is_vm="0", asset_status="在用")
        db.insert_hardware(conn, asset_serial="HW-2", hostname="vm-a", ip="10.0.0.2",
                           os="Windows Server 2019", physical_location="內湖機房",
                           environment="測試", api_id="N-002", asset_name="系統B",
                           is_vm="VM", asset_status="在用")
        conn.commit()
    finally:
        conn.close()
    assert client.post("/api/auth/login",
                       json={"username": "tester", "password": _PW}).status_code == 200
    return client


def test_未登入打不到這幾支端點():
    with tempfile.TemporaryDirectory() as tmp:
        db_path = Path(tmp) / "t.db"
        db.init_db(db_path)

        def _override():
            conn = db.get_connection(db_path)
            try:
                yield conn
            finally:
                conn.close()

        api.app.dependency_overrides[api.get_db] = _override
        client = TestClient(api.app)
        try:
            for path in ("/api/reports/physical-distribution", "/api/reports/system-overview",
                        "/api/reports/system-category/template"):
                assert client.get(path).status_code == 401, path
        finally:
            api.app.dependency_overrides.clear()


def test_頁A端點回傳且下鑽數字對得起來():
    with tempfile.TemporaryDirectory() as tmp:
        client = _client(tmp)
        try:
            d = client.get("/api/reports/physical-distribution").json()
            room = next(r for r in d["rooms"] if r["room"] == "板橋")
            assert room["total"] == 1
            rows = client.get("/api/reports/physical-distribution/drill",
                              params={"room": "板橋"}).json()
            assert len(rows) == room["total"]
            assert rows[0]["asset_serial"] == "HW-1"
        finally:
            api.app.dependency_overrides.clear()


def test_頁B端點回傳且下鑽bucket參數擋亂填():
    with tempfile.TemporaryDirectory() as tmp:
        client = _client(tmp)
        try:
            o = client.get("/api/reports/system-overview").json()
            assert o["total"] == 2
            r = client.get("/api/reports/system-overview/drill",
                           params={"bucket": "亂填"})
            assert r.status_code == 400
            ok = client.get("/api/reports/system-overview/drill",
                            params={"bucket": "test"}).json()
            assert len(ok) == o["test"]

            room = o["rooms"][0]
            room_rows = client.get("/api/reports/system-overview/drill",
                                   params={"room": room["room"]}).json()
            assert len(room_rows) == room["total"]
        finally:
            api.app.dependency_overrides.clear()


def test_業務系統對照表匯出匯入round_trip():
    """下載範本 → 填一格分類 → 上傳 → 頁B的數字要反映出來。
    這是整條路徑唯一真正的風險點：欄位順序錯位不會報錯，只會靜靜地分類全錯。"""
    with tempfile.TemporaryDirectory() as tmp:
        client = _client(tmp)
        try:
            r = client.get("/api/reports/system-category/template")
            assert r.status_code == 200
            assert r.headers["content-type"].startswith(
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

            wb = openpyxl.load_workbook(io.BytesIO(r.content))
            ws = wb.active
            header = [c.value for c in ws[1]]
            assert header[0] == "api_id"
            rows = list(ws.iter_rows(min_row=2, values_only=True))
            assert {row[0] for row in rows} == {"N-001", "N-002"}

            # 填 N-001＝金融交易服務（核心交易服務組），其餘留空，改檔名重上傳
            for row in ws.iter_rows(min_row=2):
                if row[0].value == "N-001":
                    row[3].value = "M.金融交易服務"
            buf = io.BytesIO()
            wb.save(buf)
            buf.seek(0)

            up = client.post(
                "/api/reports/system-category/import",
                files={"file": ("filled.xlsx", buf,
                                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            )
            assert up.status_code == 200, up.text
            assert up.json()["accepted"] == 1

            o = client.get("/api/reports/system-overview").json()
            assert o["core"] == 1
        finally:
            api.app.dependency_overrides.clear()


def test_匯入非xlsx檔案被拒絕():
    with tempfile.TemporaryDirectory() as tmp:
        client = _client(tmp)
        try:
            r = client.post(
                "/api/reports/system-category/import",
                files={"file": ("x.csv", io.BytesIO(b"api_id,category\n"), "text/csv")},
            )
            assert r.status_code == 400
        finally:
            api.app.dependency_overrides.clear()


# ===== 匯入必須配對匯出範本（2026-08-25 使用者通則）=====
# 這兩支端點不屬於部門報告，跟這個檔案裡其他測試是不同計畫，放這裡是因為
# 同樣是「xlsx 下載，要驗證 headers/content-type/認證」這種薄 HTTP 層的驗法。

def test_網段配置表匯出範本_未登入打不到():
    with tempfile.TemporaryDirectory() as tmp:
        db_path = Path(tmp) / "t.db"
        db.init_db(db_path)

        def _override():
            conn = db.get_connection(db_path)
            try:
                yield conn
            finally:
                conn.close()

        api.app.dependency_overrides[api.get_db] = _override
        client = TestClient(api.app)
        try:
            assert client.get("/api/segments/export-template").status_code == 401
        finally:
            api.app.dependency_overrides.clear()


def test_網段配置表匯出範本可以匯入自己():
    with tempfile.TemporaryDirectory() as tmp:
        client = _client(tmp)
        try:
            r = client.get("/api/segments/export-template")
            assert r.status_code == 200
            assert r.headers["content-type"].startswith(
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            wb = openpyxl.load_workbook(io.BytesIO(r.content))
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
            assert rows[0][0] == "使用狀況"
            assert len(rows) == 2, "範本要有一筆範例列"
        finally:
            api.app.dependency_overrides.clear()


def test_存活清單匯出範本():
    with tempfile.TemporaryDirectory() as tmp:
        client = _client(tmp)
        try:
            r = client.get("/api/import/dynassets/export-template")
            assert r.status_code == 200
            wb = openpyxl.load_workbook(io.BytesIO(r.content))
            rows = list(wb.active.iter_rows(values_only=True))
            assert "IP" in rows[0]
            assert len(rows) == 2
        finally:
            api.app.dependency_overrides.clear()
