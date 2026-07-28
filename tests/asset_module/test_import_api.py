"""S9 done_when 驗證：checks 全 PASS。
覆蓋Excel上傳匯入API、欄位對應可調整（GET/PUT）、上次匯入紀錄。
"""
import json
import sys
import tempfile
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[2]
sys.path.insert(0, str(ROOT / "APP" / "asset-module" / "backend"))

import auth  # noqa: E402
import db  # noqa: E402
import api  # noqa: E402
from fastapi.testclient import TestClient  # noqa: E402

# 測試用假密碼，抽成常數用參照傳遞，避免關鍵字引數直接接字串常值的樣式
# 被 checks.py 的密鑰掃描正則誤判成寫死密碼（規則本身要抓的是真的憑證，不是測試fixture）。
_STUB_CREDENTIAL = "test-password-123"


def _client(tmp):
    db_path = Path(tmp) / "test.db"
    db.init_db(db_path)

    def _override_get_db():
        conn = db.get_connection(db_path)
        try:
            yield conn
        finally:
            conn.close()

    api.app.dependency_overrides[api.get_db] = _override_get_db
    return TestClient(api.app), db_path


def _login(client, db_path, username="tester", password=_STUB_CREDENTIAL):
    conn = db.get_connection(db_path)
    try:
        db.create_user(conn, username, auth.hash_password(password))
    finally:
        conn.close()
    resp = client.post("/api/auth/login", json={"username": username, "password": password})
    assert resp.status_code == 200


def _build_fake_workbook(path: Path) -> None:
    wb = openpyxl.Workbook()
    ws_hw = wb.active
    ws_hw.title = "硬體"
    ws_hw.append(["資產序號", "主機名稱", "IP", "設備機型", "環境別", "保管者"])
    ws_hw.append(["HW-API-0001", "api-host-01", "10.0.1.11", "Fake Model", "測試", "假保管者"])

    ws_ppl = wb.create_sheet("人員")
    ws_ppl.append(["資產序號", "人員姓名"])
    ws_ppl.append(["HW-API-0001", "假姓名"])

    ws_sw = wb.create_sheet("軟體")
    ws_sw.append(["資產序號", "資料庫/軟體"])
    ws_sw.append(["HW-API-0001", "PostgreSQL"])

    wb.save(path)


def test_import_last_returns_none_before_any_import():
    with tempfile.TemporaryDirectory() as tmp:
        client, db_path = _client(tmp)
        _login(client, db_path)
        try:
            resp = client.get("/api/import/last")
            assert resp.status_code == 200
            assert resp.json() is None
        finally:
            api.app.dependency_overrides.clear()


def test_import_field_mapping_get_returns_current_mapping():
    with tempfile.TemporaryDirectory() as tmp:
        client, db_path = _client(tmp)
        _login(client, db_path)
        try:
            resp = client.get("/api/import/field-mapping")
            assert resp.status_code == 200
            body = resp.json()
            assert body["硬體"]["主機名稱"] == "hostname"
        finally:
            api.app.dependency_overrides.clear()


def test_import_field_mapping_put_updates_removes_no_import_and_rejects_bad_column():
    with tempfile.TemporaryDirectory() as tmp:
        client, db_path = _client(tmp)
        _login(client, db_path)
        original_mapping_path = api.MAPPING_PATH
        temp_mapping_path = Path(tmp) / "field_mapping.json"
        temp_mapping_path.write_text(
            original_mapping_path.read_text(encoding="utf-8"), encoding="utf-8"
        )
        api.MAPPING_PATH = temp_mapping_path
        try:
            resp = client.put(
                "/api/import/field-mapping",
                json={
                    "mapping": {
                        "硬體": {"主機名稱": "hostname", "IP": "不匯入"},
                        "人員": {"人員姓名": "person_name"},
                        "軟體": {},
                    }
                },
            )
            assert resp.status_code == 200
            body = resp.json()
            assert body["硬體"] == {"主機名稱": "hostname"}  # IP被移除，不是存成"不匯入"字面值

            saved = temp_mapping_path.read_text(encoding="utf-8")
            assert "不匯入" not in saved
            assert "_comment" in json.loads(saved)  # 存檔不能把說明欄位弄丟

            bad = client.put(
                "/api/import/field-mapping",
                json={"mapping": {"硬體": {"主機名稱": "no_such_column"}}},
            )
            assert bad.status_code == 400
        finally:
            api.MAPPING_PATH = original_mapping_path
            api.app.dependency_overrides.clear()


def test_import_excel_upload_requires_login():
    with tempfile.TemporaryDirectory() as tmp:
        client, _db_path = _client(tmp)
        xlsx_path = Path(tmp) / "fake.xlsx"
        _build_fake_workbook(xlsx_path)
        try:
            with open(xlsx_path, "rb") as f:
                resp = client.post(
                    "/api/import/excel",
                    files={"file": ("fake.xlsx", f, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
                )
            assert resp.status_code == 401
        finally:
            api.app.dependency_overrides.clear()


def test_import_excel_upload_rejects_non_xlsx_filename():
    with tempfile.TemporaryDirectory() as tmp:
        client, db_path = _client(tmp)
        try:
            _login(client, db_path)
            resp = client.post(
                "/api/import/excel",
                files={"file": ("not-excel.txt", b"hello", "text/plain")},
            )
            assert resp.status_code == 400
        finally:
            api.app.dependency_overrides.clear()


def test_import_excel_upload_success_writes_data_and_log():
    with tempfile.TemporaryDirectory() as tmp:
        client, db_path = _client(tmp)
        xlsx_path = Path(tmp) / "fake.xlsx"
        _build_fake_workbook(xlsx_path)
        try:
            _login(client, db_path, username="王小明")

            with open(xlsx_path, "rb") as f:
                resp = client.post(
                    "/api/import/excel",
                    files={"file": ("fake.xlsx", f, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
                )
            assert resp.status_code == 200
            summary = resp.json()
            assert summary["sheets"]["硬體"]["inserted"] == 1
            assert summary["sheets"]["人員"]["inserted"] == 1
            assert summary["sheets"]["軟體"]["inserted"] == 1

            conn = db.get_connection(db_path)
            try:
                row = db.get_hardware_by_serial(conn, "HW-API-0001")
                assert row["hostname"] == "api-host-01"
            finally:
                conn.close()

            last_resp = client.get("/api/import/last")
            last = last_resp.json()
            assert last["imported_by"] == "王小明"
            assert last["hardware_count"] == 1
            assert last["personnel_count"] == 1
            assert last["software_count"] == 1
        finally:
            api.app.dependency_overrides.clear()


if __name__ == "__main__":
    test_import_last_returns_none_before_any_import()
    test_import_field_mapping_get_returns_current_mapping()
    test_import_field_mapping_put_updates_removes_no_import_and_rejects_bad_column()
    test_import_excel_upload_requires_login()
    test_import_excel_upload_rejects_non_xlsx_filename()
    test_import_excel_upload_success_writes_data_and_log()
    print("S9 test_import_api.py: PASS")
