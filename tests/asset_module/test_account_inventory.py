"""帳號盤點寫入層與 API：只收已納管、消失標記、finding 重算、例外到期回燈、端點需登入。"""
import sys
import tempfile
from pathlib import Path

ROOT = Path(__file__).resolve().parents[2]
sys.path.insert(0, str(ROOT / "APP" / "asset-module" / "backend"))

import account_inventory  # noqa: E402
import api  # noqa: E402
import auth  # noqa: E402
import db  # noqa: E402
from fastapi.testclient import TestClient  # noqa: E402

_PW = "test-password-123"

PASSWD = """root:x:0:0:root:/root:/bin/bash
alice:x:1000:1000:Alice:/home/alice:/bin/bash
bob:x:1001:1001:Bob:/home/bob:/bin/bash
"""
PASSWD_BOB_GONE = """root:x:0:0:root:/root:/bin/bash
alice:x:1000:1000:Alice:/home/alice:/bin/bash
"""
GROUP = "wheel:x:10:alice\n"
LASTLOG = "Username Port From Latest\nalice pts/0 192.0.2.9 Jul 15, 2026\nbob   **Never logged in**\n"


def _runner(passwd=PASSWD):
    def run(host, cmd):
        if cmd.startswith("cat /etc/passwd"):
            return passwd
        if cmd.startswith("cat /etc/group"):
            return GROUP
        if "lastlog" in cmd:          # 指令帶 LC_ALL=C 前綴，不能用 startswith 比對
            return LASTLOG
        return ""          # 其餘需 root，模擬拿不到
    return run


def _db(tmp):
    p = Path(tmp) / "t.db"
    db.init_db(p)
    return db.get_connection(p)


def _seed(conn, ip="203.0.113.20", serial="A-1"):
    db.insert_hardware(conn, asset_serial=serial, hostname="app1", ip=ip, collect_ok=1)
    conn.execute("INSERT INTO scan_history (ip, scan_ok, open_ports, scan_time) "
                 "VALUES (?,1,'22','2026-07-21 10:00:00')", (ip,))
    conn.commit()


def test_收集寫入帳號並產生稽核發現():
    with tempfile.TemporaryDirectory() as tmp:
        conn = _db(tmp)
        try:
            _seed(conn)
            r = account_inventory.collect_accounts(conn, runner=_runner())
            assert r["status"] == "ok"
            assert r["accounts"] == 3
            assert r["findings"] > 0
            assert r["needs_root_hosts"] == 1        # 沒有 root，要誠實記錄
            names = {a["username"] for a in account_inventory.list_accounts(conn)}
            assert names == {"root", "alice", "bob"}
        finally:
            conn.close()


def test_帳號消失是標記不是刪除():
    """離職清掉了（好事）跟被人偷偷刪掉湮滅痕跡（大事）都要留得下來。"""
    with tempfile.TemporaryDirectory() as tmp:
        conn = _db(tmp)
        try:
            _seed(conn)
            account_inventory.collect_accounts(conn, runner=_runner())
            r2 = account_inventory.collect_accounts(conn, runner=_runner(PASSWD_BOB_GONE))
            assert r2["hosts"][0]["gone"] == 1
            live = {a["username"] for a in account_inventory.list_accounts(conn)}
            assert "bob" not in live
            allrows = account_inventory.list_accounts(conn, include_gone=True)
            assert any(a["username"] == "bob" and a["gone_at"] for a in allrows)
        finally:
            conn.close()


def test_findings每次重算不留舊紅燈():
    """已經修好卻還亮著的舊紅燈比沒有紅燈更糟——沒人會信一個會謊報的稽核系統。"""
    with tempfile.TemporaryDirectory() as tmp:
        conn = _db(tmp)
        try:
            _seed(conn)
            account_inventory.collect_accounts(conn, runner=_runner())
            first = len(account_inventory.latest_findings(conn))
            # bob 移除後，跟 bob 有關的 finding 不該再出現在「現況」
            account_inventory.collect_accounts(conn, runner=_runner(PASSWD_BOB_GONE))
            latest = account_inventory.latest_findings(conn)
            assert not any(f["username"] == "bob" for f in latest)
            assert first > len(latest)
            # 但舊 run 的紀錄仍在（要對照「上次稽核 vs 這次」）
            total = conn.execute("SELECT COUNT(*) FROM account_finding").fetchone()[0]
            assert total > len(latest)
        finally:
            conn.close()


def test_只收已納管主機():
    with tempfile.TemporaryDirectory() as tmp:
        conn = _db(tmp)
        try:
            _seed(conn, "203.0.113.20", "A-1")
            db.insert_hardware(conn, asset_serial="A-2", hostname="x", ip="203.0.113.21",
                               collect_ok=0)
            r = account_inventory.collect_accounts(conn, runner=_runner())
            assert r["candidates"] == 1
        finally:
            conn.close()


def test_例外到期後要自動回來亮燈():
    """永久例外等於把問題永遠藏起來，所以例外一定要有期限。
    狀態已改走持久的 finding_disposition（見 test_finding_lifecycle），這裡用新機制驗。"""
    with tempfile.TemporaryDirectory() as tmp:
        conn = _db(tmp)
        try:
            _seed(conn)
            account_inventory.collect_accounts(conn, runner=_runner())
            rows = account_inventory.latest_findings(conn)
            assert rows
            f = rows[0]
            key = (f["ip"], f["username"], f["rule_id"])

            def present():
                return any((x["ip"], x["username"], x["rule_id"]) == key
                           for x in account_inventory.latest_findings(conn))

            # 核准例外到未來 → 不列
            account_inventory.set_finding_disposition(
                conn, *key, "exception", exempt_until="2099-12-31", decided_by="tester")
            assert not present()

            # 過期 → 自動回來
            account_inventory.set_finding_disposition(
                conn, *key, "exception", exempt_until="2000-01-01", decided_by="tester")
            assert present()
        finally:
            conn.close()


def test_摘要要顯示收集失敗數():
    """4 台裡 3 台連不上時，failed_count 必須浮到摘要——不能被 unknown=0 蓋掉。
    （真機踩過：切 sysinfra 後遠端沒授權金鑰全部失敗，畫面卻一片綠。）"""
    with tempfile.TemporaryDirectory() as tmp:
        conn = _db(tmp)
        try:
            _seed(conn, ip="203.0.113.20", serial="A-1")
            _seed(conn, ip="203.0.113.21", serial="A-2")

            def flaky(host, cmd):
                if host == "203.0.113.21":
                    raise OSError("Permission denied (publickey)")
                if cmd.startswith("cat /etc/passwd"):
                    return PASSWD
                if "lastlog" in cmd:
                    return LASTLOG
                return ""

            account_inventory.collect_accounts(conn, runner=flaky)
            s = account_inventory.audit_summary(conn)
            assert s["failed_count"] == 1
            assert s["host_count"] == 2
            assert s["run_error"] and "Permission denied" in s["run_error"]
        finally:
            conn.close()


def test_摘要要把查不到單獨算():
    """『查不到』代表這份稽核報告有多少是空白的，藏起來會讓人以為全查過了。"""
    with tempfile.TemporaryDirectory() as tmp:
        conn = _db(tmp)
        try:
            _seed(conn)
            account_inventory.collect_accounts(conn, runner=_runner())
            s = account_inventory.audit_summary(conn)
            assert s["has_data"] is True
            assert s["unknown"] > 0
            assert s["hosts_needing_root"] == 1
            assert s["privileged"] >= 1          # root + wheel 成員
        finally:
            conn.close()


# ---- API ----

def _client(tmp):
    p = Path(tmp) / "t.db"
    db.init_db(p)

    def _ov():
        c = db.get_connection(p)
        try:
            yield c
        finally:
            c.close()

    api.app.dependency_overrides[api.get_db] = _ov
    return TestClient(api.app), p


def _login(client, p):
    c = db.get_connection(p)
    try:
        db.create_user(c, "tester", auth.hash_password(_PW))
    finally:
        c.close()
    assert client.post("/api/auth/login",
                       json={"username": "tester", "password": _PW}).status_code == 200


def test_帳號端點都要登入():
    with tempfile.TemporaryDirectory() as tmp:
        client, _ = _client(tmp)
        try:
            assert client.get("/api/accounts").status_code == 401
            assert client.get("/api/accounts/findings").status_code == 401
            assert client.post("/api/accounts/collect").status_code == 401
            assert client.get("/api/accounts/sudo-rules").status_code == 401
        finally:
            api.app.dependency_overrides.clear()


def test_findings端點帶出法規出處與門檻():
    """稽核當天要的是可交付的證據：規則、法規依據、目前門檻。"""
    with tempfile.TemporaryDirectory() as tmp:
        client, p = _client(tmp)
        try:
            _login(client, p)
            c = db.get_connection(p)
            try:
                _seed(c)
                account_inventory.collect_accounts(c, runner=_runner())
            finally:
                c.close()
            r = client.get("/api/accounts/findings").json()
            assert r["items"]
            assert all("law" in x for x in r["items"])
            assert r["thresholds"]["acct_pw_max_days"] == 90
            assert any(rule["id"] == "A2" for rule in r["rules"])
        finally:
            api.app.dependency_overrides.clear()


def test_合規表狀態欄計算與前端對齊():
    """合規表匯出的狀態字必須跟畫面一致；尤其 LK 要算成『已停用』
    （不然又踩回『鎖定帳號被當成啟用中』那個真機 bug）。"""
    disabled = account_inventory._cell_disabled
    assert disabled({"pw_status": "LK"}) == "已停用"      # 兩字母碼也要收
    assert disabled({"pw_status": "locked"}) == "已停用"
    assert disabled({"pw_status": "set"}) == "啟用中"
    assert disabled({"pw_status": None}) == "需 root"
    assert account_inventory._cell_uid0(
        {"uid": 0, "username": "backdoor"}) == "是"
    assert account_inventory._cell_uid0({"uid": 0, "username": "root"}) == "root"
    assert account_inventory._cell_keys({"authorized_keys": None}) == "需 root"
    assert account_inventory._cell_keys({"authorized_keys": 3}) == "3 把"


def test_合規表匯出只吐勾選欄位():
    """給稽核資料最小化：只勾的欄位才進 Excel，亂填欄位被擋。"""
    with tempfile.TemporaryDirectory() as tmp:
        client, p = _client(tmp)
        try:
            _login(client, p)
            c = db.get_connection(p)
            try:
                _seed(c)
                account_inventory.collect_accounts(c, runner=_runner())
            finally:
                c.close()
            # 沒選有效欄位 → 400
            assert client.get("/api/accounts/matrix/export",
                              params={"cols": ["nope"]}).status_code == 400
            r = client.get("/api/accounts/matrix/export",
                           params={"cols": ["username", "disabled", "sudo"]})
            assert r.status_code == 200
            assert "spreadsheetml" in r.headers["content-type"]
            import io as _io
            from openpyxl import load_workbook
            ws = load_workbook(_io.BytesIO(r.content)).active
            assert [c.value for c in ws[1]] == ["帳號", "帳號停用", "sudo 權限"]
        finally:
            api.app.dependency_overrides.clear()


def test_sudo規則端點不含shadow():
    with tempfile.TemporaryDirectory() as tmp:
        client, p = _client(tmp)
        try:
            _login(client, p)
            r = client.get("/api/accounts/sudo-rules").json()
            lines = [l for l in r["rules"].splitlines()
                     if l.strip() and not l.strip().startswith("#")]
            assert not any("/etc/shadow" in l for l in lines)
        finally:
            api.app.dependency_overrides.clear()


def test_備註欄位gecos有帶出且可排序():
    """帳號備註＝/etc/passwd 第 5 欄（GECOS），例如人名或帳號用途說明。"""
    with tempfile.TemporaryDirectory() as tmp:
        client, p = _client(tmp)
        try:
            _login(client, p)
            c = db.get_connection(p)
            try:
                _seed(c)
                account_inventory.collect_accounts(c, runner=_runner())
            finally:
                c.close()
            rows = client.get("/api/accounts").json()["items"]
            alice = [r for r in rows if r["username"] == "alice"][0]
            assert alice["gecos"] == "Alice"
            # 可依備註排序（走白名單，不炸）
            assert client.get("/api/accounts", params={"sort_by": "gecos"}).status_code == 200
        finally:
            api.app.dependency_overrides.clear()


def test_特權帳號篩選():
    with tempfile.TemporaryDirectory() as tmp:
        client, p = _client(tmp)
        try:
            _login(client, p)
            c = db.get_connection(p)
            try:
                _seed(c)
                account_inventory.collect_accounts(c, runner=_runner())
            finally:
                c.close()
            r = client.get("/api/accounts", params={"sudoer_only": True}).json()
            names = {x["username"] for x in r["items"]}
            assert "root" in names          # UID 0
            assert "alice" in names         # wheel 群組
            assert "bob" not in names
        finally:
            api.app.dependency_overrides.clear()
