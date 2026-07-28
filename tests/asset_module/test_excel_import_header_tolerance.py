"""匯入器表頭容錯：舊版/手改過表頭的檔案也要匯得進來。

真實故障（2026-07-18 使用者回報）：拿「表頭修正前」匯出的 assets_export.xlsx 來匯入，
整批失敗，錯誤是 `硬體：寫入失敗 (NOT NULL constraint failed: hardware.asset_serial)`。
根因不是資料壞，是表頭「資產序號（公司資產編號）」對不上匯入器認的「資產序號」，
欄位沒配對到 -> asset_serial 是空的 -> 撞 NOT NULL -> 每一列都掛。

修法是比對前先用 excel_import._norm() 正規化（去括號註解/全形斜線/空白/大小寫），
真實字差（AP ID vs API ID）則靠 field_mapping.json 的別名。
"""
import sys
import tempfile
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[2]
sys.path.insert(0, str(ROOT / "APP" / "asset-module" / "backend"))

import db  # noqa: E402
import excel_import  # noqa: E402

# 這些就是舊版匯出檔實際會出現的表頭（來自 field_meta 的 label，與 field_mapping 的鍵不一致）
OLD_STYLE_HEADERS = [
    "資產序號（公司資產編號）",  # 括號後綴——當初讓整批匯入失敗的元凶
    "主機名稱",
    "IP",
    "BIG IP／VIP",              # 全形斜線
    "保管者（SP）",              # 括號後綴
    "使用者（AP User）",         # 括號後綴（含英數字）
    "AP ID",                    # 與 mapping 的 "API ID" 是真實字差，靠別名
    "機密性 (C)",               # 括號前多一個半形空格
    "完整性 (I)",
    "可用性 (A)",
]


def _build_old_style_workbook(path: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "硬體"
    ws.append(OLD_STYLE_HEADERS)
    ws.append(
        ["HW-OLD-0001", "old-host-01", "10.0.0.21", "vip-01", "假保管者", "假使用者",
         "AP-999", "高", "中", "低"]
    )
    wb.create_sheet("人員").append(["資產序號", "人員姓名"])
    wb.create_sheet("軟體").append(["資產序號", "資料庫/軟體"])
    wb.save(path)


def test_舊版表頭也能匯入且資產序號有值():
    with tempfile.TemporaryDirectory() as tmp:
        db_path = Path(tmp) / "test.db"
        xlsx_path = Path(tmp) / "old_style.xlsx"
        db.init_db(db_path)
        _build_old_style_workbook(xlsx_path)

        conn = db.get_connection(db_path)
        try:
            summary = excel_import.import_excel(xlsx_path, conn)
            rows = conn.execute("SELECT * FROM hardware").fetchall()
        finally:
            conn.close()

        # 當初的症狀就是這裡整批 skipped、errors 有 NOT NULL constraint
        assert summary["sheets"]["硬體"]["inserted"] == 1, f"沒匯進去：{summary}"
        assert not [e for e in summary["errors"] if "NOT NULL" in e], summary["errors"]

        assert len(rows) == 1
        row = rows[0]
        assert row["asset_serial"] == "HW-OLD-0001"   # 最關鍵：不能是空的
        assert row["hostname"] == "old-host-01"
        assert row["big_ip_vip"] == "vip-01"          # 全形斜線有被橋接
        assert row["custodian"] == "假保管者"          # 括號後綴有被剝掉
        assert row["user_name"] == "假使用者"
        assert row["api_id"] == "AP-999"              # 別名有生效
        assert row["confidentiality"] == "高"          # 括號前空格有被吸收
        assert row["integrity"] == "中"
        assert row["availability"] == "低"


def test_正規化不會把不同欄位混在一起():
    """正規化太寬鬆會誤配欄位，比匯不進去更危險（資料會靜靜寫錯格）。
    確認設定檔裡沒有任何兩個標題正規化後撞在一起。"""
    mapping = excel_import.load_mapping()
    for sheet, columns in mapping.items():
        seen: dict[str, str] = {}
        for header, field in columns.items():
            norm = excel_import._norm(header)
            if norm in seen:
                assert seen[norm] == field, (
                    f"{sheet}分頁「{header}」與另一個標題正規化後都變成「{norm}」，"
                    f"卻對到不同欄位（{seen[norm]} vs {field}）——會誤配，請改標題或收緊 _norm"
                )
            seen[norm] = field


def test_別名不放進使用者可編輯的對應表():
    """別名是相容性補丁，放進 field_mapping.json 會在「欄位對應」畫面多出看似重複的列，
    使用者可能把它清成「不匯入」讓舊檔又匯不進來，行序也會影響匯出表頭。"""
    mapping = excel_import.load_mapping()
    for sheet, columns in mapping.items():
        for alias in excel_import.HEADER_ALIASES:
            assert alias not in columns, (
                f"別名「{alias}」跑進了 {sheet} 分頁的 field_mapping.json，"
                "應該只留在 excel_import.HEADER_ALIASES"
            )
        # 一個資料庫欄位在設定檔裡只該有一個正式標題（匯出表頭才不會隨行序漂移）
        fields = list(columns.values())
        dupes = {f for f in fields if fields.count(f) > 1}
        assert not dupes, f"{sheet} 分頁有欄位對到多個標題：{dupes}"


def test_別名指向的正式標題必須存在():
    """別名寫錯字（指向一個對應表裡沒有的標題）會靜靜失效，這裡直接擋下來。"""
    mapping = excel_import.load_mapping()
    all_norm_headers = {
        excel_import._norm(h) for cols in mapping.values() for h in cols
    }
    for alias, canonical in excel_import.HEADER_ALIASES.items():
        assert excel_import._norm(canonical) in all_norm_headers, (
            f"別名「{alias}」指向「{canonical}」，但對應表裡找不到這個標題"
        )
        assert excel_import._norm(alias) not in all_norm_headers, (
            f"別名「{alias}」正規化後其實就等於既有標題，不需要列為別名"
        )


def test_norm_行為():
    n = excel_import._norm
    assert n("資產序號（公司資產編號）") == n("資產序號")
    assert n("BIG IP／VIP") == n("BIG IP/VIP")
    assert n("機密性 (C)") == n("機密性(C)")
    assert n("  主機名稱  ") == n("主機名稱")
    assert n("Cloud") == n("CLOUD")
    # 不同欄位不能被壓成一樣
    assert n("完整性(I)") != n("機密性(C)")
    assert n("盤點單位-處別") != n("盤點單位-部門")
    # 非字串（空白表頭儲存格）不能爆
    assert n(None) == ""
    assert n(12345) == ""
