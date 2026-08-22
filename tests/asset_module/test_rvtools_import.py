"""S19 VC 採集器：RVTools 匯入。

核心要守的是「不亂合併、不覆蓋人填的欄位、判不準交人工」——這幾條錯了會安靜地把
兩台機器變一台或蓋掉人維護的資料，很難發現。用合成的 RVTools 格式檔把整條路走一遍。

⚠️ 這裡的樣本全是合成假資料（openpyxl 當場造），不是真實 vCenter 匯出——真檔進來對格式即可。
"""
import sys
import tempfile
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[2]
sys.path.insert(0, str(ROOT / "APP" / "asset-module" / "backend"))

import db  # noqa: E402
import rvtools_import as rv  # noqa: E402

# RVTools vInfo 標準表頭（挑我們會用到的那些）
HEADERS = ["VM", "DNS Name", "Powerstate", "IP Address", "Primary IP Address",
           "OS according to the VMware Tools", "Host", "VM UUID", "VM ID",
           "Cluster", "Path"]


def _make_xlsx(tmp: str, rows: list[dict], sheet="vInfo") -> Path:
    """造一份 RVTools 格式的 xlsx。rows 用 HEADERS 的欄名當 key。"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet
    ws.append(HEADERS)
    for r in rows:
        ws.append([r.get(h, "") for h in HEADERS])
    p = Path(tmp) / "rvtools.xlsx"
    wb.save(p)
    return p


def _conn(tmp):
    p = Path(tmp) / "t.db"
    db.init_db(p)
    return db.get_connection(p)


VM_A = {
    "VM": "web01", "DNS Name": "web01.corp.local", "Powerstate": "poweredOn",
    "Primary IP Address": "10.1.1.10", "OS according to the VMware Tools": "Ubuntu Linux (64-bit)",
    "Host": "esxi-a.corp.local", "VM UUID": "42301111-2222-3333-4444-555566667777", "VM ID": "vm-101",
}
VM_B = {
    "VM": "db01", "DNS Name": "db01.corp.local", "Powerstate": "poweredOff",
    "IP Address": "10.1.1.20", "OS according to the VMware Tools": "Microsoft Windows Server 2019",
    "Host": "esxi-b.corp.local", "VM UUID": "4230aaaa-bbbb-cccc-dddd-eeeeffff0000", "VM ID": "vm-102",
}


# ===== 解析 =====

def test_解析標準欄位_取對來源():
    with tempfile.TemporaryDirectory() as tmp:
        recs = rv.parse_rvtools(_make_xlsx(tmp, [VM_A]))
        assert len(recs) == 1
        r = recs[0]
        assert r["vm_uuid"] == "42301111-2222-3333-4444-555566667777"
        assert r["hostname"] == "web01.corp.local"   # DNS Name 優先於 VM 顯示名
        assert r["ip"] == "10.1.1.10"                 # Primary IP 優先
        assert r["os"] == "Ubuntu Linux (64-bit)"     # VMware Tools 回報的
        assert r["is_vm"] == 1
        assert r["esxi_host"] == "esxi-a.corp.local"


def test_IP退回IP_Address_當沒有Primary():
    with tempfile.TemporaryDirectory() as tmp:
        recs = rv.parse_rvtools(_make_xlsx(tmp, [VM_B]))
        assert recs[0]["ip"] == "10.1.1.20"           # 只有 IP Address 欄


def test_解析cluster與datastore_MICS切片1用():
    """cluster 直接對欄位；datastore 沒有獨立欄位，要從 Path 欄的中括號拆出來。"""
    vm = {**VM_A, "Cluster": "BQ_PROD_B_vSan_Cluster",
          "Path": "[PROD_B_vSan_Datastore] 8e9fae68-a832/web01.vmx"}
    with tempfile.TemporaryDirectory() as tmp:
        recs = rv.parse_rvtools(_make_xlsx(tmp, [vm]))
        assert recs[0]["cluster"] == "BQ_PROD_B_vSan_Cluster"
        assert recs[0]["datastore"] == "PROD_B_vSan_Datastore"


def test_解析datastore_Path缺失時不報錯():
    with tempfile.TemporaryDirectory() as tmp:
        recs = rv.parse_rvtools(_make_xlsx(tmp, [VM_A]))  # 沒填 Cluster/Path
        assert recs[0]["cluster"] is None
        assert recs[0]["datastore"] is None


def test_全空列與無名列被略過():
    with tempfile.TemporaryDirectory() as tmp:
        rows = [VM_A, {h: "" for h in HEADERS}, {"Powerstate": "poweredOn"}]
        recs = rv.parse_rvtools(_make_xlsx(tmp, rows))
        assert len(recs) == 1


# ===== 匯入：新 VM 建成資產 =====

def test_新VM建成資產_is_vm且VC序號():
    with tempfile.TemporaryDirectory() as tmp:
        conn = _conn(tmp)
        try:
            summary = rv.import_rvtools(_make_xlsx(tmp, [VM_A, VM_B]), conn)
            assert summary["total_vms"] == 2
            assert summary["inserted"] == 2 and summary["updated"] == 0
            rows = conn.execute(
                "SELECT asset_serial, hostname, ip, os, is_vm, vm_uuid FROM hardware ORDER BY hostname"
            ).fetchall()
            assert len(rows) == 2
            for row in rows:
                assert row["is_vm"] == 1
                assert row["asset_serial"].startswith("VC-")
                assert row["vm_uuid"]
            # 序號用 vm_uuid，穩定可回溯
            assert conn.execute(
                "SELECT 1 FROM hardware WHERE asset_serial = ?",
                (f"VC-{VM_A['VM UUID']}",)).fetchone() is not None
        finally:
            conn.close()


def test_每列留vcenter來源紀錄():
    with tempfile.TemporaryDirectory() as tmp:
        conn = _conn(tmp)
        try:
            rv.import_rvtools(_make_xlsx(tmp, [VM_A, VM_B]), conn)
            n = conn.execute(
                "SELECT COUNT(*) c FROM source_record WHERE source='vcenter'").fetchone()["c"]
            assert n == 2
        finally:
            conn.close()


# ===== 重匯不重複（vm_uuid 強配）=====

def test_重複匯入同一份_更新不新增():
    with tempfile.TemporaryDirectory() as tmp:
        conn = _conn(tmp)
        try:
            f = _make_xlsx(tmp, [VM_A, VM_B])
            rv.import_rvtools(f, conn)
            # 第二次：同樣兩台，vm_uuid 相同 → 應該全部走更新
            summary = rv.import_rvtools(f, conn)
            assert summary["inserted"] == 0 and summary["updated"] == 2
            assert conn.execute("SELECT COUNT(*) c FROM hardware").fetchone()["c"] == 2
            # source_record 也不該累積（UNIQUE(source, source_key)）
            assert conn.execute(
                "SELECT COUNT(*) c FROM source_record WHERE source='vcenter'").fetchone()["c"] == 2
        finally:
            conn.close()


def test_換IP換名但vm_uuid不變_仍對到同一台():
    with tempfile.TemporaryDirectory() as tmp:
        conn = _conn(tmp)
        try:
            rv.import_rvtools(_make_xlsx(tmp, [VM_A]), conn)
            moved = {**VM_A, "DNS Name": "web01-renamed.corp.local", "Primary IP Address": "10.9.9.9"}
            summary = rv.import_rvtools(_make_xlsx(tmp, [moved]), conn)
            assert summary["updated"] == 1 and summary["inserted"] == 0
            row = conn.execute("SELECT hostname, ip FROM hardware").fetchone()
            assert row["hostname"] == "web01-renamed.corp.local" and row["ip"] == "10.9.9.9"
            assert conn.execute("SELECT COUNT(*) c FROM hardware").fetchone()["c"] == 1
        finally:
            conn.close()


# ===== 只覆蓋機器事實、不碰業務欄位 =====

def test_更新不覆蓋人填的業務欄位():
    with tempfile.TemporaryDirectory() as tmp:
        conn = _conn(tmp)
        try:
            rv.import_rvtools(_make_xlsx(tmp, [VM_A]), conn)
            # 人替這台填了業務欄位
            conn.execute(
                "UPDATE hardware SET owner=?, asset_purpose=?, custodian=? WHERE vm_uuid=?",
                ("財務處", "核心帳務", "王小明", VM_A["VM UUID"]))
            conn.commit()
            # 再匯一次（OS 有變）
            rv.import_rvtools(_make_xlsx(tmp, [{**VM_A,
                "OS according to the VMware Tools": "Ubuntu Linux (64-bit) 22.04"}]), conn)
            row = conn.execute("SELECT owner, asset_purpose, custodian, os FROM hardware").fetchone()
            assert row["owner"] == "財務處" and row["asset_purpose"] == "核心帳務"
            assert row["custodian"] == "王小明"
            assert row["os"] == "Ubuntu Linux (64-bit) 22.04"   # 機器事實有更新
        finally:
            conn.close()


# ===== 判不準的不自動合併 =====

def test_同IP不同vm_uuid_判ambiguous進佇列不合併():
    """既有一台資產用 10.1.1.10，來了一台不同 vm_uuid 但同 IP——IP 可能被回收，
    絕不能自動當同一台覆蓋掉。應進 merge_review 交人工。"""
    with tempfile.TemporaryDirectory() as tmp:
        conn = _conn(tmp)
        try:
            # 既有資產：有強識別碼 vm_uuid（跟即將匯入的不同），IP 相同
            db.insert_hardware(conn, asset_serial="REAL-001", ip="10.1.1.10",
                               hostname="oldhost", vm_uuid="00000000-0000-0000-0000-000000000001")
            summary = rv.import_rvtools(_make_xlsx(tmp, [VM_A]), conn)  # VM_A 也用 10.1.1.10
            assert summary["pending_review"] == 1
            assert summary["inserted"] == 0 and summary["updated"] == 0
            # 既有那台完全沒被動到
            row = conn.execute("SELECT hostname, vm_uuid FROM hardware WHERE asset_serial='REAL-001'").fetchone()
            assert row["hostname"] == "oldhost"
            assert row["vm_uuid"] == "00000000-0000-0000-0000-000000000001"
            # 進了人工佇列
            mr = conn.execute("SELECT reason, status FROM merge_review").fetchone()
            assert mr is not None and mr["status"] == "open"
        finally:
            conn.close()


def test_找不到vInfo分頁_明確報錯():
    with tempfile.TemporaryDirectory() as tmp:
        wb = openpyxl.Workbook()
        wb.active.title = "SomethingElse"
        wb.active.append(["a", "b", "c"])
        p = Path(tmp) / "bad.xlsx"
        wb.save(p)
        import pytest
        with pytest.raises(ValueError):
            rv.parse_rvtools(p)


# ===== 2026-08-19 使用者拍板「要吃全部」：vInfo 以外的分頁也要收 =====

def _make_multi_sheet_xlsx(tmp: str) -> Path:
    """一份含 vInfo + vHost + vDatastore 三頁的合成 RVTools 匯出。"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "vInfo"
    ws.append(HEADERS)
    ws.append([VM_A.get(h, "") for h in HEADERS])

    host_ws = wb.create_sheet("vHost")
    host_ws.append(["Host", "ESX Version", "# CPU", "# Memory"])
    host_ws.append(["esxi-a.corp.local", "7.0.3", "2", "393216"])

    ds_ws = wb.create_sheet("vDatastore")
    ds_ws.append(["Name", "Capacity MiB", "Free MiB"])
    ds_ws.append(["PROD_A_Datastore", "10485760", "2097152"])

    # 不在 EXTRA_SHEETS 清單裡的分頁——確認不會被誤收
    junk_ws = wb.create_sheet("莫名其妙的分頁")
    junk_ws.append(["x"])
    junk_ws.append(["y"])

    p = Path(tmp) / "rvtools_multi.xlsx"
    wb.save(p)
    return p


def test_解析額外分頁_vHost與vDatastore():
    with tempfile.TemporaryDirectory() as tmp:
        p = _make_multi_sheet_xlsx(tmp)
        extra = rv.parse_extra_sheets(p)
        assert set(extra.keys()) == {"vHost", "vDatastore"}
        assert extra["vHost"][0]["Host"] == "esxi-a.corp.local"
        assert extra["vHost"][0]["ESX Version"] == "7.0.3"
        assert extra["vDatastore"][0]["Name"] == "PROD_A_Datastore"
        # 不認識的分頁不收
        assert "莫名其妙的分頁" not in extra


def test_額外分頁存進source_record不動hardware():
    with tempfile.TemporaryDirectory() as tmp:
        conn = _conn(tmp)
        try:
            before = conn.execute("SELECT COUNT(*) c FROM hardware").fetchone()["c"]
            p = _make_multi_sheet_xlsx(tmp)
            summary = rv.import_rvtools(p, conn)

            assert summary["extra_sheets"] == {"vHost": 1, "vDatastore": 1}

            rows = conn.execute(
                "SELECT payload FROM source_record WHERE source='vcenter_extra:vHost'"
            ).fetchall()
            assert len(rows) == 1
            import json
            payload = json.loads(rows[0]["payload"])
            assert payload["Host"] == "esxi-a.corp.local"
            assert payload["ESX Version"] == "7.0.3"

            # 額外分頁不應該動到 hardware 表——vInfo 那台VM本身有寫入是正常的，
            # 但 vHost/vDatastore 這兩頁不該再多建/多改任何一筆
            after = conn.execute("SELECT COUNT(*) c FROM hardware").fetchone()["c"]
            assert after == before + 1   # 只有 vInfo 那台 VM
        finally:
            conn.close()


def test_重複匯入額外分頁_更新不累加():
    with tempfile.TemporaryDirectory() as tmp:
        conn = _conn(tmp)
        try:
            p = _make_multi_sheet_xlsx(tmp)
            rv.import_rvtools(p, conn)
            rv.import_rvtools(p, conn)
            rows = conn.execute(
                "SELECT COUNT(*) c FROM source_record WHERE source='vcenter_extra:vHost'"
            ).fetchone()
            assert rows["c"] == 1   # 沒有因為匯兩次變成兩筆
        finally:
            conn.close()
