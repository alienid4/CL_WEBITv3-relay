"""S2：Excel 匯入功能。

讀取 CIA 表（硬體/人員/軟體三分頁），依 field_mapping.json 的欄位對應（可調整，不寫死）
把每一列轉成 DB 欄位並寫入。欄位對應用「標題文字比對」，不是固定欄位位置，所以 CIA 表
換欄位順序不會壞（換欄位名稱才需要改 field_mapping.json）。
"""
from __future__ import annotations

import json
import re
import sqlite3
from pathlib import Path
from typing import Any

import openpyxl

MAPPING_PATH = Path(__file__).parent / "field_mapping.json"

# 表頭正規化：同一個欄位在不同版本的檔案裡會長得不一樣，逐字比對會整批匯入失敗。
# 實際踩過的坑：使用者拿表頭修正前匯出的檔來匯入，「資產序號（公司資產編號）」對不上
# 匯入器認的「資產序號」-> asset_serial 變空 -> NOT NULL constraint failed -> 整批掛掉。
# 已知的差異型態（全部來自 field_meta 的 label 與 field_mapping 的鍵不一致，共 8 欄）：
#   資產序號（公司資產編號） / 保管者（SP） / 使用者（AP User）  -> 括號後綴
#   機密性 (C) vs 機密性(C) / 完整性 (I) / 可用性 (A)          -> 括號前多空格
#   BIG IP／VIP vs BIG IP/VIP                                 -> 全形斜線
#   AP ID vs API ID                                            -> 真實字差，正規化橋不了，靠別名
_PAREN_RE = re.compile(r"[（(][^（()）]*[）)]")

# 舊表頭別名：正規化橋不了的「真實字差」才放這裡（差在字母本身，不是標點空白）。
# key/value 都會再經過 _norm 比對，所以這裡照人看得懂的寫法列就好。
#
# 刻意放在程式碼、不放進 field_mapping.json：這是為了讀舊檔的相容性補丁，不是使用者
# 該調的業務設定。放進設定檔的話會在「欄位對應」畫面多出一列長得像重複的項目，
# 使用者很可能把它設成「不匯入」清掉，舊檔就又匯不進來了；而且設定檔的行序會影響
# 匯出表頭，多一個別名等於多一個把正式標題蓋掉的機會。
HEADER_ALIASES = {
    "AP ID": "API ID",  # 舊版匯出用 field_meta 的 label「AP ID」，對應表的正式標題是「API ID」
}


def _norm(header: Any) -> str:
    """把表頭壓成比對用的正規形式：去括號註解、全形轉半形、去空白、大小寫不敏感。

    只用於「比對」，不影響寫回檔案的原始標題。"""
    if not isinstance(header, str):
        return ""
    text = _PAREN_RE.sub("", header)
    text = text.replace("／", "/").replace("　", " ")
    text = re.sub(r"\s+", "", text)
    return text.casefold()

# 這三個分頁對應到哪張表、用哪組欄位辨識「同一筆」。
#
# key          = 單欄主鍵（硬體用資產序號）
# natural_key  = 複合自然鍵，給沒有單一主鍵的分頁用
#
# 為什麼人員/軟體要有 natural_key：原本兩者的 key 都是 None，_upsert 於是一律 INSERT，
# 同一份檔案匯入三次，人員/軟體就變成三筆一模一樣的資料（硬體因為有 key 所以正常）。
# 「改個錯字重新匯入一次」是最常見的操作，等於每修一次錯字資料就髒一次。
#
# 鍵的選法刻意「寧可多列」：鍵取太少會把兩筆真正不同的資料誤判成同一筆而覆蓋掉（資料遺失、
# 事後救不回來）；鍵取太多最多是留下看得見的重複（清得掉）。兩種錯誤的代價不對稱。
SHEET_CONFIG = {
    "硬體": {"table": "hardware", "key": "asset_serial", "natural_key": None},
    "人員": {"table": "personnel", "key": None,
             "natural_key": ("asset_serial", "person_name")},
    "軟體": {"table": "software", "key": None,
             "natural_key": ("asset_serial", "asset_name", "db_software")},
}


def load_mapping() -> dict[str, dict[str, str]]:
    data = json.loads(MAPPING_PATH.read_text(encoding="utf-8"))
    data.pop("_comment", None)
    return data


def _read_sheet_rows(ws, mapping: dict[str, str]) -> list[dict[str, Any]]:
    """用標題列文字比對欄位，不依賴固定欄位位置。

    比對分兩段：先原文精確比對（最快也最不會誤判），對不上才退到 _norm 正規化比對，
    這樣舊版/手改過表頭的檔案也匯得進來，又不會因為正規化太寬鬆而亂配。
    """
    header_cells = next(ws.iter_rows(min_row=1, max_row=1))
    # 正規化後的對照表；同一個正規形式若對到多個欄位，保留第一個（設定檔由上而下為準）
    norm_mapping: dict[str, str] = {}
    for header, field in mapping.items():
        norm_mapping.setdefault(_norm(header), field)
    # 別名指向正式標題，正式標題再查回欄位；別名對不到欄位就當沒這條，不讓它擋路
    for alias, canonical in HEADER_ALIASES.items():
        field = norm_mapping.get(_norm(canonical))
        if field is not None:
            norm_mapping.setdefault(_norm(alias), field)

    col_to_field = {}
    for cell in header_cells:
        raw = cell.value
        header_text = raw.strip() if isinstance(raw, str) else raw
        if header_text in mapping:
            col_to_field[cell.column - 1] = mapping[header_text]
            continue
        field = norm_mapping.get(_norm(raw))
        if field is not None:
            col_to_field[cell.column - 1] = field

    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        record = {}
        for col_idx, field in col_to_field.items():
            if col_idx < len(row):
                value = row[col_idx]
                if value is not None:
                    record[field] = value
        if record:
            rows.append(record)
    return rows


def _upsert(
    conn: sqlite3.Connection,
    table: str,
    key: str | None,
    record: dict[str, Any],
    natural_key: tuple[str, ...] | None = None,
) -> str:
    """判斷這筆是新增還是更新既有資料，回傳 'inserted' 或 'updated'。

    辨識順序：
      1. key（單欄，硬體的 asset_serial）
      2. natural_key（複合欄，人員/軟體）——這組欄位在這筆資料裡必須「全部有值」才算數，
         少一個就無法安全辨識，寧可當新資料 INSERT，也不要拿不完整的鍵去 UPDATE 錯的列。
    兩者都對不上就 INSERT。
    """
    match_cols: tuple[str, ...] | None = None
    if key and key in record:
        match_cols = (key,)
    elif natural_key and all(record.get(col) not in (None, "") for col in natural_key):
        match_cols = tuple(natural_key)

    if match_cols:
        where = " AND ".join(f"{col} = ?" for col in match_cols)
        where_values = [record[col] for col in match_cols]
        existing = conn.execute(
            f"SELECT id FROM {table} WHERE {where}", where_values
        ).fetchone()
        if existing:
            # 鍵本身不用寫回去（值一樣），只更新其餘欄位；若整筆只有鍵欄位就沒東西可改
            updatable = {k: v for k, v in record.items() if k not in match_cols}
            if updatable:
                set_clause = ", ".join(f"{k} = ?" for k in updatable)
                conn.execute(
                    f"UPDATE {table} SET {set_clause} WHERE id = ?",
                    [*updatable.values(), existing["id"]],
                )
            return "updated"

    columns = ", ".join(record.keys())
    placeholders = ", ".join("?" for _ in record)
    conn.execute(f"INSERT INTO {table} ({columns}) VALUES ({placeholders})", tuple(record.values()))
    return "inserted"


def import_excel(xlsx_path: Path, conn: sqlite3.Connection) -> dict[str, Any]:
    mapping = load_mapping()
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)

    summary: dict[str, Any] = {"sheets": {}, "errors": []}

    for sheet_name, cfg in SHEET_CONFIG.items():
        if sheet_name not in wb.sheetnames:
            summary["errors"].append(f"找不到分頁：{sheet_name}")
            continue

        ws = wb[sheet_name]
        rows = _read_sheet_rows(ws, mapping.get(sheet_name, {}))
        inserted = updated = skipped = 0

        for record in rows:
            if cfg["table"] in ("personnel", "software") and "asset_serial" not in record:
                skipped += 1
                continue
            try:
                if cfg["table"] in ("personnel", "software"):
                    exists = conn.execute(
                        "SELECT 1 FROM hardware WHERE asset_serial = ?", (record["asset_serial"],)
                    ).fetchone()
                    if not exists:
                        skipped += 1
                        summary["errors"].append(
                            f"{sheet_name}：資產序號 {record['asset_serial']} 在硬體分頁找不到，略過"
                        )
                        continue
                result = _upsert(
                    conn, cfg["table"], cfg["key"], record, cfg.get("natural_key")
                )
                if result == "updated":
                    updated += 1
                else:
                    inserted += 1
            except sqlite3.IntegrityError as exc:
                skipped += 1
                summary["errors"].append(f"{sheet_name}：寫入失敗 ({exc})")

        summary["sheets"][sheet_name] = {"inserted": inserted, "updated": updated, "skipped": skipped}

    conn.commit()
    return summary


if __name__ == "__main__":
    import sys

    from db import get_connection, init_db

    if len(sys.argv) < 2:
        print("用法：python excel_import.py <ica表.xlsx>")
        raise SystemExit(2)

    init_db()
    connection = get_connection()
    try:
        result = import_excel(Path(sys.argv[1]), connection)
        print(json.dumps(result, ensure_ascii=False, indent=2))
    finally:
        connection.close()
