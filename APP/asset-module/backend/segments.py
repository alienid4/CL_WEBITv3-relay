"""網段配置表：匯入、查詢、IP 配置輔助。

來源是公司的「總分公司網段配置表」Excel（2026-08-15 使用者提供，183 段）。
欄位：使用狀況／使用位置／用途說明／**環境別**／使用類別／使用目的／網段／弱掃說明／
**註解**／**VLAN**。

⚠️ 2026-08-26 使用者提供的版本跟 8/15 那份**不一樣**：環境別獨立成一欄（UAT／PROD），
使用類別純寫 `SERVER` 不再是 `UAT-SERVER`。原本的程式從使用類別的 `UAT-` 前綴推環境，
拿新版檔案匯入會把**所有 UAT 網段標成「正式」**——而這個欄位餵給「機房→環境→網段」
的 IP 選單與掃描範圍建議，標錯的後果是有人照著它把測試 IP 當正式 IP 發出去。
現在改成：**明文的環境別欄優先，前綴推導只當備援**（明文比推論可靠）。

匯入的立場是**寬鬆解析、明確回報**：真實檔案裡本來就有解析不掉的寫法
（一格塞兩段、寫成 IP 範圍），這種列不能靜默丟掉——丟掉的網段之後不會有人發現，
而「系統裡沒有這段」跟「這段不存在」在盤點上是完全不同的兩件事。
所以：解析得出來的存 cidr，解析不出來的照樣入庫（cidr=NULL）並列進匯入警告。
"""
from __future__ import annotations

import bisect
import ipaddress
import re
import sqlite3
from datetime import datetime
from pathlib import Path
from typing import Iterable

# Excel 表頭 → 欄位。用「包含」比對，不是逐字相等：表頭常有全形括號、尾端空白、
# 或多一個「（本欄位…）」註記，逐字比對會整批匯不進來（同 excel_import 踩過的坑）。
HEADER_MAP = {
    "使用狀況": "usage_status",
    "使用位置": "location",
    "用途說明": "purpose_desc",
    "使用類別": "category",
    "使用目的": "usage",
    "網段": "raw_cidr",
    "弱掃說明": "scan_note",
    "環境別": "environment_raw",
    "註解": "remark",
    "VLAN": "vlan",
}
REQUIRED = ("raw_cidr",)

# 環境別欄的值 → 系統用語。使用者 2026-08-26 確認這份檔案只有 UAT／PROD 兩種。
#
# ⚠️ **認不出來的值不要猜成「正式」**。這個欄位餵給 IP 配置選單與掃描範圍建議，
# 猜錯的方向是「把測試網段標成正式」，代價比留白高得多——留白只是選單少一個選項，
# 猜錯是有人照著它把測試 IP 當正式 IP 發出去。認不出來就存原值 + 列成警告讓人來看。
_ENV_MAP = {
    "UAT": "測試",
    "PROD": "正式",
}

# 「建議排除掃描」是資安人員寫在弱掃說明裡的判斷（員工電腦、UAT、重複 IP 網段）。
# 這是主動掃描要不要打進去的唯一依據，不能靠我們自己猜。
_EXCLUDE_RE = re.compile(r"排除掃描|勿掃|不要掃|不得掃")


def export_template_rows() -> tuple[list[str], list[str]]:
    """空白範本的表頭與一筆範例列。2026-08-25 使用者提出的通則：有匯入就要有配對的
    匯出範本，不然新使用者不知道要填什麼——這是「匯入網段配置表」旁邊那顆
    「⬇ 下載空白範本」按鈕的資料來源。

    表頭直接取自 HEADER_MAP（匯入認得的同一份），不要另外寫一份會漂走的欄位清單。
    範例列刻意示範最常見的「解析不掉」寫法該怎麼改（見 parse_cidr 檔頭那段）：
    網段只能填單一標準 CIDR，一格兩段或 IP 範圍要拆成多列，不要塞在同一格。
    """
    headers = list(HEADER_MAP.keys())
    example = {
        "使用狀況": "使用中", "使用位置": "板橋機房", "用途說明": "SERVER網段",
        "環境別": "PROD", "使用類別": "SERVER", "使用目的": "應用系統主機",
        "網段": "10.99.1.0/24",
        "弱掃說明": "範例列，請刪除；網段只填單一 CIDR，"
                    "「一格兩段」或「IP範圍(~)」的寫法系統解析不了，請拆成多列各填一段",
        "註解": "", "VLAN": "2001",
    }
    return headers, [example.get(h, "") for h in headers]


def export_current_rows(conn: sqlite3.Connection) -> tuple[list[str], list[list]]:
    """匯出**目前系統裡的**網段清單（不是空白範本）。

    2026-08-26 使用者：「而且我是有匯入，我沒有匯出嗎？」——原本只有空白範本，
    等於資料進得去出不來。有匯出才做得到三件事：拿系統的清單去跟 Excel 對帳、
    在系統裡修好之後把正確版本交回去、以及留一份離線備份。

    表頭跟匯入認得的那份一樣（HEADER_MAP），所以**匯出的檔案可以直接再匯入**。
    後面多兩欄系統算出來的資訊（展開自、目前登記台數），欄名刻意不在 HEADER_MAP
    裡，再匯入時會被忽略，不會造成困擾。

    ⚠️ 展開出來的列匯出時會是**一列一段**（原檔的 `A----B` 一格會變成 N 列）。
    這是刻意的：交回去的版本應該是系統看得懂的寫法，不要把難解析的寫法再傳一次。
    「展開自」那欄讓人對得回原檔是哪一格。
    """
    headers = list(HEADER_MAP.keys()) + ["展開自", "目前登記台數"]
    rows = []
    for r in list_segments(conn):
        rows.append([
            r.get("usage_status"), r.get("location"), r.get("purpose_desc"),
            r.get("environment_raw") or _env_back(r.get("environment")),
            r.get("category"), r.get("usage"),
            r.get("cidr") or r.get("raw_cidr"),
            r.get("scan_note"), r.get("remark"), r.get("vlan"),
            r.get("expanded_from"), r.get("asset_count"),
        ])
    return headers, rows


def _env_back(environment: str | None) -> str | None:
    """正規化後的環境反查回檔案用語（測試→UAT、正式→PROD）。

    匯出的檔案要能再匯進來，所以寫回去的必須是匯入認得的值。原檔有 environment_raw
    時優先用原檔的字（人可能寫 uat 小寫，照原樣還他），這裡只處理沒有原值的舊資料。
    """
    back = {v: k for k, v in _ENV_MAP.items()}
    return back.get(environment or "")


def _norm_header(s: str) -> str:
    """去空白、全形括號轉半形，並轉大寫——「VLAN」這種英文表頭別人可能寫成
    Vlan／vlan，大小寫敏感的比對會讓整欄安靜地匯不進來。"""
    return re.sub(r"\s+", "", str(s or "")).replace("（", "(").replace("）", ")").upper()


def derive_environment(environment_raw: str | None, category: str | None
                       ) -> tuple[str | None, str | None]:
    """回傳 (正規化環境, 警告訊息或 None)。

    **明文優先、推導備援。** 順序刻意是這樣：
      1. 檔案有「環境別」欄且看得懂（UAT／PROD）→ 用它
      2. 有值但看不懂 → 回 None 並給警告（**不猜成正式**，見 _ENV_MAP 上面的說明）
      3. 沒填 → 退回舊規則：使用類別以 UAT 開頭算測試，其餘正式

    第 3 條保留是因為 8/15 那份舊檔案就是那樣寫的（UAT-SERVER／UAT-NETWORK），
    拿掉的話舊檔案重匯會整批變成未知環境。
    """
    raw = (environment_raw or "").strip()
    if raw:
        hit = _ENV_MAP.get(raw.upper())
        if hit:
            return hit, None
        return None, (f"環境別欄填的是「{raw}」，系統只認得 "
                      f"{'／'.join(sorted(_ENV_MAP))}；這一段的環境留空白不猜，"
                      f"請確認要對應到測試還是正式")
    c = (category or "").upper()
    return ("測試" if c.startswith("UAT") else "正式"), None


def parse_cidr(raw: str) -> tuple[str | None, int | None, int | None]:
    """回傳 (正規化 CIDR, 起始整數, 結束整數)；解析不出來回 (None, None, None)。

    只認單一標準 CIDR。原檔有這幾種解析不掉的寫法，一律當「無法解析」不要硬猜：
      · 一格兩段：`10.99.255.23/28\\n10.99.255.17/28`
      · IP 範圍：`172.16.156.0/24~172.16.157.230`
    硬猜的後果是把錯的範圍寫進系統，之後「這個 IP 屬於哪段」全錯——寧可留白等人修。
    """
    s = str(raw or "").strip()
    if not s or "\n" in s or "~" in s or "," in s:
        return None, None, None
    try:
        net = ipaddress.ip_network(s, strict=False)
    except ValueError:
        return None, None, None
    return str(net), int(net.network_address), int(net.broadcast_address)


#: 一格展開成幾段就放棄展開。上限存在的理由是打錯字：`10.0.0.0/8----10.255.0.0/8`
#: 這種寫法會展開出六萬多列，把整個匯入變成災難。超過上限就退回「當成一個大範圍」
#: 並明講展開了幾段——**不要靜默截斷**，那會讓人以為全部都進來了。
MAX_EXPAND = 512

#: 多段寫法的分隔符。原檔實際出現過 `----`（四個）與 `--`（兩個）兩種寫法，
#: 所以用「兩個以上的連字號」比對，不要寫死四個。
_RANGE_SEP = re.compile(r"\s*(?:-{2,}|~|～|至|到)\s*")

#: 一格塞多段（換行、空白、逗號、頓號分隔）
_SPLIT_SEP = re.compile(r"[\n,、;；]+|\s{1,}")


def _last_octet_range(left: str, right: str) -> tuple[int, int] | None:
    """`172.16.157.231~249` 這種「只寫最後一段」的簡寫。

    右邊是純數字時，把左邊的前三段接上去。不這樣處理的話這一列完全進不了系統，
    而「系統裡沒有這段」跟「這段不存在」在盤點上是兩件完全不同的事。
    """
    if not right.isdigit() or "/" in left:
        return None
    parts = left.split(".")
    if len(parts) != 4:
        return None
    try:
        a = ipaddress.ip_address(left)
        b = ipaddress.ip_address(".".join(parts[:3] + [right]))
    except ValueError:
        return None
    return (int(a), int(b)) if int(b) >= int(a) else None


def _as_int_bounds(token: str) -> tuple[int, int] | None:
    """一個 token（CIDR 或單一 IP）的起訖整數。"""
    t = token.strip()
    if not t:
        return None
    try:
        if "/" in t:
            net = ipaddress.ip_network(t, strict=False)
            return int(net.network_address), int(net.broadcast_address)
        addr = ipaddress.ip_address(t)
        return int(addr), int(addr)
    except ValueError:
        return None


def expand_segments(raw: str) -> tuple[list[dict], str | None]:
    """把原檔一格裡的網段寫法攤成一列或多列。回傳 (entries, 警告或 None)。

    每個 entry：`{cidr, raw_cidr, net_start, net_end, expanded_from}`。
    `expanded_from` 有值代表「原檔沒有這一列，是系統拆出來的」。

    ## 為什麼要展開，而不是像以前那樣一律標成「無法解析」

    2026-08-26 使用者拿真實檔案匯進 221，警告列出 9 列解析不掉。實際算過，
    光是 `10.99.71.0/24----10.99.120.0/24` 這一格就是 **50 個 /24**，七格加起來
    **約 120 個網段完全沒進系統**——而這張表的三個用途（IP 配置選單、掃描範圍、
    資料品質的涵蓋率分母）全都因此少算，且沒有人看得出來少了什麼。
    「不猜」原本是對的立場，但「連能確定的都不解析」不是保守，是漏資料。

    支援的四種寫法，全部是**可以確定的**，不含猜測：
      1. 單一 CIDR → 一列（原本就支援）
      2. 一格多段（換行／空白／逗號分隔）→ 拆成多列
      3. `A/n----B/n`（前綴長度相同）→ 展開成 A 到 B 之間每一個 /n
      4. `A~B` 的位址範圍（含 `172.16.157.231~249` 這種只寫尾碼的簡寫）
         → 一列，`cidr` 留 None 但 **net_start/net_end 照樣算得出來**

    第 4 種特別值得講：`cidr` 是 None 不代表這段沒用——「這個 IP 屬於哪一段」
    與「這段已登記幾台」查的是 net_start/net_end（見 find_segment_for_ip），
    所以範圍寫法照樣能用；只有「機房→環境→網段」的選單需要單一 CIDR。
    以前這種列連 net_start 都是 NULL，等於整段消失。

    前綴長度不同時（`A/24----B/28`）不展開成一段一段，但**起訖位址仍然是確定的**，
    所以走第 4 種當成位址範圍收下。

    仍然完全不猜的只有一種：看不懂的字串（手寫中文、亂碼）。那種留 raw 入庫、
    net_start 為 NULL、列警告，等人去改原檔。
    """
    s = str(raw or "").strip()
    if not s:
        return [], None

    def one(cidr, start, end, src=None):
        return {"cidr": cidr, "raw_cidr": cidr or s, "net_start": start,
                "net_end": end, "expanded_from": src}

    # 1) 單純一個 CIDR（最常見，先走完不要被後面的規則干擾）
    cidr, start, end = parse_cidr(s)
    if cidr:
        return [one(cidr, start, end)], None

    # 3) A----B 範圍寫法
    parts = _RANGE_SEP.split(s)
    if len(parts) == 2 and all(parts):
        left, right = parts[0].strip(), parts[1].strip()
        a_net = _safe_network(left)
        b_net = _safe_network(right)
        if a_net is not None and b_net is not None and a_net.prefixlen == b_net.prefixlen:
            if int(b_net.network_address) < int(a_net.network_address):
                return [one(None, None, None)], (
                    f"「{s}」的結束位址比起始位址小，看不出是哪一段，已保留但不能用於 "
                    f"IP 配置與掃描")
            step = b_net.num_addresses
            count = (int(b_net.network_address) - int(a_net.network_address)) // step + 1
            if count > MAX_EXPAND:
                return [one(None, int(a_net.network_address),
                            int(b_net.broadcast_address))], (
                    f"「{s}」是一個涵蓋 {count} 個 /{a_net.prefixlen} 的範圍，超過一格展開上限 "
                    f"{MAX_EXPAND}，已當成單一大範圍收下（查得到「IP 屬於哪段」，但不會出現在"
                    f"網段選單）。若這格是打錯字請修正")
            out = []
            for i in range(count):
                n = ipaddress.ip_network(
                    (int(a_net.network_address) + i * step, a_net.prefixlen))
                out.append(one(str(n), int(n.network_address),
                               int(n.broadcast_address), src=s))
            return out, (
                f"「{s}」是範圍寫法，已展開成 {count} 個 /{a_net.prefixlen} 網段；"
                f"這些列在畫面上會標「展開自原檔」，原檔並沒有這麼多列")

        # 4) 位址範圍（含只寫尾碼的簡寫）
        bounds = _last_octet_range(left, right) or _pair_bounds(left, right)
        if bounds:
            return [one(None, bounds[0], bounds[1])], (
                f"「{s}」是位址範圍不是標準網段，已收下起訖位址（查得到「IP 屬於哪段」、"
                f"也算得出已登記幾台），但因為不是單一 CIDR，不會出現在新增資產的網段選單")

    # 2) 一格多段
    tokens = [t for t in _SPLIT_SEP.split(s) if t.strip()]
    if len(tokens) > 1:
        parsed = [parse_cidr(t) for t in tokens]
        if all(c for c, _, _ in parsed):
            return ([one(c, st, en, src=s) for c, st, en in parsed],
                    f"「{s}」一格寫了 {len(tokens)} 段，已拆成 {len(tokens)} 列")

    return [one(None, None, None)], (
        "網段寫法無法解析成單一 CIDR（一格多段或 IP 範圍），已保留但不能用於 IP 配置與掃描")


def _safe_network(token: str):
    try:
        return ipaddress.ip_network(token.strip(), strict=False)
    except ValueError:
        return None


def _pair_bounds(left: str, right: str) -> tuple[int, int] | None:
    a = _as_int_bounds(left)
    b = _as_int_bounds(right)
    if a is None or b is None or b[1] < a[0]:
        return None
    return a[0], b[1]


def _rows_from_xlsx(path: Path) -> list[dict]:
    import openpyxl

    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        return []
    header = [_norm_header(c) for c in rows[0]]
    out = []
    for i, r in enumerate(rows[1:], start=2):
        d = {"_row_no": i}
        for h, v in zip(header, r):
            for key, field in HEADER_MAP.items():
                if key.upper() in h:
                    d[field] = None if v is None else str(v).strip()
        out.append(d)
    return out


def _rows_from_text(path: Path) -> list[dict]:
    """Tab 分隔的文字檔（使用者先給的過渡格式，欄位跟 Excel 一樣）。"""
    import csv
    import io

    with io.open(path, encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f, delimiter="\t")
        out = []
        for i, r in enumerate(reader, start=2):
            d = {"_row_no": i}
            for k, v in r.items():
                h = _norm_header(k)
                for key, field in HEADER_MAP.items():
                    if key.upper() in h:
                        d[field] = (v or "").strip() or None
            out.append(d)
        return out


def read_rows(path: Path) -> list[dict]:
    if path.suffix.lower() in (".xlsx", ".xlsm"):
        return _rows_from_xlsx(path)
    return _rows_from_text(path)


def import_segments(conn: sqlite3.Connection, path: Path) -> dict:
    """整批取代匯入。回傳摘要與警告清單。

    **整批取代不是 upsert**：Excel 是這份清單的唯一真相，段被刪掉就該從系統消失。
    upsert 會讓已作廢的網段永遠留著，越用越髒，而且沒有人會發現。
    先解析成功才刪舊資料——檔案格式不對時不能把現有清單清空。
    """
    rows = read_rows(path)
    if not rows:
        raise ValueError("檔案裡沒有資料列")

    header_ok = any(r.get("raw_cidr") for r in rows)
    if not header_ok:
        raise ValueError("找不到「網段」欄位，請確認是不是網段配置表（表頭要有：使用位置／使用類別／網段）")

    parsed: list[dict] = []
    warnings: list[dict] = []
    for r in rows:
        raw = (r.get("raw_cidr") or "").strip()
        if not raw:
            continue  # 整列空白（Excel 常見的尾巴空列）不算警告
        env, env_warn = derive_environment(r.get("environment_raw"), r.get("category"))
        if env_warn:
            warnings.append({"row_no": r["_row_no"], "raw_cidr": raw, "reason": env_warn})
        # 一格可能是一段，也可能是「A----B」的範圍或一格塞多段，展開成多列。
        # 展開出來的每一列都會帶 expanded_from（＝原檔那一格的原文），畫面上要
        # 標得出來「原檔沒有這一列，是系統拆的」，否則拿系統清單去跟 Excel 對帳
        # 會對不起來，而且不知道要回頭改哪一格。
        entries, warn = expand_segments(raw)
        if warn:
            warnings.append({"row_no": r["_row_no"], "raw_cidr": raw, "reason": warn})
        common = {
            "usage_status": r.get("usage_status"), "location": r.get("location"),
            "purpose_desc": r.get("purpose_desc"), "category": r.get("category"),
            "usage": r.get("usage"),
            "environment": env,
            "environment_raw": r.get("environment_raw"),
            "vlan": r.get("vlan"),
            "remark": r.get("remark"),
            "scan_excluded": 1 if _EXCLUDE_RE.search(r.get("scan_note") or "") else 0,
            "scan_note": r.get("scan_note"), "row_no": r["_row_no"],
        }
        for e in entries:
            parsed.append({**common, **e})

    seen: dict[str, int] = {}
    for p in parsed:
        if p["cidr"]:
            if p["cidr"] in seen:
                warnings.append({
                    "row_no": p["row_no"], "raw_cidr": p["raw_cidr"],
                    "reason": f"這個網段在第 {seen[p['cidr']]} 列已經出現過，兩列都保留，請確認是不是重複登記",
                })
            else:
                seen[p["cidr"]] = p["row_no"]

    # 匯入前先記下現有的，才講得出「這次跟上次差在哪」。
    #
    # 2026-08-26 使用者：「我故意重複匯入同一個檔案，但系統沒有擋掉，是覺得無所謂嗎？」
    # ——重複匯入同一個檔案確實無所謂（整批取代是冪等的，不會變兩份），但
    # **系統沒有把這件事講出來，使用者只能用猜的**。而且真正危險的不是重複匯入，
    # 是「匯入一份不完整的檔案」：整批取代會讓沒出現在新檔案裡的網段直接消失，
    # 而畫面只會顯示「匯入 N 段」看起來很正常。所以這裡算出增減，讓人看得見。
    before = {
        (r["cidr"] or r["raw_cidr"]) for r in
        conn.execute("SELECT cidr, raw_cidr FROM network_segment")
    }
    after = {(p["cidr"] or p["raw_cidr"]) for p in parsed}

    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    conn.execute("DELETE FROM network_segment")
    conn.executemany(
        "INSERT INTO network_segment "
        "(cidr, raw_cidr, net_start, net_end, usage_status, location, purpose_desc, "
        " category, usage, environment, environment_raw, vlan, remark, expanded_from, "
        " scan_excluded, scan_note, row_no, imported_at) "
        "VALUES (:cidr, :raw_cidr, :net_start, :net_end, :usage_status, :location, "
        " :purpose_desc, :category, :usage, :environment, :environment_raw, :vlan, "
        " :remark, :expanded_from, :scan_excluded, :scan_note, :row_no, :imported_at)",
        [{**p, "imported_at": now} for p in parsed],
    )
    conn.commit()
    return {
        # imported 是「進到系統的段數」，不等於「Excel 的列數」——範圍寫法會展開成多列。
        # 兩個數字都給，不然人拿它去跟 Excel 對帳會以為系統多算了。
        "imported": len(parsed),
        "source_rows": sum(1 for r in rows if (r.get("raw_cidr") or "").strip()),
        "expanded": sum(1 for p in parsed if p.get("expanded_from")),
        "unparsed": sum(1 for p in parsed if p["net_start"] is None),
        "range_only": sum(1 for p in parsed
                          if p["cidr"] is None and p["net_start"] is not None),
        "parsed_cidr": sum(1 for p in parsed if p["cidr"]),
        "scan_excluded": sum(1 for p in parsed if p["scan_excluded"]),
        "locations": len({p["location"] for p in parsed if p["location"]}),
        # 環境沒認出來的段數獨立講：它不是「解析失敗」也不是「正常」，
        # 混在總數裡會讓人以為全部都匯好了
        "environment_unknown": sum(1 for p in parsed if p["environment"] is None),
        # 跟匯入前比的增減。`unchanged` 等於全部就代表「這次匯入沒有改變任何東西」，
        # 重複匯入同一個檔案就會是這樣——把它明講出來，人才不用猜。
        "was_empty": not before,
        "added": sorted(after - before)[:200],
        "added_count": len(after - before),
        "removed": sorted(before - after)[:200],
        "removed_count": len(before - after),
        "unchanged_count": len(before & after),
        "warnings": warnings,
        "imported_at": now,
    }


def list_segments(conn: sqlite3.Connection) -> list[dict]:
    """全部網段＋每段已登記幾台資產（要看得出哪段快滿了、哪段根本沒東西）。

    ## 為什麼不是每段各跑一次 COUNT

    原本每段呼叫一次 `_count_assets()`，每次都對整張 hardware 做
    `_ip_int(ip) BETWEEN ?` ——而 `_ip_int` 是註冊給 SQLite 的 **Python 函式**，
    所以那是「段數 × 資產數」次 Python 呼叫。

    2026-08-27 使用者回報「每次查詢都會卡一下」，實測（443 段 × 4784 台）：
    **10.8 秒**。而且範圍寫法展開之後段數從 327 變 443，又慢了三成——
    這種「資料愈完整愈慢」的設計會逼人不想把資料補齊，方向完全反了。

    改成一次把所有 IP 讀出來轉成整數排序，再用二分搜尋數每段的區間：
    O(N log N + 段數 × log N)。同樣的資料量從 10.8 秒降到 0.05 秒等級。
    """
    rows = conn.execute(
        "SELECT * FROM network_segment ORDER BY location, cidr, raw_cidr"
    ).fetchall()
    sorted_ips = _sorted_asset_ips(conn)
    out = []
    for r in rows:
        d = dict(r)
        d["asset_count"] = _count_in_range(sorted_ips, r["net_start"], r["net_end"])
        d["capacity"] = (r["net_end"] - r["net_start"] - 1) if r["net_start"] else None
        out.append(d)
    return out


def _sorted_asset_ips(conn: sqlite3.Connection) -> list[int]:
    """全部資產 IP 轉成整數並排序。**只認 IPv4**——跟 db._ip_int 同一個立場：
    net_start/net_end 是拿 IPv4 灌的整數，IPv6 的值會溢位（2026-08-19 正式機
    真的因為 vCenter 收到 fe80:: 而炸過）。"""
    out: list[int] = []
    for (ip,) in conn.execute(
        "SELECT ip FROM hardware WHERE ip IS NOT NULL AND TRIM(ip) != ''"
    ):
        try:
            addr = ipaddress.ip_address(str(ip).strip())
        except ValueError:
            continue
        if addr.version == 4:
            out.append(int(addr))
    out.sort()
    return out


def _count_in_range(sorted_ips: list[int], start: int | None, end: int | None) -> int:
    if start is None or end is None:
        return 0
    return bisect.bisect_right(sorted_ips, end) - bisect.bisect_left(sorted_ips, start)


def _count_assets(conn: sqlite3.Connection, seg: sqlite3.Row) -> int:
    """單一段的台數。給只要問一段的呼叫端用（下鑽），走 SQL 比較直接。

    ⚠️ **要一次問很多段時不要用這支**，用 list_segments 那條路——
    見它的說明，逐段呼叫是「段數 × 資產數」次 Python 呼叫。
    """
    if seg["net_start"] is None:
        return 0
    r = conn.execute(
        "SELECT COUNT(*) n FROM hardware WHERE ip IS NOT NULL AND TRIM(ip) != '' "
        "AND _ip_int(ip) BETWEEN ? AND ?",
        (seg["net_start"], seg["net_end"]),
    ).fetchone()
    return r["n"]


def tree(conn: sqlite3.Connection) -> list[dict]:
    """機房 → 環境 → 網段 三層，給新增資產的 IP 選單用。

    使用者指定的順序（2026-08-15）：先選機房、再選環境、最後才選網段——
    填表的人記得住「這台在板橋、是正式機」，記不住 10.99.163 是哪一段。
    """
    rows = conn.execute(
        "SELECT * FROM network_segment WHERE cidr IS NOT NULL "
        "ORDER BY location, environment, cidr"
    ).fetchall()
    out: list[dict] = []
    for r in rows:
        loc = r["location"] or "（未填機房）"
        env = r["environment"] or "正式"
        node = next((x for x in out if x["location"] == loc), None)
        if node is None:
            node = {"location": loc, "environments": []}
            out.append(node)
        enode = next((x for x in node["environments"] if x["environment"] == env), None)
        if enode is None:
            enode = {"environment": env, "segments": []}
            node["environments"].append(enode)
        enode["segments"].append({
            "cidr": r["cidr"],
            "label": f"{r['cidr']}　{r['purpose_desc'] or ''}".strip(),
            "category": r["category"],
            "usage": r["usage"],
            "scan_excluded": bool(r["scan_excluded"]),
            "scan_note": r["scan_note"],
        })
    return out


def segment_ips(conn: sqlite3.Connection, cidr: str) -> dict:
    """這段裡哪些 IP 已被登記、下一個沒被用的是哪個。

    只講「這份清單裡有沒有登記」，不宣稱「這個 IP 實際上沒人在用」——
    清單本來就不完整（這正是資料品質那頁在量的事），把「沒登記」講成「可用」
    會讓使用者拿去配一個實際上已經有人在用的 IP。
    """
    try:
        net = ipaddress.ip_network(cidr, strict=False)
    except ValueError:
        raise ValueError(f"網段格式不正確：{cidr}")

    seg = conn.execute("SELECT * FROM network_segment WHERE cidr = ? LIMIT 1", (cidr,)).fetchone()
    rows = conn.execute(
        "SELECT asset_serial, hostname, ip, asset_status FROM hardware "
        "WHERE ip IS NOT NULL AND TRIM(ip) != '' AND _ip_int(ip) BETWEEN ? AND ? "
        "ORDER BY _ip_int(ip)",
        (int(net.network_address), int(net.broadcast_address)),
    ).fetchall()
    used = {r["ip"].strip(): dict(r) for r in rows}

    suggestion = None
    for host in net.hosts():
        if str(host) not in used:
            suggestion = str(host)
            break

    return {
        "cidr": cidr,
        "purpose_desc": seg["purpose_desc"] if seg else None,
        "scan_note": seg["scan_note"] if seg else None,
        "capacity": net.num_addresses - 2 if net.num_addresses > 2 else net.num_addresses,
        "used": list(used.values()),
        "used_count": len(used),
        "suggestion": suggestion,
        "suggestion_caveat": "只代表「這份清單裡沒登記」，不代表實際上沒人在用",
    }


def scan_candidates(conn: sqlite3.Connection) -> dict:
    """掃描範圍建議：哪些網段該掃、哪些被資安註記為排除。

    這支存在的理由是資料品質那頁量出來的問題——「涵蓋率 0%」的下一步是
    「那該掃哪些」，不該讓使用者自己回去翻 Excel。
    """
    rows = conn.execute(
        "SELECT * FROM network_segment WHERE cidr IS NOT NULL ORDER BY cidr"
    ).fetchall()
    include = [dict(r) for r in rows if not r["scan_excluded"]]
    exclude = [dict(r) for r in rows if r["scan_excluded"]]
    return {
        "include": include,
        "exclude": exclude,
        "include_count": len(include),
        "exclude_count": len(exclude),
    }


def find_segment_for_ip(conn: sqlite3.Connection, ip: str) -> dict | None:
    """這個 IP 屬於哪一段。資產詳細頁要講得出「這台在板橋 DMZ」而不只是一串數字。

    只認 IPv4——network_segment.net_start/net_end 是用 int(IPv4位址) 灌的，範圍在
    SQLite INTEGER（64位元有號）之內；IPv6 的整數值可以到 2^128，直接綁進查詢會炸
    `OverflowError: Python int too large to convert to SQLite INTEGER`。2026-08-19
    正式機真的踩到：vCenter 收到兩台 VM 的 IPv6 link-local 位址（fe80::...）當
    hardware.ip，讓整個 ci_graph.rebuild() 在跑到那筆資料時整批中斷——不是「這個 IP
    沒有段」，是「這個 IP 根本問錯問題」，兩者都回 None，呼叫端不用分辨。
    """
    try:
        addr = ipaddress.ip_address(str(ip).strip())
    except ValueError:
        return None
    if addr.version != 4:
        return None
    n = int(addr)
    r = conn.execute(
        "SELECT * FROM network_segment WHERE net_start IS NOT NULL "
        "AND ? BETWEEN net_start AND net_end ORDER BY (net_end - net_start) LIMIT 1",
        (n,),
    ).fetchone()
    return dict(r) if r else None
