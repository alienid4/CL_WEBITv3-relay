"""業務系統對照表：`api_id` → 系統名稱、AP 部門、AP 負責人。

## 為什麼需要這張表

資產庫裡只有 `api_id`（`N-008` 這種代碼），沒有中文名稱。畫面上到處是看不懂的
代碼，而帳號盤點要交出去的 Excel 第一欄就是 `system_id` ＋ `system`。

在這之前，系統名稱是拿同一個 `api_id` 底下 **`MIN(asset_name)`** 湊出來的
（見 api.py 的全域搜尋與 blast_radius）——那是「隨便挑一台機器的名字當系統名」，
猜對是運氣。有了這張表就有正式來源。

## 為什麼 AP 部門與 AP 負責人也放這裡

使用者提供的範例裡，`system_id`／`system`／`ap_department`／`ap_owner` 四欄
在同一個系統的每一列都**重複同一組值**——那是**業務系統的屬性**，不是機器的屬性。
放在機器欄位上會有 N 份副本，改一次要改 N 台。

## 空白的兩種原因要分得開

匯出時 `system` 欄空白有兩種完全不同的意思：
  · 這台機器**沒填 api_id**            → 要去補資產資料
  · 有 api_id，但**對照表裡沒有這個代碼** → 要去補對照表
兩者長得一樣的話，人不知道該補哪一邊。`lookup()` 回的 dict 用 `found` 分開。
"""
from __future__ import annotations

import sqlite3
from pathlib import Path
from typing import Any

#: 匯入時可接受的欄名（大小寫、全半形、前後空白都會正規化後比對）。
#: 不寫死單一欄名是專案慣例（決策 D14）——來源 Excel 的表頭常常換。
_COLUMN_CANDIDATES: dict[str, tuple[str, ...]] = {
    "api_id": ("system_id", "ap_id", "apid", "api_id", "系統代碼", "系統編號"),
    "name": ("system", "system_name", "系統名稱", "業務系統", "系統"),
    "ap_department": ("ap_department", "ap_dept", "AP部門", "AP 部門", "應用部門"),
    "ap_owner": ("ap_owner", "AP負責人", "AP 負責人", "應用負責人", "負責人"),
}


def _norm(s: Any) -> str:
    return str(s or "").strip().lower().replace(" ", "").replace("　", "")


def upsert(conn: sqlite3.Connection, api_id: str, name: str | None = None,
           ap_department: str | None = None, ap_owner: str | None = None) -> None:
    """寫入一筆對照。以 api_id 為鍵——對照表會重匯（改名、加新系統），
    重匯必須是更新同一筆而不是長出重複。"""
    conn.execute(
        "INSERT INTO business_system (api_id, name, ap_department, ap_owner) "
        "VALUES (?,?,?,?) "
        "ON CONFLICT(api_id) DO UPDATE SET "
        "  name = COALESCE(excluded.name, name), "
        "  ap_department = COALESCE(excluded.ap_department, ap_department), "
        "  ap_owner = COALESCE(excluded.ap_owner, ap_owner), "
        "  updated_at = datetime('now','localtime')",
        (api_id.strip(), (name or "").strip() or None,
         (ap_department or "").strip() or None, (ap_owner or "").strip() or None),
    )


def lookup(conn: sqlite3.Connection, api_id: str | None) -> dict:
    """查一個 api_id。**空白的原因要分得開**（見模組 docstring）。

    回 `{"found": bool, "reason": str|None, ...}`：
      · api_id 是空的      → found=False, reason="機器沒填 api_id"
      · 對照表裡查不到     → found=False, reason="對照表沒有這個代碼"
      · 查到               → found=True,  reason=None
    """
    if not (api_id or "").strip():
        return {"found": False, "reason": "機器沒填 api_id",
                "api_id": None, "name": None, "ap_department": None, "ap_owner": None}
    row = conn.execute(
        "SELECT api_id, name, ap_department, ap_owner FROM business_system "
        "WHERE api_id = ?", (api_id.strip(),)).fetchone()
    if row is None:
        return {"found": False, "reason": "對照表沒有這個代碼",
                "api_id": api_id.strip(), "name": None,
                "ap_department": None, "ap_owner": None}
    return {"found": True, "reason": None, **dict(row)}


def list_all(conn: sqlite3.Connection) -> list[dict]:
    """全部對照，附「這個系統目前有幾台機器」——匯入後要看得出對得上多少。"""
    rows = conn.execute(
        "SELECT b.api_id, b.name, b.ap_department, b.ap_owner, b.updated_at, "
        "       (SELECT COUNT(*) FROM hardware h WHERE h.api_id = b.api_id) AS asset_count "
        "FROM business_system b ORDER BY b.api_id").fetchall()
    return [dict(r) for r in rows]


def coverage(conn: sqlite3.Connection) -> dict:
    """對帳用：資產庫裡的 api_id 有多少對得到對照表。

    只給「已匯入 N 筆」沒有用——人要知道的是**還有多少台查不到名字**，
    以及那是「對照表缺代碼」還是「機器沒填 api_id」。
    """
    total = conn.execute("SELECT COUNT(*) FROM hardware").fetchone()[0]
    no_apid = conn.execute(
        "SELECT COUNT(*) FROM hardware "
        "WHERE api_id IS NULL OR TRIM(api_id) = ''").fetchone()[0]
    unmatched = conn.execute(
        "SELECT COUNT(*) FROM hardware h WHERE h.api_id IS NOT NULL AND TRIM(h.api_id) != '' "
        "AND NOT EXISTS (SELECT 1 FROM business_system b WHERE b.api_id = h.api_id)"
    ).fetchone()[0]
    missing_codes = [r[0] for r in conn.execute(
        "SELECT DISTINCT h.api_id FROM hardware h "
        "WHERE h.api_id IS NOT NULL AND TRIM(h.api_id) != '' "
        "AND NOT EXISTS (SELECT 1 FROM business_system b WHERE b.api_id = h.api_id) "
        "ORDER BY h.api_id LIMIT 50")]
    return {
        "mapped_systems": conn.execute(
            "SELECT COUNT(*) FROM business_system").fetchone()[0],
        "assets_total": total,
        "assets_without_api_id": no_apid,          # 要去補資產資料
        "assets_with_unmapped_api_id": unmatched,  # 要去補對照表
        "unmapped_codes": missing_codes,           # 具名列出，不要只給數字
    }


def import_xlsx(path: Path, conn: sqlite3.Connection) -> dict:
    """吃一份對照表 Excel。第一列當表頭，欄名比對 `_COLUMN_CANDIDATES`。

    `api_id` 是必要欄；其餘缺了就留空（COALESCE 不會把既有值洗掉）。
    """
    import openpyxl

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb[wb.sheetnames[0]]
        rows = ws.iter_rows(values_only=True)
        header = next(rows, None)
        if not header:
            raise ValueError("這份檔案是空的（連表頭都沒有）")

        idx: dict[str, int] = {}
        norm_header = {_norm(h): i for i, h in enumerate(header) if h}
        for field, cands in _COLUMN_CANDIDATES.items():
            for c in cands:
                if _norm(c) in norm_header:
                    idx[field] = norm_header[_norm(c)]
                    break
        if "api_id" not in idx:
            raise ValueError(
                f"找不到系統代碼欄。可接受的欄名："
                f"{'、'.join(_COLUMN_CANDIDATES['api_id'])}。"
                f"這份檔案的表頭是：{[h for h in header if h]}")

        seen, skipped = 0, 0
        for row in rows:
            if not any(v not in (None, "") for v in row):
                continue
            def get(f):
                i = idx.get(f)
                return row[i] if i is not None and i < len(row) else None
            api_id = str(get("api_id") or "").strip()
            if not api_id:
                skipped += 1        # 有內容但沒代碼——不能寫，也不能安靜吞掉
                continue
            upsert(conn, api_id, get("name"), get("ap_department"), get("ap_owner"))
            seen += 1
        conn.commit()
    finally:
        wb.close()

    return {
        "imported": seen,
        "skipped_no_api_id": skipped,
        "columns_found": sorted(idx),
        # 匯完立刻對帳：對得上多少、還差哪些代碼。只回「匯了 N 筆」等於沒回答問題。
        "coverage": coverage(conn),
    }
