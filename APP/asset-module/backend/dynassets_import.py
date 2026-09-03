"""dynassets 匯入：把「存活掃描 + CMDB」清單吃進資產清單，跟 CIA 登記對帳。

## 這份資料是什麼

另一個系統（Red Hat Satellite + nmap 存活探測 + CMDB）匯出的存活主機清單，
一列一台「網路上實際存活」的主機，含 IP/Hostname/MAC/OS/部門/用途/所在地/生命週期。
涵蓋實體 + 虛擬，補上 RVTools（只有 VM）補不到的實體機。

## 角色：對帳的「實況」那一半

CIA 登記表＝「我登記了什麼」；dynassets＝「網路上實際有什麼」。兩者對帳：
  dynassets 有、CIA 沒 → 漏登記（帳外資產）
  CIA 有、dynassets 沒 → 帳實不符（登記卻沒掃到；不代表不存在，可能關機/沒掃到）

## 三條原則（沿用既有地基，跟 rvtools_import 一致）

1. 不亂合併：每列走 identity.resolve，判不準進 merge_review，不自動合併錯。
2. 只覆蓋機器事實（ip/hostname/mac/os/is_vm），不碰人填的業務欄位。
3. 每列留來源紀錄（source='dynassets'，source_key=IP）。

## 跟 rvtools 的差別

dynassets 有 MAC（identity 的 MEDIUM 階，比 rvtools 多一個網路識別碼），解析會優先用它；
沒有 vm_uuid/hw_serial（STRONG），所以主要靠 mac + hostname/ip。新資產序號用 DYN- 前綴。
"""
from __future__ import annotations

import csv
import json
import re
import sqlite3
from pathlib import Path
from typing import Any

import openpyxl

# dynassets 表頭 → 我們的欄位。標題文字比對（正規化容忍大小寫/空白），不靠固定欄位位置。
_COLUMN_CANDIDATES: dict[str, tuple[str, ...]] = {
    "ip": ("IP",),
    "hostname": ("Hostname", "主機名稱"),
    "mac": ("MAC",),
    "os": ("作業系統", "OS 標準名稱", "OS Vendor"),
    "device_model": ("設備型號",),
    "department": ("部門",),
    "location": ("所在地",),
    "purpose": ("用途說明",),
    "lifecycle": ("Lifecycle Environment", "環境"),
    "os_eos": ("OS EOS Stage",),
    "ap": ("項目名稱", "AP"),
}

# 只覆蓋這些「機器事實」欄位；業務欄位（用途/保管者/環境/資產狀態…）不碰。
FACT_FIELDS = ("hostname", "ip", "os", "mac", "is_vm")

# 這些字串視為「沒有值」——dynassets 大量用 N/A 佔位。
_NULLISH = {"", "n/a", "na", "none", "null", "-"}


# 匯出範本用的顯示標題：優先選中文別名（這是中文語系的操作團隊，範本要讓「新
# 使用者不知道要填什麼」這件事真的被解決，不是塞一堆英文縮寫）；沒有中文別名的
# 欄位（IP／MAC 這種本來就是英文慣例、或 OS EOS Stage 只有英文別名）才退回第一個。
# 2026-08-25 使用者提出的通則：有匯入就要有配對的匯出範本。畫面上原本只寫「需含
# IP 欄」，但實際上還吃主機名／MAC／OS 這些欄位——查證時發現的落差，範本要把
# 認得的全部欄位列出來，不能讓人以為只填 IP 就夠。
def _prefer_chinese(candidates: tuple[str, ...]) -> str:
    for c in candidates:
        if any("一" <= ch <= "鿿" for ch in c):
            return c
    return candidates[0]


_TEMPLATE_HEADERS = {f: _prefer_chinese(c) for f, c in _COLUMN_CANDIDATES.items()}
_TEMPLATE_EXAMPLE = {
    "ip": "10.99.1.1", "hostname": "server01", "mac": "00:1A:2B:3C:4D:5E",
    "os": "Rocky Linux 9.7", "device_model": "Dell PowerEdge R740",
    "department": "資訊部", "location": "板橋機房", "purpose": "應用系統主機",
    "lifecycle": "Production", "os_eos": "", "ap": "資產編號範例A001",
}


def export_template_rows() -> tuple[list[str], list[str]]:
    """空白範本的表頭與一筆範例列——跟 segments.export_template_rows() 同一個理由。
    只有 IP 是必填（parse_dynassets 用它當主鍵），其餘都是可選但建議填，
    範例列刻意每欄都示範，讓人知道系統其實吃得到這麼多欄，不是只有 IP。"""
    fields = list(_COLUMN_CANDIDATES.keys())
    headers = [_TEMPLATE_HEADERS[f] for f in fields]
    example = [_TEMPLATE_EXAMPLE.get(f, "") for f in fields]
    return headers, example


def _clean(v: Any) -> Any:
    if v is None:
        return None
    s = str(v).strip()
    return None if s.casefold() in _NULLISH else s


def _norm(header: Any) -> str:
    if not isinstance(header, str):
        return ""
    text = re.sub(r"\s+", " ", header.replace("　", " ")).strip()
    return text.casefold()


def _read_rows(path: Path) -> tuple[list[str], list[list]]:
    """讀 dynassets（.xlsx / .xlsm / .csv），回 (表頭, 資料列)。"""
    if path.suffix.lower() in (".xlsx", ".xlsm"):
        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
        try:
            ws = wb[wb.sheetnames[0]]
            rows = [list(r) for r in ws.iter_rows(values_only=True)]
        finally:
            wb.close()
    else:
        with open(path, encoding="utf-8-sig", newline="") as f:
            rows = [r for r in csv.reader(f)]
    if not rows:
        return [], []
    header = [str(h) if h is not None else "" for h in rows[0]]
    return header, rows[1:]


def parse_dynassets(path: Path) -> list[dict[str, Any]]:
    """讀 dynassets 存活清單，一列一台轉一筆記錄。以 IP 為主鍵（沒 IP 的列跳過）。"""
    header, data = _read_rows(path)
    if not header:
        raise ValueError("dynassets 檔案是空的或讀不到表頭")

    norm_to_idx: dict[str, int] = {}
    for idx, h in enumerate(header):
        n = _norm(h)
        if n:
            norm_to_idx.setdefault(n, idx)
    if _norm("IP") not in norm_to_idx:
        raise ValueError("找不到 IP 欄——這不像 dynassets 存活清單")

    field_indices: dict[str, list[int]] = {}
    for field, candidates in _COLUMN_CANDIDATES.items():
        idxs = [norm_to_idx[_norm(c)] for c in candidates if _norm(c) in norm_to_idx]
        if idxs:
            field_indices[field] = idxs

    records = []
    for row in data:
        if not any(_clean(v) is not None for v in row):
            continue

        def get(field: str):
            for idx in field_indices.get(field, ()):
                if idx < len(row):
                    v = _clean(row[idx])
                    if v is not None:
                        return v
            return None

        ip = get("ip")
        if not ip:
            continue

        model = get("device_model")
        # 設備型號含 VM → 虛擬；有型號但不含 VM → 實體；判不出來留 None（不猜）
        is_vm = 1 if (model and "vm" in model.casefold()) else (0 if model else None)

        records.append({
            "ip": ip,
            "hostname": get("hostname"),
            "mac": get("mac"),
            "os": get("os"),
            "is_vm": is_vm,
            # 供 remark／asset_name 用（非機器事實，只在建新資產時參考）
            "department": get("department"),
            "location": get("location"),
            "purpose": get("purpose"),
            "lifecycle": get("lifecycle"),
            "os_eos": get("os_eos"),
            "device_model": model,
            "ap": get("ap"),
        })
    return records


def _make_remark(rec: dict) -> str:
    bits = []
    if rec.get("device_model"):
        bits.append(f"機型={rec['device_model']}")
    if rec.get("os_eos"):
        bits.append(f"OS生命週期={rec['os_eos']}")
    if rec.get("lifecycle"):
        bits.append(f"環境={rec['lifecycle']}")
    if rec.get("department"):
        bits.append(f"部門={rec['department']}")
    if rec.get("location"):
        bits.append(f"位置={rec['location']}")
    if rec.get("purpose"):
        bits.append(f"用途={rec['purpose']}")
    bits.append("來源=dynassets存活掃描")
    return "；".join(bits)


def _synth_serial(rec: dict) -> str:
    """新資產序號：DYN- 前綴標明來源。MAC（穩定）優先，退到 IP。"""
    return f"DYN-{rec.get('mac') or rec.get('ip')}"


def _stage_source_record(conn: sqlite3.Connection, rec: dict, match) -> int | None:
    """留來源紀錄。source='dynassets'、source_key=IP（這份的自然鍵）；
    UNIQUE(source, source_key) 讓同一 IP 重匯是更新同一筆、不累積。"""
    source_key = rec.get("ip")
    conn.execute(
        "INSERT INTO source_record (source, source_key, payload, resolved_status, "
        "resolved_hardware_id, resolved_rule, resolved_confidence) VALUES ('dynassets',?,?,?,?,?,?) "
        "ON CONFLICT(source, source_key) DO UPDATE SET payload=excluded.payload, "
        "resolved_status=excluded.resolved_status, resolved_hardware_id=excluded.resolved_hardware_id, "
        "resolved_rule=excluded.resolved_rule, resolved_confidence=excluded.resolved_confidence, "
        "collected_at=datetime('now','localtime')",
        (source_key, json.dumps(rec, ensure_ascii=False),
         match.status, match.hardware_id, match.rule, match.confidence),
    )
    row = conn.execute(
        "SELECT id FROM source_record WHERE source='dynassets' AND source_key=?", (source_key,)
    ).fetchone()
    return row["id"] if row else None


def import_dynassets(path: Path, conn: sqlite3.Connection) -> dict[str, Any]:
    """把 dynassets 存活清單吃進資產清單。
    matched→更新機器事實／new→建 DYN- 資產／ambiguous→進 merge_review 待審。"""
    import identity
    from db import _now_local

    records = parse_dynassets(path)
    inserted = updated = pending = 0
    errors: list[str] = []

    for rec in records:
        try:
            match = identity.resolve(conn, rec)
            sr_id = _stage_source_record(conn, rec, match)

            if match.status == identity.MATCHED:
                existing = conn.execute(
                    "SELECT asset_serial FROM hardware WHERE id = ?", (match.hardware_id,)
                ).fetchone()
                if existing is None:
                    errors.append(f"{rec.get('hostname') or rec['ip']}：對到的資產已不存在，略過")
                    continue
                setters = {k: rec[k] for k in FACT_FIELDS if rec.get(k) not in (None, "")}
                if setters:
                    assigns = ", ".join(f"{k} = ?" for k in setters)
                    conn.execute(
                        f"UPDATE hardware SET {assigns}, updated_at = ? WHERE id = ?",
                        (*setters.values(), _now_local(), match.hardware_id),
                    )
                updated += 1

            elif match.status == identity.NEW:
                serial = _synth_serial(rec)
                exists = conn.execute(
                    "SELECT id FROM hardware WHERE asset_serial = ?", (serial,)).fetchone()
                fields = {k: rec[k] for k in FACT_FIELDS if rec.get(k) not in (None, "")}
                if exists:
                    if fields:
                        assigns = ", ".join(f"{k} = ?" for k in fields)
                        conn.execute(
                            f"UPDATE hardware SET {assigns}, updated_at = ? WHERE id = ?",
                            (*fields.values(), _now_local(), exists["id"]),
                        )
                    updated += 1
                    continue
                fields.update({
                    "asset_serial": serial,
                    "asset_name": rec.get("ap") or rec.get("hostname"),
                    "remark": _make_remark(rec),
                })
                cols = ", ".join(fields.keys())
                ph = ", ".join("?" for _ in fields)
                cur = conn.execute(
                    f"INSERT INTO hardware ({cols}) VALUES ({ph})", tuple(fields.values()))
                if sr_id:
                    conn.execute(
                        "UPDATE source_record SET resolved_hardware_id = ? WHERE id = ?",
                        (cur.lastrowid, sr_id))
                inserted += 1

            else:  # AMBIGUOUS — 不自動合併，進人工佇列
                conn.execute(
                    "INSERT INTO merge_review (source_record_id, reason, candidates, status) "
                    "VALUES (?,?,?, 'open')",
                    (sr_id, match.reason, json.dumps(match.candidates, ensure_ascii=False)))
                pending += 1

        except sqlite3.IntegrityError as exc:
            errors.append(f"{rec.get('hostname') or rec['ip']}：寫入失敗（{exc}）")

    conn.commit()
    return {
        "source": "dynassets",
        "total": len(records),
        "inserted": inserted,
        "updated": updated,
        "pending_review": pending,
        "errors": errors,
    }


if __name__ == "__main__":
    import sys
    from db import get_connection, init_db

    if len(sys.argv) < 2:
        print("用法：python dynassets_import.py <dynassets.csv|.xlsx>")
        raise SystemExit(2)
    init_db()
    connection = get_connection()
    try:
        print(json.dumps(import_dynassets(Path(sys.argv[1]), connection),
                         ensure_ascii=False, indent=2))
    finally:
        connection.close()
