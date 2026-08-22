"""S19 VC 採集器：把 RVTools 匯出的 vCenter 盤點吃進資產清單。

## 為什麼走 RVTools 而不是先接 vCenter API

RVTools 是業界收 vCenter 盤點最常見的工具：連上 vCenter 按一下就匯出一份 Excel，
`vInfo` 分頁一列一台 VM，欄位版本間幾乎沒變（公開且穩定）。而我們系統本來就吃 Excel，
所以「先用 RVTools 拿到真實盤點」是最短路徑。等這條走順，再往「直接連 vCenter API、
每晚自己收」推（那才是不用人點的版本）——同一套身分解析與寫入，不會白做。

## 三條原則（沿用既有地基）

1. **不亂合併**：每一列都走 identity.resolve（強識別碼 vm_uuid 相符才定案）。判不準的
   一律進 merge_review 交人工，**寧留兩筆也不自動合併錯**（合併錯不噴錯、很難救）。
2. **只覆蓋機器事實、不碰人填的業務欄位**：vCenter 對 VM 的主機名/IP/OS/UUID 是權威，
   但資產用途、保管者那些是人維護的，收集不覆蓋（沿用 manage_state 的 FACT_FIELDS 精神）。
3. **每一列留來源紀錄**（source_record，source='vcenter'）：判錯時能回溯是哪一步、哪條規則。

## VM 是真實資料，不是假資料

RVTools 匯出的是 vCenter 裡真的 VM——這正是我們要盤的東西，跟「demo 假資料」是兩回事。
新 VM 會建成資產，序號用 `VC-<vm_uuid>` 前綴標明來源（像納管的 AUTO-），一眼看得出這台
是從 vCenter 收進來的、還沒對到公司資產序號。
"""
from __future__ import annotations

import datetime as _dt
import json
import re
import sqlite3
from decimal import Decimal
from pathlib import Path
from typing import Any

import openpyxl


def export_time_from_filename(name: str | None) -> str | None:
    """從 RVTools 的檔名取出「這份是哪天從 vCenter 匯出的」。

    RVTools 預設檔名帶匯出時間，例如：
        BQ_10.99.169.191_RVTools_export_all_2026-07-30_10.34.59.xlsx
                                            ^^^^^^^^^^ ^^^^^^^^

    為什麼非要這個不可：`imported_at`（匯進系統的時間）跟匯出時間是兩件事。
    2026-08-20 匯進來的五個檔全部是 07-30 匯出的，中間隔了三週——拿三週前的快照
    算爆炸半徑，會漏掉搬過來的 VM、也會多算搬走的。畫面只寫「最後匯入 8/20」
    會讓人以為資料是新的。

    ⚠️ 取不到就回 None，**不要拿檔案時間或匯入時間頂替**。「不知道這份多舊」
    跟「這份是今天的」是完全不同的兩句話，猜一個看起來合理的值等於製造假證據。
    vSource 分頁裡沒有匯出時間（只有 vCenter 版本），所以檔名是唯一來源。
    """
    if not name:
        return None
    m = re.search(r"(\d{4}-\d{2}-\d{2})[_ ](\d{2})[.\-:](\d{2})[.\-:](\d{2})", str(name))
    if m:
        return f"{m.group(1)} {m.group(2)}:{m.group(3)}:{m.group(4)}"
    m = re.search(r"(\d{4}-\d{2}-\d{2})", str(name))
    return m.group(1) if m else None


def _jsonable(v: Any) -> Any:
    """把 Excel 儲存格的值轉成 json.dumps 吞得下的型別。

    2026-08-20 使用者實測五個檔全部炸在這裡：
    `匯入失敗，請確認是 RVTools 匯出的檔：Object of type datetime is not JSON serializable`
    ——而那五個就是 RVTools 匯出的檔，訊息在怪使用者，實際上是我們自己的 bug。

    原因：額外分頁（vHost/vSnapshot/vInfo…）是「整列原樣存進 source_record」，
    而 RVTools 有一堆日期時間欄（快照建立時間、開機時間、憑證到期日…），
    openpyxl 會把它們讀成 datetime 物件，json.dumps 直接拋例外。
    因為是**整份檔案共用一次 json.dumps**，只要任何一列有任何一個日期欄，
    整個檔案就全軍覆沒——這就是為什麼五個檔一個都進不來。

    轉成 ISO 字串而不是丟掉：payload 的用途是「這份匯出還告訴我們什麼」，
    快照建立時間正是之後要做快照稽核時最需要的欄位，丟了就白存了。
    """
    if isinstance(v, _dt.datetime):
        return v.isoformat(sep=" ")
    if isinstance(v, (_dt.date, _dt.time)):
        return v.isoformat()
    if isinstance(v, _dt.timedelta):
        return str(v)
    if isinstance(v, Decimal):
        return float(v)
    if isinstance(v, bytes):
        return v.decode("utf-8", "replace")
    return v

VINFO_SHEET_CANDIDATES = ("vInfo", "vinfo", "tabvInfo")

# RVTools vInfo 標準表頭 → 我們的欄位。用「正規化後比對」容忍大小寫/空白/版本差異。
# 同一個目標欄位可有多個來源候選，依序取第一個「有值」的（真值優先於次選）。
# 這些是 RVTools 各版本都在的欄位，刻意只挑「盤點真的需要」的，不貪多。
_COLUMN_CANDIDATES: dict[str, tuple[str, ...]] = {
    # 最強識別碼：VM 的 BIOS/SMBIOS UUID，換 IP 換名都不變（identity.py 的最強階）。
    # ⚠️ 不可用 "VI SDK UUID"——那是 vCenter 伺服器的 UUID，不是這台 VM 的。
    "vm_uuid": ("VM UUID", "SMBIOS UUID", "BIOS UUID"),
    # 主機名：優先 guest 的 DNS Name（真主機名），退回 VM 顯示名稱
    "hostname": ("DNS Name", "VM"),
    # IP：優先 Primary IP Address，退回 IP Address
    "ip": ("Primary IP Address", "IP Address"),
    # OS：優先 VMware Tools 回報的（實際跑的），退回設定檔宣告的
    "os": ("OS according to the VMware Tools", "OS according to the configuration file",
           "OS according to the VMware tools"),
    # VM 顯示名稱（進 asset_name，也當 hostname 的退路）
    "vm_name": ("VM",),
    # 這台 VM 跑在哪台實體 ESXi 主機上（進 remark，沒有對應欄位）
    "esxi_host": ("Host",),
    # 開機狀態（進 remark；不映射到 asset_status——那是業務生命週期、人維護的另一條軸）
    "powerstate": ("Powerstate", "Power State", "powerstate"),
    # vCenter 內部 moref（vm-123），當 source_key 退路（vm_uuid 缺時）
    "moref": ("VM ID", "MoRef", "Object ID"),
    # MICS 切片1：這台 VM 所在的 cluster（vInfo 本來就有獨立欄位，直接對應）
    "cluster": ("Cluster",),
    # MICS 切片1：資料存放在哪個 datastore——vInfo 沒有獨立的 Datastore 欄位，
    # 是包在 VMX 檔案的 Path 欄裡（格式固定：中括號包住 datastore 名稱，例如
    # "[PROD_B_vSan_Datastore] uuid/VMNAME.vmx"），存原始 Path，datastore 名稱
    # 在 parse_rvtools() 用正則從裡面拆出來，不是直接欄位比對。
    "datastore_path": ("Path",),
}

# 只覆蓋這些「機器事實」欄位，不碰人填的業務欄位（用途/保管者/環境/資產狀態…）。
FACT_FIELDS = ("hostname", "ip", "os", "vm_uuid", "is_vm")

# 2026-08-19 使用者拍板「要吃全部」：RVTools 匯出通常不只 vInfo 一頁，還有主機層
# （vHost：ESXi版本/CPU/記憶體總量）、儲存層（vDatastore：容量/剩餘空間）、
# 快照層（vSnapshot：誰有沒清的舊快照）、VM 細項（vCPU/vMemory/vDisk/vPartition/
# vNetwork/vTools）等十幾頁，先前全部被忽略——上傳了也沒進系統。
#
# 這些分頁沒有像 vInfo 那樣對到 hardware 表的固定欄位（vInfo 那組是機器身分
# ident 用的核心欄位），所以不做 identity.resolve/hardware UPDATE 那一套：整列
# 原樣存進 source_record（source='vcenter_extra:<分頁名>'）當「這份匯出還告訴
# 我們什麼」的紀錄，供之後的功能（容量儀表板、快照稽核、ESXi硬體規格…）查詢，
# 不會憑空消失——資料先進來，用途可以之後再長。
EXTRA_SHEETS = (
    "vHost", "vCluster", "vDatastore", "vCPU", "vMemory", "vDisk", "vPartition",
    "vNetwork", "vSnapshot", "vTools", "vSource", "vNIC", "vSwitch", "vPort",
    "vRP", "vHBA", "vMultiPath", "vLicense", "vFolder", "dvSwitch", "dvPort", "vHealth",
)


def _norm(header: Any) -> str:
    """表頭比對用正規化：全形轉半形、collapse 空白、小寫。只用於比對，不改原值。"""
    if not isinstance(header, str):
        return ""
    text = header.replace("　", " ")
    text = re.sub(r"\s+", " ", text).strip()
    return text.casefold()


def _find_vinfo_sheet(wb) -> Any:
    for name in wb.sheetnames:
        if _norm(name) in {_norm(c) for c in VINFO_SHEET_CANDIDATES}:
            return wb[name]
    # 有些匯出把 VM 清單放在唯一/第一個分頁——退而求其次用第一個含 "VM UUID" 表頭的分頁
    for ws in wb.worksheets:
        header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ())
        norms = {_norm(h) for h in header}
        if _norm("VM UUID") in norms or _norm("VM") in norms:
            return ws
    return None


def parse_rvtools(xlsx_path: Path) -> list[dict[str, Any]]:
    """讀 RVTools 匯出的 vInfo 分頁，一列 VM 轉一筆記錄。

    回傳的每筆含：vm_uuid/hostname/ip/os/is_vm，以及供 remark 與 source_key 用的
    vm_name/esxi_host/powerstate/moref。欄位用標題文字比對，不靠固定欄位位置——
    RVTools 換版本或欄位順序不會壞。
    """
    wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
    try:
        ws = _find_vinfo_sheet(wb)
        if ws is None:
            raise ValueError("找不到 RVTools 的 vInfo 分頁（也沒有含 VM/VM UUID 表頭的分頁）")

        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ())
        # 正規化表頭 → 欄索引
        norm_to_idx: dict[str, int] = {}
        for idx, h in enumerate(header_row):
            n = _norm(h)
            if n:
                norm_to_idx.setdefault(n, idx)

        # 目標欄位 → 所有存在的候選欄索引（依候選順序）。
        # ⚠️ 逐「列」取值而非只選一欄：RVTools 可能同時有 Primary IP Address 與 IP Address
        # 兩欄，但某台 VM 的 Primary 為空、IP Address 才有值——只鎖第一個存在的欄會漏掉。
        field_indices: dict[str, list[int]] = {}
        for field, candidates in _COLUMN_CANDIDATES.items():
            idxs = [norm_to_idx[_norm(c)] for c in candidates if _norm(c) in norm_to_idx]
            if idxs:
                field_indices[field] = idxs

        records = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not any(v not in (None, "") for v in row):
                continue  # 全空列跳過

            def get(field: str):
                for idx in field_indices.get(field, ()):
                    if idx >= len(row):
                        continue
                    v = row[idx]
                    if isinstance(v, str):
                        v = v.strip()
                    if v not in (None, ""):
                        # vInfo 這幾欄正常都是字串，但 rec 一樣會被 json.dumps
                        # 存進 source_record，同樣不能讓 datetime 漏進去。
                        return _jsonable(v)      # 候選欄裡第一個有值的
                return None

            hostname = get("hostname")
            datastore_path = get("datastore_path")
            m = re.match(r"^\[([^\]]+)\]", str(datastore_path)) if datastore_path else None
            rec = {
                "vm_uuid": get("vm_uuid"),
                "hostname": hostname,
                "ip": get("ip"),
                "os": get("os"),
                "is_vm": 1,
                "vm_name": get("vm_name"),
                "esxi_host": get("esxi_host"),
                "powerstate": get("powerstate"),
                "moref": get("moref"),
                "cluster": get("cluster"),
                "datastore": m.group(1) if m else None,
            }
            # 整列連個名字都沒有就不算一台
            if not (rec["vm_uuid"] or rec["hostname"] or rec["vm_name"]):
                continue
            records.append(rec)
        return records
    finally:
        wb.close()


def parse_extra_sheets(xlsx_path: Path) -> dict[str, list[dict]]:
    """讀 vInfo 以外、EXTRA_SHEETS 裡有出現的分頁，一列轉一個 dict（欄名當key，原始值，
    不做欄位對應/型別轉換——這些分頁欄位多、版本間差異也大，不像 vInfo 那樣值得手刻
    對應表；先整列存下來，之後哪個欄位真的要拿來用，再從 payload 裡挑。
    """
    wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
    try:
        wanted = {_norm(s): s for s in EXTRA_SHEETS}
        result: dict[str, list[dict]] = {}
        for name in wb.sheetnames:
            target = wanted.get(_norm(name))
            if not target:
                continue
            ws = wb[name]
            header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ())
            headers = [str(h).strip() if h not in (None, "") else f"col{i}" for i, h in enumerate(header_row)]
            rows: list[dict] = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not any(v not in (None, "") for v in row):
                    continue
                rec: dict[str, Any] = {}
                for i, h in enumerate(headers):
                    if i >= len(row):
                        continue
                    v = row[i]
                    if isinstance(v, str):
                        v = v.strip()
                    if v not in (None, ""):
                        rec[h] = _jsonable(v)
                if rec:
                    rows.append(rec)
            if rows:
                result[target] = rows
        return result
    finally:
        wb.close()


def _stage_extra_sheet_records(conn: sqlite3.Connection, sheet: str, rows: list[dict]) -> int:
    """存進 source_record，source=`vcenter_extra:<分頁名>`。

    這些分頁常常一個實體對多列（一台VM好幾顆硬碟/好幾張快照），不像 vInfo 一列
    一台 VM 那樣天然唯一，所以 source_key 用「自然鍵 + 列序號」——序號防同鍵多列
    互相覆蓋，代價是同一份檔案重複匯入時舊列不會被真的去重，但至少資料不會遺失
    或互相蓋掉，這比「假裝能完美去重結果蓋掉不該蓋的列」安全。
    """
    for i, rec in enumerate(rows):
        key = (
            rec.get("VM UUID") or rec.get("Host") or rec.get("Host Name")
            or rec.get("Name") or rec.get("Object ID") or str(i)
        )
        source_key = f"{key}#{i}"
        payload = json.dumps(rec, ensure_ascii=False)
        conn.execute(
            "INSERT INTO source_record (source, source_key, payload, resolved_status) "
            "VALUES (?,?,?, 'extra') "
            "ON CONFLICT(source, source_key) DO UPDATE SET payload=excluded.payload, "
            "collected_at=datetime('now','localtime')",
            (f"vcenter_extra:{sheet}", source_key, payload),
        )
    return len(rows)


def _make_remark(rec: dict) -> str | None:
    """把沒有對應欄位、但盤點有用的資訊收進 remark（實體主機、開機狀態）。"""
    bits = []
    if rec.get("esxi_host"):
        bits.append(f"vHost={rec['esxi_host']}")
    if rec.get("powerstate"):
        bits.append(f"powerState={rec['powerstate']}")
    bits.append("來源=vCenter/RVTools")
    return "；".join(bits)


def _synth_serial(rec: dict) -> str:
    """新 VM 的資產序號：VC- 前綴標明來源（像納管的 AUTO-）。

    優先用 vm_uuid（穩定、換 IP 換名都不變），沒有才退到顯示名稱——退路會比較弱，
    但至少能建起來讓人看到、之後補真序號。"""
    base = rec.get("vm_uuid") or rec.get("moref") or rec.get("vm_name") or rec.get("hostname")
    return f"VC-{base}"


def _stage_source_record(conn: sqlite3.Connection, rec: dict, match) -> int:
    """留一筆來源紀錄（source='vcenter'）。判錯時能回溯是哪一步、哪條規則。

    source_key 用 vm_uuid（沒有退到 moref／顯示名），UNIQUE(source, source_key) 讓
    同一台 VM 重複匯入是更新同一筆、不會累積。"""
    source_key = rec.get("vm_uuid") or rec.get("moref") or rec.get("vm_name") or rec.get("hostname")
    payload = json.dumps(rec, ensure_ascii=False)
    cur = conn.execute(
        "INSERT INTO source_record (source, source_key, payload, resolved_status, "
        "resolved_hardware_id, resolved_rule, resolved_confidence) VALUES ('vcenter',?,?,?,?,?,?) "
        "ON CONFLICT(source, source_key) DO UPDATE SET payload=excluded.payload, "
        "resolved_status=excluded.resolved_status, resolved_hardware_id=excluded.resolved_hardware_id, "
        "resolved_rule=excluded.resolved_rule, resolved_confidence=excluded.resolved_confidence, "
        "collected_at=datetime('now','localtime')",
        (source_key, payload, match.status, match.hardware_id, match.rule, match.confidence),
    )
    # ON CONFLICT 更新時 lastrowid 不可靠，回查一次拿 id
    row = conn.execute(
        "SELECT id FROM source_record WHERE source='vcenter' AND source_key=?", (source_key,)
    ).fetchone()
    return row["id"] if row else cur.lastrowid


def import_rvtools(xlsx_path: Path, conn: sqlite3.Connection) -> dict[str, Any]:
    """把一份 RVTools 匯出吃進資產清單。回統計＋逐項結果。

    每列走 identity.resolve：
      matched   → 更新既有資產的機器事實欄位（不碰業務欄位）
      new       → 建一筆 VC- 前綴的新資產（is_vm=1）
      ambiguous → 進 merge_review 交人工，**不寫 hardware**（不自動合併錯）
    """
    import identity
    from db import _now_local

    records = parse_rvtools(xlsx_path)
    inserted = updated = pending = 0
    errors: list[str] = []

    for rec in records:
        try:
            match = identity.resolve(conn, rec)
            sr_id = _stage_source_record(conn, rec, match)

            if match.status == identity.MATCHED:
                existing = conn.execute(
                    "SELECT asset_serial FROM hardware WHERE id = ?", (match.hardware_id,)
                ).fetchone()
                if existing is None:
                    errors.append(f"{rec.get('vm_name') or rec.get('hostname')}：對到的資產已不存在，略過")
                    continue
                setters = {k: rec[k] for k in FACT_FIELDS if rec.get(k) not in (None, "")}
                if setters:
                    assigns = ", ".join(f"{k} = ?" for k in setters)
                    conn.execute(
                        f"UPDATE hardware SET {assigns}, updated_at = ? WHERE id = ?",
                        (*setters.values(), _now_local(), match.hardware_id),
                    )
                updated += 1

            elif match.status == identity.NEW:
                serial = _synth_serial(rec)
                # 序號可能已存在（同一份重匯、或退路序號撞名）→ 當更新處理，不硬塞出 IntegrityError
                exists = conn.execute(
                    "SELECT id FROM hardware WHERE asset_serial = ?", (serial,)).fetchone()
                fields = {k: rec[k] for k in FACT_FIELDS if rec.get(k) not in (None, "")}
                if exists:
                    if fields:
                        assigns = ", ".join(f"{k} = ?" for k in fields)
                        conn.execute(
                            f"UPDATE hardware SET {assigns}, updated_at = ? WHERE id = ?",
                            (*fields.values(), _now_local(), exists["id"]),
                        )
                    updated += 1
                    continue
                fields.update({
                    "asset_serial": serial,
                    "asset_name": rec.get("vm_name"),
                    "is_vm": 1,
                    "environment": "正式",       # 預設，人可改；不猜業務欄位
                    "remark": _make_remark(rec),
                })
                cols = ", ".join(fields.keys())
                ph = ", ".join("?" for _ in fields)
                cur = conn.execute(
                    f"INSERT INTO hardware ({cols}) VALUES ({ph})", tuple(fields.values()))
                conn.execute(
                    "UPDATE source_record SET resolved_hardware_id = ? WHERE id = ?",
                    (cur.lastrowid, sr_id))
                inserted += 1

            else:  # AMBIGUOUS —— 不自動合併，進人工佇列
                conn.execute(
                    "INSERT INTO merge_review (source_record_id, reason, candidates, status) "
                    "VALUES (?,?,?, 'open')",
                    (sr_id, match.reason, json.dumps(match.candidates, ensure_ascii=False)))
                pending += 1

        except sqlite3.IntegrityError as exc:
            errors.append(f"{rec.get('vm_name') or rec.get('hostname')}：寫入失敗（{exc}）")

    # vInfo 以外的分頁（vHost/vDatastore/vSnapshot…）——2026-08-19 使用者拍板全部收進來，
    # 不只吃 VM 清單那一頁。存進 source_record，不動 hardware（見 EXTRA_SHEETS 說明）。
    extra_sheets = parse_extra_sheets(xlsx_path)
    extra_sheet_counts = {
        sheet: _stage_extra_sheet_records(conn, sheet, rows)
        for sheet, rows in extra_sheets.items()
    }

    conn.commit()
    return {
        "source": "vcenter/rvtools",
        "total_vms": len(records),
        "inserted": inserted,
        "updated": updated,
        "pending_review": pending,   # 判不準、等人工決定的
        "errors": errors,
        "extra_sheets": extra_sheet_counts,  # {分頁名: 列數}，vInfo以外收到的其他分頁
    }


if __name__ == "__main__":
    import sys

    from db import get_connection, init_db

    if len(sys.argv) < 2:
        print("用法：python rvtools_import.py <RVTools匯出.xlsx>")
        raise SystemExit(2)

    init_db()
    connection = get_connection()
    try:
        result = import_rvtools(Path(sys.argv[1]), connection)
        print(json.dumps(result, ensure_ascii=False, indent=2))
    finally:
        connection.close()
