"""S5：後端 REST API 層（D29 API優先設計）。

只做資料存取／整合既有模組（db／scanner／comparison_engine），不重複實作比對或掃描邏輯——
這層單純把既有函式包成 HTTP 介面，供 Nuxt3 前端（S7-S9）與之後其他模組重用（D29 note：
系統拓撲模組可直接呼叫這裡的 API 重用資料，不用重新收集）。

sort_by／order 一律經白名單檢查才拼進 SQL ORDER BY，避免使用者輸入直接串進查詢字串。
"""
from __future__ import annotations

import csv
import io
import ipaddress
import json
import os
import re
import socket
import sqlite3
import subprocess
import tempfile
from datetime import datetime
from pathlib import Path
from typing import Iterator

from fastapi import (
    Cookie, Depends, FastAPI, File, Form, Header, HTTPException, Query, Response, UploadFile,
)
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, StreamingResponse
from openpyxl import Workbook, load_workbook
from pydantic import BaseModel

import agent_auth
import auth
# 收集器身分（金鑰與自己的位址）集中在這裡解析——五個端點都要用，
# 各自 import 過一次容易漏掉，漏掉的那個就會走回舊的寫死預設值。
import onboard_engine
import backup
import cmdb_gateway
import golive
import scan_service
from db import (
    create_connection_record,
    get_db_path,
    create_import_log,
    delete_connection_record,
    get_connection,
    get_connection_by_id,
    get_feature_flag,
    get_latest_import_log,
    insert_hardware,
    list_connections,
    list_import_log,
    list_feature_flags,
    mark_comparison_read,
    set_connection_enabled,
    set_feature_flag,
    update_connection_record,
    update_connection_status,
    update_hardware,
)
from excel_import import MAPPING_PATH, SHEET_CONFIG, import_excel, load_mapping

app = FastAPI(title="資產盤點模組 API")

# 本行程的啟動時間：只有真的重啟過才會變，是「新程式碼有沒有生效」最可信的證據。
# （version.json / build_info.json 都是即時讀檔，服務沒重啟時照樣顯示新版，會騙人。）
_PROCESS_STARTED_AT = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

# 開發環境前端跑在 3000 埠（Nuxt3預設），可用環境變數 ASSET_API_CORS_ORIGINS 覆蓋（部署到
# 內網IP時就是靠它指定實際來源）。
#
# 預設同時放行 localhost 與 127.0.0.1 兩種寫法：這兩個在瀏覽器眼中是不同的 origin，
# 少放一個就會在開發時撞到 CORS 錯誤，而畫面只會顯示「載入失敗」，很難聯想到是位址寫法問題。
# （本機實測：這台 localhost 先解析到 IPv6 ::1，開發時改用 127.0.0.1 才連得到 dev server，
#  結果就踩到只放行 localhost 的 CORS。）
_DEFAULT_DEV_ORIGINS = "http://localhost:3000,http://127.0.0.1:3000"
_allowed_origins = [
    origin.strip()
    for origin in os.environ.get("ASSET_API_CORS_ORIGINS", _DEFAULT_DEV_ORIGINS).split(",")
    if origin.strip()
]
app.add_middleware(
    CORSMiddleware,
    allow_origins=_allowed_origins,
    allow_credentials=True,  # 帶cookie（session_token）必須開這個，前端fetch才要用credentials:"include"
    allow_methods=["*"],
    allow_headers=["*"],
)

FIELD_GROUPS_PATH = Path(__file__).parent / "field_groups.json"

ALLOWED_ASSET_SORT = {
    "hostname", "ip", "device_model", "rack_no", "group_name", "api_id",
    "asset_purpose", "custodian", "usage_unit", "asset_status", "environment",
    "asset_serial",
}
ALLOWED_ISSUE_SORT = {"detected_at", "hostname", "issue_type", "is_read"}
ALLOWED_PERSONNEL_SORT = {
    "person_name", "phone", "job_desc", "belong_division", "belong_department",
    "asset_serial",
}
ALLOWED_SOFTWARE_SORT = {
    "asset_name", "hostname", "ip", "db_software", "backup_frequency", "cloud",
    "custodian", "asset_serial",
}


def get_db() -> Iterator[sqlite3.Connection]:
    conn = get_connection()
    try:
        yield conn
    finally:
        conn.close()


def require_auth(
    session_token: str | None = Cookie(default=None),
    conn: sqlite3.Connection = Depends(get_db),
) -> sqlite3.Row:
    """要保護一支端點，在路由函式簽章加上 Depends(require_auth) 即可。

    現況：除了 login／logout／version 這三支（見 test_auth_coverage.PUBLIC 白名單），
    所有 /api 端點都掛了這個依賴。

    歷史教訓：S6 當初只用它保護 /api/auth/me，其餘端點「不在這個切片範圍」而留待後續，
    結果一直沒補——前端每頁都有登入守衛，看起來很安全，但 API 全裸長達數個切片，
    任何碰得到後端埠的人都能撈走全部資產與人員姓名電話，或 PUT 改欄位對應。
    現已全數補上，並由 tests/asset_module/test_auth_coverage.py 自動涵蓋新端點防止再退化。
    新增路由請預設掛上；真要公開必須明確加進該白名單。
    """
    session = auth.resolve_session(conn, session_token)
    if session is None:
        raise HTTPException(401, "未登入或登入已過期")
    return session


def require_host_key(
    x_agent_key: str | None = Header(default=None),
    conn: sqlite3.Connection = Depends(get_db),
) -> str:
    """Push agent 端點專用的驗證——跟 require_auth（人的 session cookie）是不同層級，
    這裡驗的是「機器」用 header 帶的一次性核發 key。回傳驗證出的 asset_serial，
    路由函式一定要用這個回傳值，不能信任 request body 裡任何自稱的主機識別欄位
    （否則 agent 就能冒充寫別台的資料，見 agent_auth.resolve_host_key 的說明）。
    """
    asset_serial = agent_auth.resolve_host_key(conn, x_agent_key)
    if asset_serial is None:
        raise HTTPException(401, "agent key 缺失、無效或已撤銷")
    return asset_serial


class MarkReadBody(BaseModel):
    is_read: bool = True


class LoginBody(BaseModel):
    username: str
    password: str


@app.post("/api/auth/login")
def login(body: LoginBody, response: Response, conn: sqlite3.Connection = Depends(get_db)):
    user = auth.authenticate(conn, body.username, body.password)
    if user is None:
        # 帳號不存在／密碼錯誤都回同一句話，不透露「帳號不存在」讓人拿去枚舉帳號
        raise HTTPException(401, "帳號或密碼錯誤")
    token, _expires_at = auth.issue_session(conn, user["id"])
    response.set_cookie(
        key=auth.SESSION_COOKIE_NAME,
        value=token,
        httponly=True,
        samesite="lax",
        secure=False,  # 部署目標是內網IP走http（D11/D18已定案的內網簡化精神），改https要記得打開
        max_age=int(auth.SESSION_TTL.total_seconds()),
    )
    return {"username": user["username"]}


@app.post("/api/auth/logout")
def logout(
    response: Response,
    session_token: str | None = Cookie(default=None),
    conn: sqlite3.Connection = Depends(get_db),
):
    if session_token:
        from db import delete_session

        delete_session(conn, session_token)
    response.delete_cookie(auth.SESSION_COOKIE_NAME)
    return {"ok": True}


@app.get("/api/auth/me")
def me(session: sqlite3.Row = Depends(require_auth)):
    return {"username": session["username"]}


# ===== CIA(登記) × 掃描 的比對：儀表板磚塊與各個下鑽清單共用同一套判定 =====
# 儀表板每個數字都可以點進去看細項，所以「數字」跟「清單」必須用同一段邏輯算，
# 否則點進去筆數對不上，比不能點還糟。

def _latest_scan_time(conn: sqlite3.Connection) -> str | None:
    row = conn.execute("SELECT MAX(scan_time) AS t FROM scan_history").fetchone()
    return row["t"] if row else None


def _scanned_alive_rows(conn: sqlite3.Connection, scan_time: str | None) -> list[sqlite3.Row]:
    """最新一次掃描中「掃得到（存活）」的主機。scan_ok=0 是網段掃描失敗，不是主機死掉。"""
    if not scan_time:
        return []
    return conn.execute(
        "SELECT * FROM scan_history WHERE scan_time = ? AND scan_ok = 1", (scan_time,)
    ).fetchall()


def _scan_keys(scanned_rows: list[sqlite3.Row]) -> tuple[set[str], set[str]]:
    return (
        {r["ip"] for r in scanned_rows if r["ip"]},
        {r["hostname"] for r in scanned_rows if r["hostname"]},
    )


def _row_in_keys(row, ips: set[str], hostnames: set[str]) -> bool:
    """IP 或主機名稱任一對得上就算同一台（沿用 /api/scan/unregistered 既有的判定方式）。"""
    return bool(
        (row["ip"] and row["ip"] in ips) or (row["hostname"] and row["hostname"] in hostnames)
    )


def _check_sort(sort_by: str, order: str, allowed: set[str]) -> None:
    if sort_by not in allowed:
        raise HTTPException(400, f"不支援的排序欄位：{sort_by}")
    if order not in ("asc", "desc"):
        raise HTTPException(400, "order 只接受 asc 或 desc")


_COLUMN_CACHE: dict[str, set[str]] = {}


def _sortable_columns(conn: sqlite3.Connection, table: str) -> set[str]:
    """回傳該資料表**實際存在的欄位名**當排序白名單。

    為什麼不寫死一份清單：使用者的鐵則是「表格每一欄都要能排」，而畫面上的欄位
    是由 field_groups/field_meta 動態決定的（含數十個進階欄位）。寫死的白名單只涵蓋
    12 欄，點到其他欄就回 400、畫面顯示「資產資料載入失敗」——比不能排更糟。

    安全性不變：來源是 PRAGMA 的真實欄位名，使用者輸入只能「命中或不命中」，
    永遠不會有自由字串被拼進 ORDER BY。新增欄位也自動涵蓋，不用記得回來改。
    """
    if table not in _COLUMN_CACHE:
        _COLUMN_CACHE[table] = {
            r[1] for r in conn.execute(f"PRAGMA table_info({table})")
        }
    return _COLUMN_CACHE[table]


# D25：儀表板環境篩選預設「正式」，可切換包含測試/備援。HTML Mock v4三段式選單，
# key用ASCII"+"避免query string編碼問題，None代表不篩選（=全部含備援）。
ENV_FILTER_PRESETS: dict[str, list[str] | None] = {
    "正式": ["正式"],
    "正式+測試": ["正式", "測試"],
    "全部": None,
}


@app.get("/api/dashboard/stats")
def dashboard_stats(
    environment: str = "正式",
    session: sqlite3.Row = Depends(require_auth),
    conn: sqlite3.Connection = Depends(get_db),
):
    """D25：環境篩選預設「正式」。重疊圖三區塊：僅CIA登記／重疊／僅掃描到（=漏登記候選）。"""
    if environment not in ENV_FILTER_PRESETS:
        raise HTTPException(400, f"不支援的環境篩選：{environment}")
    env_list = ENV_FILTER_PRESETS[environment]

    if env_list is None:
        ica_rows = conn.execute("SELECT ip, hostname FROM hardware").fetchall()
    else:
        placeholders = ",".join("?" for _ in env_list)
        ica_rows = conn.execute(
            f"SELECT ip, hostname FROM hardware WHERE environment IN ({placeholders})",
            env_list,
        ).fetchall()
    ica_ips = {r["ip"] for r in ica_rows if r["ip"]}
    ica_hostnames = {r["hostname"] for r in ica_rows if r["hostname"]}

    last_scan_time = _latest_scan_time(conn)
    scanned_rows = _scanned_alive_rows(conn, last_scan_time)
    failed_segments: list[str] = []
    if last_scan_time:
        failed_segments = [
            r["segment"]
            for r in conn.execute(
                "SELECT DISTINCT segment FROM scan_history WHERE scan_time = ? AND scan_ok = 0",
                (last_scan_time,),
            ).fetchall()
        ]

    scan_ips, scan_hostnames = _scan_keys(scanned_rows)

    # 「相符」從**資產側**數：有幾台登記的資產這次掃得到。
    # 原本是從掃描側數（幾筆掃描結果對得上 CIA），再用 ica_count - overlap 反推「登記卻掃不到」——
    # 一旦一台資產被多筆掃描結果對到（同機多 IP／IP 與主機名分別命中），那個減法就會少算甚至變負數。
    # 現在每個數字都可以點進去看清單，磚塊上的數字必須跟清單筆數一致，所以改成各自獨立算：
    #   overlap   = 登記且掃得到（資產側）      -> /assets?scan_status=overlap
    #   ica_only  = 登記但掃不到（資產側）      -> /assets?scan_status=ica_only
    #   scan_only = 掃到但沒登記（掃描側）      -> /adopt（與 /api/scan/unregistered 同一套判定）
    overlap_count = sum(1 for r in ica_rows if _row_in_keys(r, scan_ips, scan_hostnames))
    ica_only_count = len(ica_rows) - overlap_count

    # ⚠️ scan_only 要對「全部」資產比，不能只對環境篩選後的子集：
    # 「掃到卻沒登記」的意思是「這台機器在網路上，但整個資產清單裡都沒有它」。
    # 一台登記為「測試」環境的機器是有登記的，只是不屬於目前檢視的環境——
    # 拿環境子集去比，會把它算成未登記，磚塊顯示 4 但點進去的 /adopt（不分環境）只有 3。
    # 數字對不上的下鑽比不能點更糟，實測踩到過。
    all_ica_rows = conn.execute("SELECT ip, hostname FROM hardware").fetchall()
    all_ica_ips = {r["ip"] for r in all_ica_rows if r["ip"]}
    all_ica_hostnames = {r["hostname"] for r in all_ica_rows if r["hostname"]}
    scan_only_count = sum(
        1 for r in scanned_rows if not _row_in_keys(r, all_ica_ips, all_ica_hostnames)
    )

    # 一致率專用的「全站」數字：不受環境下拉影響。
    # 理由：一致率的分母要含「掃到卻沒登記」，而那些機器**不屬於任何環境**（沒登記過，
    # 哪來的環境欄位）。拿環境篩選過的登記數去配全站的未登記數，兩邊尺規不同，
    # 換個環境分數就跳動，但實際盤點狀況根本沒變。頭條數字要能一眼比較，所以固定全站。
    total_overlap_count = sum(
        1 for r in all_ica_rows if _row_in_keys(r, scan_ips, scan_hostnames)
    )

    issue_counts = {"異常新增": 0, "異常消失": 0, "漏登記": 0}
    for row in conn.execute(
        "SELECT issue_type, COUNT(*) AS c FROM comparison_result WHERE is_read = 0 GROUP BY issue_type"
    ).fetchall():
        issue_counts[row["issue_type"]] = row["c"]

    return {
        "environment": environment,
        "ica_count": len(ica_rows),
        "scanned_count": len(scanned_rows),
        "overlap_count": overlap_count,
        "ica_only_count": ica_only_count,
        "scan_only_count": scan_only_count,
        # 一致率用這兩個（全站、不隨環境變動），別拿上面那組環境篩選過的去算
        "total_ica_count": len(all_ica_rows),
        "total_overlap_count": total_overlap_count,
        "last_scan_time": last_scan_time,
        "last_scan_ok": len(failed_segments) == 0,
        "failed_segments": failed_segments,
        "issue_counts": issue_counts,
    }


@app.get("/api/issues")
def list_issues(
    issue_type: str | None = None,
    is_read: bool | None = None,
    sort_by: str = "detected_at",
    order: str = "desc",
    session: sqlite3.Row = Depends(require_auth),
    conn: sqlite3.Connection = Depends(get_db),
):
    _check_sort(sort_by, order, ALLOWED_ISSUE_SORT)

    query = "SELECT * FROM comparison_result WHERE 1=1"
    params: list = []
    if issue_type is not None:
        query += " AND issue_type = ?"
        params.append(issue_type)
    if is_read is not None:
        query += " AND is_read = ?"
        params.append(int(is_read))
    query += f" ORDER BY {sort_by} {order.upper()}"

    rows = conn.execute(query, params).fetchall()
    return [dict(r) for r in rows]


@app.patch("/api/issues/{issue_id}")
def mark_issue_read(
    issue_id: int,
    body: MarkReadBody,
    session: sqlite3.Row = Depends(require_auth),
    conn: sqlite3.Connection = Depends(get_db),
):
    """標記已處理（契約已定案功能）。目前只支援標記為已讀，不支援改回未讀。"""
    if not body.is_read:
        raise HTTPException(400, "目前只支援標記已處理（is_read=true）")
    existing = conn.execute(
        "SELECT id FROM comparison_result WHERE id = ?", (issue_id,)
    ).fetchone()
    if existing is None:
        raise HTTPException(404, "查無此問題紀錄")
    mark_comparison_read(conn, issue_id)
    row = conn.execute("SELECT * FROM comparison_result WHERE id = ?", (issue_id,)).fetchone()
    return dict(row)


@app.get("/api/assets/field-groups")
def asset_field_groups(session: sqlite3.Row = Depends(require_auth)):
    """D31精神延伸：常用/進階分層來自設定檔，前端不用自己硬編欄位清單。"""
    return json.loads(FIELD_GROUPS_PATH.read_text(encoding="utf-8"))


# ⚠️ 這支一定要定義在 /api/assets/{asset_serial} 之前。
# FastAPI 依「定義順序」比對路由，先中先贏：放在參數路由後面的話，
# /api/assets/facets 會被當成「查詢序號叫 facets 的資產」而回 404，
# 前端拿不到數字又被 catch 吞掉，畫面只是靜靜地少了數字，很難查。
# （2026-07-29 實際踩到，篩選按鈕的台數一直不出現就是這個原因。）
@app.get("/api/assets/facets")
def asset_facets(
    session: sqlite3.Row = Depends(require_auth), conn: sqlite3.Connection = Depends(get_db)
):
    """各平台／各類型分別有幾台，給篩選按鈕直接標在上面。

    沒有數字的話，使用者得一個一個點進去才知道哪個分類有東西、哪個是空的。

    刻意跟 list_assets 共用 asset_classify：兩邊若各算各的，按鈕上寫「5 台」
    點進去卻是 3 筆，比不顯示數字更糟。

    注意角色是多值的（一台同時聽 3306 和 80 就既是資料庫也是 Web），
    所以類型的數字加總會大於資產總數 —— 那是實情，不是重複計算的 bug。
    """
    import asset_classify

    cls = asset_classify.classify_all(conn)
    by_platform: dict[str, int] = {}
    by_role: dict[str, int] = {}
    for info in cls.values():
        p = info.get("platform") or "未知"
        by_platform[p] = by_platform.get(p, 0) + 1
        for r in info.get("roles") or ["unknown"]:
            by_role[r] = by_role.get(r, 0) + 1
    return {"platform": by_platform, "role": by_role, "total": len(cls)}


def _dismissed_keys(conn: sqlite3.Connection) -> set[tuple[str, str]]:
    """人工標記過「這組不是重複」的 (hostname, ip) key 集合。"""
    return {
        (r["hostname"], r["ip"])
        for r in conn.execute("SELECT hostname, ip FROM duplicate_dismiss").fetchall()
    }


def _duplicate_groups(conn: sqlite3.Connection) -> list[dict]:
    """找出「主機名稱與 IP 都一樣」的多筆資產——幾乎可以確定是同一台被登記了兩次。

    只認兩者都相同，是刻意保守：
      · 只有 IP 相同 → 可能是 DHCP 回收後給了別台，或真的位址衝突，都不是重複登記。
      · 只有主機名相同 → 可能是同一台換過網段，那是搬遷不是重複。
    誤報一筆重複，使用者就得花時間去比對兩台到底哪裡不一樣，寧可少報也不要亂報。

    比對時去空白、主機名不分大小寫：Excel 匯入常常帶進尾端空白或大小寫不一致，
    那種「看起來一模一樣卻被當成兩台」正是最該抓出來的情況。

    停用/報廢/閒置＝退役資產排除在外：不然「停用舊料＋使用中新料同IP」會被誤判成重複登記，
    使用者得花時間去比對兩台根本不是同一回事（見 manage_state.RETIRED_STATUS）。

    人工標記過「這組不是重複」（`duplicate_dismiss`）的組別也排除在外——使用者比機器更清楚
    哪些是同一台合法掛多個業務系統/VIP，不該每次都再跳出來煩他。

    抽成獨立函式讓 release/dismiss 兩個動作也能重跑同一份判準，在伺服器端重新驗證
    使用者操作當下這組資料是不是還跟畫面上看到的一樣（不信任前端傳來的舊快照）。
    """
    import manage_state as ms

    retired_placeholders = ",".join("?" for _ in ms.RETIRED_STATUS)
    dup_keys = conn.execute(
        f"""
        SELECT lower(trim(hostname)) AS h, trim(ip) AS i, COUNT(*) AS n
        FROM hardware
        WHERE hostname IS NOT NULL AND length(trim(hostname)) > 0
          AND ip IS NOT NULL AND length(trim(ip)) > 0
          AND COALESCE(asset_status, '') NOT IN ({retired_placeholders})
        GROUP BY h, i
        HAVING n > 1
        ORDER BY n DESC, h
        """,
        tuple(ms.RETIRED_STATUS),
    ).fetchall()

    dismissed = _dismissed_keys(conn)
    groups = []
    for k in dup_keys:
        if (k["h"], k["i"]) in dismissed:
            continue
        members = conn.execute(
            f"SELECT * FROM hardware WHERE lower(trim(hostname)) = ? AND trim(ip) = ? "
            f"AND COALESCE(asset_status, '') NOT IN ({retired_placeholders}) "
            "ORDER BY asset_serial",
            (k["h"], k["i"], *ms.RETIRED_STATUS),
        ).fetchall()
        if not members:
            continue
        rows = [dict(m) for m in members]
        # 標出這組成員之間「值不一樣」的欄位：若完全一致，刪掉多的那筆即可；
        # 若有差異，就得由人判斷該留哪一筆，不能盲目合併。
        diff_fields = []
        for col in rows[0].keys():
            if col in ("id", "asset_serial", "created_at", "updated_at"):
                continue
            vals = {("" if r.get(col) is None else str(r.get(col)).strip()) for r in rows}
            if len(vals) > 1:
                diff_fields.append(col)
        groups.append({
            "hostname": rows[0].get("hostname"),
            "ip": rows[0].get("ip"),
            "count": len(rows),
            "identical": not diff_fields,
            "diff_fields": diff_fields,
            # 給人工管理介面逐欄比較用：完整欄位，不截斷。
            "members": rows,
        })
    return groups


# 同樣必須排在 /api/assets/{asset_serial} 之前，理由見上面 facets 的說明。
@app.get("/api/assets/duplicates")
def asset_duplicates(
    session: sqlite3.Row = Depends(require_auth), conn: sqlite3.Connection = Depends(get_db)
):
    groups = _duplicate_groups(conn)
    return {
        "groups": groups,
        # 「多出來的筆數」＝真的清乾淨可以少掉幾筆，比「幾組」更能說明問題規模
        "extra_rows": sum(g["count"] - 1 for g in groups),
    }


class DuplicateReleaseBody(BaseModel):
    hostname: str
    ip: str
    release_serials: list[str]
    note: str | None = None


@app.post("/api/assets/duplicates/release")
def release_duplicate_serials(
    body: DuplicateReleaseBody,
    session: sqlite3.Row = Depends(require_auth), conn: sqlite3.Connection = Depends(get_db),
):
    """人工確認「這幾個序號是同一台被重複登記」，把多出來的序號真正刪掉、還給號碼池。

    序號釋放＝直接刪除該 asset_serial 在 hardware/software/personnel 三張表的資料列
    （後兩張表對 hardware.asset_serial 有外鍵約束，必須先刪才能刪 hardware）。
    這是使用者明確要的語意：不是「報廢/停用」（那是資產生命週期狀態，序號還留在表裡，
    UNIQUE 約束會擋住之後同號的新資產進來），而是讓這個序號值真正空出來可以重用。

    這只是內部管理工具，使用者確認不需要保留刪除稽核記錄，故不做快照。

    伺服器端重新查一次目前這組（hostname, ip）真實的成員，不信任前端傳來的序號清單是不是
    還有效——避免兩個人同時操作、或畫面資料已經過期時誤刪（比照 merge_review 批次合併同樣
    的「不信任前端快照」原則）。
    """
    h, i = body.hostname.strip().lower(), body.ip.strip()
    current = {
        r["asset_serial"] for r in conn.execute(
            "SELECT asset_serial FROM hardware WHERE lower(trim(hostname)) = ? AND trim(ip) = ?",
            (h, i),
        ).fetchall()
    }
    release_set = set(body.release_serials)
    if not release_set:
        raise HTTPException(400, "沒有要釋放的序號")
    missing = release_set - current
    if missing:
        raise HTTPException(400, f"這幾個序號已經不在這組裡了，資料可能已變更，請重新整理：{'、'.join(sorted(missing))}")
    if len(release_set) >= len(current):
        raise HTTPException(400, "釋放後這組至少要保留一筆，不能全部釋放")

    for serial in release_set:
        conn.execute("DELETE FROM software WHERE asset_serial = ?", (serial,))
        conn.execute("DELETE FROM personnel WHERE asset_serial = ?", (serial,))
        conn.execute("DELETE FROM hardware WHERE asset_serial = ?", (serial,))
    conn.commit()
    return {"ok": True, "released": sorted(release_set), "remaining": len(current) - len(release_set)}


class DuplicateDismissBody(BaseModel):
    hostname: str
    ip: str
    note: str | None = None


@app.post("/api/assets/duplicates/dismiss")
def dismiss_duplicate_group(
    body: DuplicateDismissBody,
    session: sqlite3.Row = Depends(require_auth), conn: sqlite3.Connection = Depends(get_db),
):
    """標記「這組（同主機名+IP）其實不是重複登記」，例如 DB 主機本來就會掛多個業務系統/VIP。

    之後 /api/assets/duplicates 不會再列出這組。UNIQUE(hostname, ip) 防止重複點擊堆記錄。
    """
    h, i = body.hostname.strip().lower(), body.ip.strip()
    conn.execute(
        "INSERT OR IGNORE INTO duplicate_dismiss (hostname, ip, note, dismissed_by) "
        "VALUES (?, ?, ?, ?)",
        (h, i, body.note, session["username"]),
    )
    conn.commit()
    return {"ok": True}


@app.get("/api/assets")
def list_assets(
    response: Response,
    environment: str | None = None,
    q: str | None = None,
    scan_status: str | None = None,
    filter_field: str | None = None,
    filter_value: str | None = None,
    virtual: str | None = None,
    platform: str | None = None,
    role: str | None = None,
    canonical_os: str | None = None,
    canonical_model: str | None = None,
    location_group: str | None = None,
    environment_group: str | None = None,
    sort_by: str = "hostname",
    order: str = "asc",
    limit: int | None = None,
    offset: int = 0,
    session: sqlite3.Row = Depends(require_auth),
    conn: sqlite3.Connection = Depends(get_db),
):
    """scan_status：讓儀表板的「兩邊相符 / 登記卻掃不到」磚塊可以點進來看是哪幾台。
    判定與 /api/dashboard/stats 共用 _row_in_keys，數字才會跟磚塊對得起來。

    environment 這裡吃單一環境值；儀表板用的是 ENV_FILTER_PRESETS（正式／正式+測試／全部），
    前端下鑽時會把 preset 展開成對應的單一環境或不帶此參數。
    """
    _check_sort(sort_by, order, ALLOWED_ASSET_SORT | _sortable_columns(conn, "hardware"))

    if scan_status is not None and scan_status not in ("overlap", "ica_only"):
        raise HTTPException(400, "scan_status 只接受 overlap（登記且掃得到）或 ica_only（登記但掃不到）")

    query = "SELECT * FROM hardware WHERE 1=1"
    params: list = []
    # 預設排除退役（停用/報廢/閒置），口徑要跟 composition()/eos_summary() 對得起來——
    # 那兩支頭條數字都已經排除退役，這裡不排除的話，點進來的筆數會比頭條數字多，
    # 使用者會以為算錯了（首頁「看全部有效資產」的連結標題就寫著「不含退役」）。
    # 例外：使用者自己用 filter_field=asset_status 明確要看某個狀態（含退役徽章本身
    # 的「另有 N 台退役」連結），這時候該尊重他要看的就是那個狀態，不能反過來擋掉。
    import manage_state as _manage_state

    if filter_field != "asset_status":
        retired_placeholders = ",".join("?" for _ in _manage_state.RETIRED_STATUS)
        query += f" AND COALESCE(asset_status, '') NOT IN ({retired_placeholders})"
        params.extend(_manage_state.RETIRED_STATUS)
    if environment is not None:
        # 同時吃「單一環境值」與儀表板用的 preset（正式／正式+測試／全部）。
        # 儀表板磚塊要能原封不動把目前的環境選擇帶進下鑽，否則點進來的筆數會跟磚塊對不上。
        if environment in ENV_FILTER_PRESETS:
            env_list = ENV_FILTER_PRESETS[environment]
            if env_list is not None:  # None＝全部，不加條件
                placeholders = ",".join("?" for _ in env_list)
                query += f" AND environment IN ({placeholders})"
                params.extend(env_list)
        else:
            query += " AND environment = ?"
            params.append(environment)
    # .strip()：使用者常從 Excel/終端機複製貼上，尾隨空白會讓搜尋直接 0 筆，
    # 看起來像「查無此主機」，其實只是多了一個空格——這是實測踩到的體感問題。
    if q and q.strip():
        query += " AND (hostname LIKE ? OR ip LIKE ? OR asset_serial LIKE ?)"
        like = f"%{q.strip()}%"
        params.extend([like, like, like])
    # 「同值篩選」：畫面上任何一個類別型的值（設備機型／群組名稱／保管者…）都要能點進來
    # 看「還有誰跟它一樣」。欄位名走真實欄位白名單，值走參數化，不會有字串拼進 SQL。
    if filter_field is not None:
        if filter_field not in _sortable_columns(conn, "hardware"):
            raise HTTPException(400, f"不支援的篩選欄位：{filter_field}")
        if filter_value in (None, ""):
            # 點「空值」也要能看——「哪些機器這欄是空的」本身就是有用的盤點問題
            query += f" AND ({filter_field} IS NULL OR {filter_field} = '')"
        elif filter_field == "asset_status" and "," in filter_value:
            # 多值篩選（如「退役」＝停用,報廢,閒置 三種狀態合看）：逗號分隔當 IN 清單。
            # ⚠️ 只限 asset_status 這種有限列舉值欄位——備註/用途說明這類自由文字欄若
            # 剛好存了含逗號的內容，會被誤拆成多值查詢，完全比對不到，寧可不支援。
            values = [v for v in filter_value.split(",") if v]
            placeholders = ",".join("?" for _ in values)
            query += f" AND {filter_field} IN ({placeholders})"
            params.extend(values)
        else:
            query += f" AND {filter_field} = ?"
            params.append(filter_value)
    # 虛擬／實體篩選：is_vm 在資料裡混了 0/1 與 'VM' 字串（納管表單存字串），
    # 不能直接用 filter_field 比值。這裡統一判定：非空且不是 0/none 就算虛擬機。
    if virtual in ("yes", "no"):
        want_vm = virtual == "yes"
        vm_true = "(is_vm IS NOT NULL AND LOWER(CAST(is_vm AS TEXT)) NOT IN ('0','','none','false','否'))"
        query += f" AND {vm_true}" if want_vm else f" AND NOT {vm_true}"
    query += f" ORDER BY {sort_by} {order.upper()}"

    rows = conn.execute(query, params).fetchall()

    if scan_status is not None:
        # 掃描比對用 Python 做而不是塞進 SQL：判定是「IP 或主機名稱任一命中」，
        # 跟 /api/scan/unregistered、儀表板同一套；寫成 SQL 反而會分岔成兩套規則。
        scanned = _scanned_alive_rows(conn, _latest_scan_time(conn))
        ips, hostnames = _scan_keys(scanned)
        want_matched = scan_status == "overlap"
        rows = [r for r in rows if _row_in_keys(r, ips, hostnames) == want_matched]

    # 平台／角色篩選。兩者都在 Python 端做，理由跟 scan_status 一樣：
    # 平台判定是「真 OS > 掃描推測 > 機型」的優先序邏輯，角色來自 host_service 的
    # 監聽埠——都不是 hardware 表上的單一欄位，硬塞進 SQL 會分岔成兩套規則。
    if platform or role:
        import asset_classify

        cls = asset_classify.classify_all(conn)
        if platform:
            wanted = {p.strip() for p in platform.split(",") if p.strip()}
            rows = [r for r in rows if cls.get(r["asset_serial"], {}).get("platform") in wanted]
        if role:
            wanted_r = {x.strip() for x in role.split(",") if x.strip()}
            rows = [r for r in rows
                    if wanted_r & set(cls.get(r["asset_serial"], {}).get("roles", ["unknown"]))]

    # 平台下鑽點某個 OS 版本（如「Windows Server 2022」）：跟 platform 同理，
    # 正規化是「原值 → 收斂寫法」的邏輯，不是資料庫裡的單一欄位，只能在這層過濾。
    # 只比對有真 OS 的資產——composition() 的 by_platform_os 對純推測的機器標「未知版本／(推測)」，
    # 那些沒有穩定的原始字串可比對，這裡就不硬篩，維持只看點得到來源的部分。
    if canonical_os:
        import normalize

        rows = [
            r for r in rows
            if r["os"] and normalize.normalize_os(r["os"], conn, r["device_model"])["canonical"] == canonical_os
        ]

    # EOS 頁點某個硬體型號（如「Dell PowerEdge R740」）：跟 canonical_os 同理。
    # ⚠️ 硬體型號分頁的項目其實有兩種不同來源，篩選要兩種都檢查，缺一都會篩到
    # 0 筆（2026-08-13 實際發現，兩個都是這輪才修）：
    # 1. device_model 側：normalize_model() 算出來的——之前漏傳 hint，靠 hint
    #    才解析出來的型號（如「IBM DS3524 SAN Switch」）永遠篩不到自己。
    # 2. os 側被併過來的（HW_ROUTED_PRODUCTS，例：Dell iDRAC (BMC)／Unisphere
    #    Central）：這些 canonical 其實是 normalize_os() 算出來的，不是
    #    normalize_model()，原本的篩選完全沒檢查這條路徑，這類資產點「N台→」
    #    100% 篩到 0 筆，不是個案。
    if canonical_model:
        import normalize

        def _row_matches_hw_canonical(r: sqlite3.Row) -> bool:
            if r["os"]:
                os_info = normalize.normalize_os(r["os"], conn, r["device_model"])
                if os_info["product"] in normalize.HW_ROUTED_PRODUCTS \
                        and os_info["canonical"] == canonical_model:
                    return True
            if r["device_model"]:
                model_info = normalize.normalize_model(r["device_model"], conn, _model_hint(r))
                if model_info["canonical"] == canonical_model:
                    return True
            return False

        rows = [r for r in rows if _row_matches_hw_canonical(r)]

    # 機房分組（板橋／敦南／內湖／分公司）不是資料庫裡存的值，是 location_groups.json
    # 從 physical_location 原值（01_板橋機房…）收斂出來的，所以只能在這層過濾。
    # 儀表板點「板橋 120 台」要能跳進來看是哪 120 台，用 filter_field 精確比對做不到。
    if location_group:
        import manage_state

        rows = [r for r in rows
                if manage_state.group_location(r["physical_location"]) == location_group]

    # 環境別分組同理（UAT/DEV/OA 併進「測試」）。儀表板的機房×環境別交叉表點格子時
    # 一定要用這個而不是 filter_field=environment 精確比對——「測試」那格算的是合併後
    # 的 705 台，精確比對只撈得到原值剛好是「測試」的 695 台。
    # 格子寫 705、點進去列 695，看的人只會認定系統在騙他，這張表就白做了。
    if environment_group:
        import manage_state

        rows = [r for r in rows
                if manage_state.group_environment(r["environment"]) == environment_group]

    # 分頁：不帶 limit 時行為與以前完全相同（回全部、純陣列），既有呼叫端不受影響。
    # 總筆數放在 X-Total-Count 標頭而不是包成 {items, total}：改回應結構會弄壞所有
    # 現有前端；標頭是加法，不是改法。
    total = len(rows)
    if limit is not None:
        if limit < 1 or limit > 1000:
            raise HTTPException(400, "limit 只接受 1–1000")
        if offset < 0:
            raise HTTPException(400, "offset 不能是負數")
        rows = rows[offset:offset + limit]
    response.headers["X-Total-Count"] = str(total)
    response.headers["Access-Control-Expose-Headers"] = "X-Total-Count"

    # 平台／角色一併回傳（加欄位是加法，不動既有結構）。分類在分頁之後才算，
    # 只算這一頁的那幾十筆，不用為了顯示兩個標籤把全表都跑一遍。
    import asset_classify

    cls = asset_classify.classify_all(conn) if rows else {}
    out = []
    for r in rows:
        d = dict(r)
        c = cls.get(r["asset_serial"], {})
        d["platform"] = c.get("platform", "未知")
        d["roles"] = c.get("roles", ["unknown"])
        # 使用者 2026-08-13 要求：「未分類」清單點進來看到的逐台列表，每一台都要
        # 顯示系統推測，不能只在單筆資產詳情頁才有——不然像「N/A」這種大雜燴群組
        # 裡面明明有能猜到的資產，卻被群組彙總層級擋住看不到。
        guess_value, guess_kind, guess_confirmed = _identity_guess(r, conn)
        d["identity_guess"] = guess_value
        d["identity_guess_kind"] = guess_kind
        d["identity_guess_confirmed"] = guess_confirmed
        out.append(d)
    return out


class BatchLookupBody(BaseModel):
    terms: list[str]


@app.post("/api/assets/batch-lookup")
def batch_lookup_assets(
    body: BatchLookupBody,
    session: sqlite3.Row = Depends(require_auth),
    conn: sqlite3.Connection = Depends(get_db),
):
    """值班用批次查詢（貼上多筆主機名稱/IP，一行一筆）：精確比對，不是like模糊搜尋。"""
    terms = [t.strip() for t in body.terms if t.strip()]
    if not terms:
        return []
    placeholders = ",".join("?" for _ in terms)
    rows = conn.execute(
        f"SELECT * FROM hardware WHERE hostname IN ({placeholders}) OR ip IN ({placeholders})",
        terms + terms,
    ).fetchall()
    return [dict(r) for r in rows]


@app.get("/api/personnel")
def list_personnel(
    q: str | None = None,
    sort_by: str = "person_name",
    order: str = "asc",
    session: sqlite3.Row = Depends(require_auth),
    conn: sqlite3.Connection = Depends(get_db),
):
    _check_sort(sort_by, order, ALLOWED_PERSONNEL_SORT | _sortable_columns(conn, "personnel"))

    query = "SELECT * FROM personnel WHERE 1=1"
    params: list = []
    if q and q.strip():
        query += " AND (person_name LIKE ? OR asset_serial LIKE ?)"
        like = f"%{q.strip()}%"
        params.extend([like, like])
    query += f" ORDER BY {sort_by} {order.upper()}"

    rows = conn.execute(query, params).fetchall()
    return [dict(r) for r in rows]


@app.get("/api/software")
def list_software(
    q: str | None = None,
    sort_by: str = "asset_name",
    order: str = "asc",
    session: sqlite3.Row = Depends(require_auth),
    conn: sqlite3.Connection = Depends(get_db),
):
    _check_sort(sort_by, order, ALLOWED_SOFTWARE_SORT | _sortable_columns(conn, "software"))

    query = "SELECT * FROM software WHERE 1=1"
    params: list = []
    if q and q.strip():
        query += " AND (asset_name LIKE ? OR hostname LIKE ? OR ip LIKE ?)"
        like = f"%{q.strip()}%"
        params.extend([like, like, like])
    query += f" ORDER BY {sort_by} {order.upper()}"

    rows = conn.execute(query, params).fetchall()
    return [dict(r) for r in rows]


@app.get("/api/assets/{asset_serial}")
def asset_detail(
    asset_serial: str,
    session: sqlite3.Row = Depends(require_auth),
    conn: sqlite3.Connection = Depends(get_db),
):
    """主機詳細頁：D24分層邏輯不外露給使用者（見D31）——這裡回傳全部欄位，
    常用/進階顯示由前端依 field-groups 設定檔決定，不是後端替使用者過濾資料。
    """
    hardware = conn.execute(
        "SELECT * FROM hardware WHERE asset_serial = ?", (asset_serial,)
    ).fetchone()
    if hardware is None:
        raise HTTPException(404, "查無此資產")

    personnel = conn.execute(
        "SELECT * FROM personnel WHERE asset_serial = ?", (asset_serial,)
    ).fetchall()
    software = conn.execute(
        "SELECT * FROM software WHERE asset_serial = ?", (asset_serial,)
    ).fetchall()
    history: list[sqlite3.Row] = []
    if hardware["hostname"]:
        history = conn.execute(
            "SELECT * FROM comparison_result WHERE hostname = ? ORDER BY detected_at DESC",
            (hardware["hostname"],),
        ).fetchall()

    import eos
    import normalize

    hw = dict(hardware)
    os_eos_hit = None
    os_info = None
    if hw.get("os"):
        os_info = normalize.normalize_os(hw["os"], conn, hw.get("device_model"))
        os_eos_hit = eos.lookup_os_eos(os_info["canonical"])
    hw_eos_hit = None
    model_info = None
    if hw.get("device_model"):
        # ⚠️ 這裡之前漏傳 hint（資產名稱/資產用途），跟 eos_summary() 統計頁用的
        # 不是同一套邏輯——同一台 device_model 只填「(VM)」的資產，統計頁能靠
        # hint 解析出「HPE 3PAR 8400 storage」，這支查單一資產卻查不到，兩邊
        # 結果對不起來。使用者 2026-08-13 實際發現、補上。
        model_info = normalize.normalize_model(hw["device_model"], conn, _model_hint(hw))
        hw_eos_hit = eos.lookup_hardware_eos(model_info["canonical"])

    # 使用者 2026-08-13 要求：來源系統的 os／device_model 欄位常常空白或認不出來，
    # 而且這個系統對來源資料沒有編輯權限，改不了。這裡唯讀顯示「系統猜測」，
    # 不寫回任何欄位、也不影響上面的 os_eos／hardware_eos 查詢結果（那兩個本來
    # 就已經用了 hint，猜測欄只是把「這是猜的」這件事明講出來，讓人知道不是
    # 來源系統本身寫的）。
    os_guess = None
    if hw.get("asset_purpose") and (not hw.get("os") or not (os_info and os_info["matched"])):
        os_guess = normalize.suggest_os_canonical(hw["asset_purpose"])
    # ⚠️ 使用者 2026-08-13 再次要求：不管是規則直接對到（method="rule"/"alias"，
    # 可信）還是靠 hint 猜的（"hint"/"hint-vendor"，僅供參考），都要顯示——
    # 之前只顯示「靠猜的」，理由是「規則直接對到的已經看得到不用重複」，但這樣
    # 使用者沒辦法分辨「系統沒查到」跟「已經查到、只是覺得不用顯示」，兩種空白
    # 長一樣。confirmed 旗標讓前端用不同樣式區分兩種狀況。
    model_guess = None
    model_guess_confirmed = False
    if model_info and model_info["method"] != "unmatched" \
            and model_info["canonical"] not in normalize._HW_VM_PRODUCTS:
        model_guess = model_info["canonical"]
        model_guess_confirmed = model_info["method"] in ("rule", "alias")

    return {
        "hardware": hw,
        "personnel": [dict(r) for r in personnel],
        "software": [dict(r) for r in software],
        "history": [dict(r) for r in history],
        # 這台的軟硬體 EOS 狀態：查不到官方資料就是 null，不是 0 或空字串——
        # 前端要能分辨「查過沒公開日期」跟「還沒查」。
        "os_eos": ({**os_eos_hit, "status": eos.eos_status(os_eos_hit.get("eos_date"))}
                   if os_eos_hit else None),
        "hardware_eos": ({**hw_eos_hit, "status": eos.eos_status(hw_eos_hit.get("eos_date"))}
                          if hw_eos_hit else None),
        "os_guess": os_guess,
        "model_guess": model_guess,
        "model_guess_confirmed": model_guess_confirmed,
    }


class OnboardBody(BaseModel):
    ip: str
    platform: str          # linux / windows
    username: str
    password: str          # ⚠️ 只用於這一次納管，處理完即丟，絕不寫任何地方


@app.post("/api/onboard")
def onboard_endpoint(
    body: OnboardBody,
    session: sqlite3.Row = Depends(require_auth),
    conn: sqlite3.Connection = Depends(get_db),
):
    """A：UI 一鍵納管。使用者在畫面輸入 sysctl 帳密 → 系統用它去目標機建 webit3scan。

    ⚠️ 憑證不落地是寫死的底線：body.password 只傳給引擎、用完即丟——
    不寫 DB、不寫 log、不進稽核紀錄。稽核只記帳號名、平台、成敗，不記密碼。
    """
    import manage_state
    import onboard_engine

    collector_ip = onboard_engine.resolve_collector_ip(conn)
    # 開一格進度：執行器會逐行把目標主機的輸出寫進去，畫面同時輪詢 /api/onboard/progress。
    # 這支端點本身仍然是同步的（回傳最終結果），進度只是讓等待中的人看得到在做什麼。
    onboard_engine.progress_start(body.ip)
    try:
        result = onboard_engine.onboard(
            host=body.ip, platform=body.platform,
            username=body.username, password=body.password,   # 用完即丟，下面不再引用
            collector_ip=collector_ip,
            # 建出來的帳號名要跟收集端之後拿去登入的一致，否則建完仍然連不進去
            account=manage_state.get_collect_account(conn, body.platform),
        )
    finally:
        onboard_engine.progress_done(body.ip)
    # 稽核：明確只寫無憑證的欄位。body.password 到這裡就結束生命週期，不進 SQL。
    conn.execute(
        "INSERT INTO onboard_audit (target_ip, platform, login_user, trigger, "
        "triggered_by, ok, stage, message, output) VALUES (?,?,?,?,?,?,?,?,?)",
        (body.ip, body.platform, body.username, "manual", session["username"],
         1 if result.ok else 0, result.stage, result.message, result.output),
    )
    conn.commit()

    # 納管成功後：
    # 1) 「發現但沒登記」的主機要先補一筆最小資產，否則收到的 facts 無處可落。
    #    只填系統知道的（IP／掃到的主機名／預設環境），業務欄位留空給人補。
    # 2) 試連＋收 facts，讓使用者立刻看到狀態變「已納管」＋真 OS/序號進來。
    if result.ok:
        try:
            import manage_state
            exists = conn.execute("SELECT 1 FROM hardware WHERE ip = ?", (body.ip,)).fetchone()
            if not exists:
                scan_hn = conn.execute(
                    "SELECT hostname FROM scan_history WHERE ip = ? AND scan_ok = 1 "
                    "ORDER BY scan_time DESC LIMIT 1", (body.ip,)).fetchone()
                insert_hardware(
                    conn, asset_serial=f"AUTO-{body.ip}", ip=body.ip,
                    hostname=(scan_hn["hostname"] if scan_hn and scan_hn["hostname"] else None),
                    environment="正式", asset_status="使用中",
                )
            manage_state.refresh_collect_status(conn)
            manage_state.collect_facts_into_assets(conn)
        except Exception:  # noqa: BLE001 - 收集失敗不影響納管本身已完成的事實
            pass

    return {"ok": result.ok, "stage": result.stage, "message": result.message,
            "output": result.output}


# ===== Push agent（2026-08-14 設計）：主機自己排程回報 disk/memory 狀態，
# collector 端完全不主動連進主機——防火牆只需要「主機 → collector 單一 port」。
# 詳見 backend/agent_scripts/README.md 的兩段式安裝流程（sysinfra 上傳＋root 排程落地）。

class AgentStageBody(BaseModel):
    asset_serial: str


@app.post("/api/agent/stage")
def agent_stage(
    body: AgentStageBody,
    session: sqlite3.Row = Depends(require_auth),
    conn: sqlite3.Connection = Depends(get_db),
):
    """核發一把新的 agent key，組出 Stage 1 要上傳的檔案內容（呼叫端負責實際 SCP
    上傳到目標主機的 /tmp，這支端點本身不連任何主機）。asset_serial 必須已經是
    hardware 表裡的一筆資產——沒有資產記錄就沒地方掛這把 key。

    `push_agent.sh`／`install.sh` 是靜態檔案（見 backend/agent_scripts/），唯二會變的
    是 key 跟 collector URL，各自用獨立小檔案帶過去，不對腳本內容做字串樣板替換——
    腳本本身固定不變才好稽核/比對版本，會變的東西放設定檔。
    """
    scripts_dir = Path(__file__).parent / "agent_scripts"
    exists = conn.execute(
        "SELECT 1 FROM hardware WHERE asset_serial = ?", (body.asset_serial,)
    ).fetchone()
    if not exists:
        raise HTTPException(404, "找不到這筆資產，請先建立資產記錄再種 agent")

    key = agent_auth.issue_host_key(conn, body.asset_serial)
    collector_url = f"http://{onboard_engine.resolve_collector_ip(conn)}:8000"
    return {
        "asset_serial": body.asset_serial,
        "files": {
            "agent_key": key,
            "collector_url": collector_url,
            "push_agent.sh": (scripts_dir / "push_agent.sh").read_text(encoding="utf-8"),
            "install.sh": (scripts_dir / "install.sh").read_text(encoding="utf-8"),
        },
    }


class AgentFactsBody(BaseModel):
    metrics: list[dict]


@app.post("/api/agent/facts")
def agent_facts(
    body: AgentFactsBody,
    asset_serial: str = Depends(require_host_key),
    conn: sqlite3.Connection = Depends(get_db),
):
    """Push agent 每日回報進來的資料。asset_serial 一律用 require_host_key 驗證出來的
    那個，不信任 body 裡任何主機識別欄位——agent 只能寫自己的資料，這是防冒充的關鍵。
    """
    for m in body.metrics:
        key = m.get("key")
        if not key:
            continue
        conn.execute(
            "INSERT INTO host_metric_latest (asset_serial, metric_key, value, unit, collected_at) "
            "VALUES (?, ?, ?, ?, ?) "
            "ON CONFLICT(asset_serial, metric_key) DO UPDATE SET "
            "value = excluded.value, unit = excluded.unit, collected_at = excluded.collected_at, "
            "received_at = datetime('now','localtime')",
            (asset_serial, key, m.get("value"), m.get("unit"), m.get("collected_at")),
        )
    conn.commit()
    return {"ok": True}


@app.get("/api/onboard/hint")
def onboard_hint(
    ip: str,
    session: sqlite3.Row = Depends(require_auth),
    conn: sqlite3.Connection = Depends(get_db),
):
    """納管前的平台研判——**即時探測，不信登記值**。

    為什麼不用 hardware.os：登記的 OS 可能過時、可能是假資料（實例：.101 登記成
    Ubuntu，實際 banner 是 OpenSSH_for_Windows）。拿錯平台的腳本去打，一定失敗。
    這裡即時抓 SSH banner（不需帳密），banner 自報最準；並回報跟登記值是否衝突，
    讓使用者看得到「畫面說 Ubuntu 但實測是 Windows」這種矛盾。
    """
    import fingerprint

    banner = fingerprint.grab_banner(ip, 22)
    scan = conn.execute(
        "SELECT open_ports, os_guess FROM scan_history WHERE ip = ? AND scan_ok = 1 "
        "ORDER BY scan_time DESC LIMIT 1", (ip,)).fetchone()
    ports = [int(p) for p in (scan["open_ports"] or "").split(",") if p.strip().isdigit()] \
        if scan else []
    os_guess = scan["os_guess"] if scan else None
    method = fingerprint.onboard_method(os_guess=os_guess, open_ports=ports, banner=banner)

    registered = conn.execute("SELECT os FROM hardware WHERE ip = ?", (ip,)).fetchone()
    registered_os = registered["os"] if registered else None
    # 衝突：登記說 Linux 但實測 Windows（或反過來）
    conflict = None
    if registered_os and method["method"] in ("linux", "windows"):
        reg_is_win = any(k in registered_os.lower() for k in ("windows", "microsoft"))
        detected_win = method["method"] == "windows"
        if reg_is_win != detected_win:
            conflict = f"登記為「{registered_os}」，但實測是 {'Windows' if detected_win else 'Linux'}"

    return {
        "ip": ip,
        "platform": method["method"] if method["method"] in ("linux", "windows") else "",
        "confidence": method["confidence"],
        "evidence": method["evidence"],
        "banner": banner,
        "registered_os": registered_os,
        "conflict": conflict,
    }


class CredentialBody(BaseModel):
    name: str
    kind: str = "winrm"
    username: str
    password: str          # ⚠️ 加密後才進 DB，回應永不含它
    scope: str | None = None
    note: str | None = None


@app.get("/api/credentials")
def list_credentials(
    session: sqlite3.Row = Depends(require_auth), conn: sqlite3.Connection = Depends(get_db)
):
    """列出收集用憑證——**永不含密碼，連密文都不給**。"""
    import credential_store

    return credential_store.list_public(conn)


@app.post("/api/credentials")
def save_credential(
    body: CredentialBody,
    session: sqlite3.Row = Depends(require_auth),
    conn: sqlite3.Connection = Depends(get_db),
):
    """新增／更新收集用憑證（如 Windows WinRM 服務帳號）。

    密碼加密後才進 DB，加密金鑰另存 0600 檔案（不在 DB 裡——DB 會被備份、被複製，
    金鑰跟著走就等於沒加密）。回應只回 metadata。
    """
    import credential_store

    credential_store.save(conn, body.name, body.kind, body.username, body.password,
                          scope=body.scope, note=body.note)
    return {"ok": True, "credentials": credential_store.list_public(conn)}


@app.delete("/api/credentials/{name}")
def delete_credential(
    name: str,
    session: sqlite3.Row = Depends(require_auth),
    conn: sqlite3.Connection = Depends(get_db),
):
    import credential_store

    n = credential_store.delete(conn, name)
    if not n:
        raise HTTPException(404, "查無此憑證")
    return {"ok": True}


@app.get("/api/onboard/script")
def onboard_script(
    platform: str,
    fmt: str = "oneliner",
    session: sqlite3.Row = Depends(require_auth),
    conn: sqlite3.Connection = Depends(get_db),
):
    """產生「在目標機本機自己跑」的一行指令，或給維運批次佈署的 Ansible playbook。

    為什麼需要這條路：遠端納管要目標機的管理員帳密，但很多情況拿不到——
    Windows 沒有維運帳號慣例、單機環境沒網域、或人就坐在那台機器前面（根本不用遠端）。
    這時給他一行可貼的指令最實際，且**完全不需要任何密碼**。
    腳本內容與遠端用的完全同一份（同一個引擎產出），不會有兩套行為分岔。

    fmt=ansible（僅 Linux）：把同一組動作翻成 playbook，交給資安／維運一次佈完整批。
    ⚠️ playbook 也是即時從收集公鑰組出來的，不是 repo 裡的靜態檔——換過金鑰之後
    靜態檔會安靜地變成錯的（佈下去每台都成功，只是收集全部連不進來）。
    AIX 沒有這個選項：Ansible 不支援 AIX，只能給可貼的腳本。
    """
    import base64

    import onboard_engine

    if platform not in ("linux", "aix", "windows"):
        raise HTTPException(400, "platform 只接受 linux、aix 或 windows")
    collector_ip = onboard_engine.resolve_collector_ip(conn)
    import manage_state

    account = manage_state.get_collect_account(conn, platform)

    if fmt == "ansible":
        if platform != "linux":
            raise HTTPException(
                400, "Ansible playbook 只適用 Linux——AIX 不支援 Ansible（改用可貼的腳本），"
                     "Windows 走 WinRM 不需要在目標機建帳號")
        try:
            content = onboard_engine.build_linux_playbook(
                onboard_engine.collector_pubkey(), collector_ip, account)
        except ValueError as exc:      # 收集金鑰還沒產生
            raise HTTPException(400, str(exc))
        return {
            "platform": platform, "fmt": "ansible",
            "filename": f"{account}_bootstrap.yml",
            "content": content,
            "note": "交給資安／維運：ansible-playbook -i <inventory> "
                    f"{account}_bootstrap.yml",
        }
    if fmt != "oneliner":
        raise HTTPException(400, "fmt 只接受 oneliner 或 ansible")

    try:
        script = onboard_engine.build_script(
            platform, onboard_engine.collector_pubkey(), collector_ip, account)
    except ValueError as exc:      # AIX 帳號名過長、或收集金鑰還沒產生
        raise HTTPException(400, str(exc))
    b64 = base64.b64encode(script.encode()).decode()

    if platform == "linux":
        cmd = f"echo '{b64}' | base64 -d | sudo bash"
        note = "在該機器以 root（或可 sudo 的帳號）執行"
    elif platform == "aix":
        # AIX 沒有 GNU coreutils 的 base64，也未必裝 sudo——用 openssl（標配）＋ ksh，
        # 並且要求以 root 登入。硬套 Linux 那行在 AIX 上一定失敗。
        cmd = f"echo '{b64}' | openssl base64 -d -A | ksh"
        note = "在該機器以 root 執行（AIX 未必裝 sudo，所以不走 sudo）"
    else:
        cmd = (f"$s='{b64}'; $f=\"$env:TEMP\\wb.ps1\"; "
               f"[IO.File]::WriteAllText($f,"
               f"[Text.Encoding]::UTF8.GetString([Convert]::FromBase64String($s)),"
               f"(New-Object Text.UTF8Encoding $true)); "
               f"powershell -ExecutionPolicy Bypass -File $f")
        note = "在該機器開「系統管理員 PowerShell」執行"
    return {"platform": platform, "command": cmd, "note": note}


@app.get("/api/onboard/progress")
def onboard_progress(
    ip: str,
    session: sqlite3.Row = Depends(require_auth),
):
    """納管執行中的即時進度：目標主機現在做到哪一步。

    為什麼需要（使用者 2026-08-16）：畫面原本只能顯示「已經幾秒」，而腳本本來就會
    逐行印「已建立帳號」「佈署收集公鑰」這些話——資訊一直都在，只是被 subprocess
    一次收完、等結束才回來。改成邊跑邊寫進進度表，這裡讀出來給畫面輪詢。

    ⚠️ 只回腳本的 stdout，永不含憑證（腳本本身設計上就不含機密）。
    """
    return onboard_engine.progress_of(ip)


class OnboardVerifyBody(BaseModel):
    ip: str
    platform: str          # linux / aix / windows


@app.post("/api/onboard/verify")
def onboard_verify(
    body: OnboardVerifyBody,
    session: sqlite3.Row = Depends(require_auth),
    conn: sqlite3.Connection = Depends(get_db),
):
    """本機執行納管指令跑完後，立刻試連驗證——不用等重新掃描、不用猜「有沒有生效」。

    使用者 2026-08-17 明確要求：「本機執行」這條路（PAM 代登、拿不到密碼那種情境）
    使用者是在系統看不到的地方跑完指令，系統原本要等下一輪「重新掃描」才會發現這台，
    完成與否完全不明顯。這支端點對單一 IP 立即試一次真收集，成功就直接看到
    hostname/os，失敗就直接看到原因——不用等、不用猜。

    跟 /api/onboard（帳密流程）不同：那支是「系統代你連進去執行納管腳本」，這支是
    「你已經自己執行過了，系統只是幫你驗證有沒有生效」，語意完全不同，所以是獨立端點。
    """
    import manage_state

    if body.platform not in ("linux", "aix", "windows"):
        raise HTTPException(400, "platform 只接受 linux、aix 或 windows")

    row = conn.execute("SELECT asset_serial FROM hardware WHERE ip = ?", (body.ip,)).fetchone()
    if row:
        asset_serial = row["asset_serial"]
    else:
        # 還沒有資產記錄（本機執行流程不像帳密流程會自動先補一筆）——
        # 驗證的前提是「這台在系統裡掛得上號」，沒有就先補到最小記錄，跟
        # /api/onboard 成功後的自動補記錄邏輯一致。
        asset_serial = f"AUTO-{body.ip}"
        conn.execute(
            "INSERT OR IGNORE INTO hardware (asset_serial, ip, environment, asset_status) "
            "VALUES (?, ?, '正式', '使用中')",
            (asset_serial, body.ip),
        )

    # 驗證的本質就是「現在馬上試連一次」，不受之前 collect_ok 的舊狀態影響——
    # 這正是使用者剛執行完腳本、還沒被任何排程重新探測過的那一刻。
    conn.execute("UPDATE hardware SET collect_ok = 1 WHERE asset_serial = ?", (asset_serial,))
    conn.commit()

    result = manage_state.collect_facts_into_assets(conn, only_serial=asset_serial)
    row = conn.execute(
        "SELECT hostname, os, device_model FROM hardware WHERE asset_serial = ?", (asset_serial,)
    ).fetchone()

    if result["updated"] >= 1:
        return {
            "ok": True, "asset_serial": asset_serial,
            "hostname": row["hostname"], "os": row["os"], "device_model": row["device_model"],
        }
    error = result["failed"][0]["error"] if result["failed"] else "收不到任何欄位，原因不明"
    return {"ok": False, "asset_serial": asset_serial, "error": error}


@app.get("/api/onboard/audit")
def onboard_audit_endpoint(
    session: sqlite3.Row = Depends(require_auth), conn: sqlite3.Connection = Depends(get_db)
):
    """納管稽核紀錄（不含任何憑證）。"""
    return [dict(r) for r in conn.execute(
        "SELECT * FROM onboard_audit ORDER BY id DESC LIMIT 100")]


class SegmentBody(BaseModel):
    prefix: str
    enabled: bool = True
    note: str | None = None


class SegmentEnabledBody(BaseModel):
    enabled: bool


class AutoOnboardEnabledBody(BaseModel):
    enabled: bool


@app.get("/api/auto-onboard")
def auto_onboard_state(
    session: sqlite3.Row = Depends(require_auth), conn: sqlite3.Connection = Depends(get_db)
):
    """B 排程自動納管的現況：總開關、授權網段、最近的自動納管稽核。

    授權網段是這功能的安全閘門——排程只碰列在這裡且啟用的網段前綴。稽核只含帳號名/成敗，
    永不含密碼。
    """
    import auto_onboard

    return {
        "enabled": auto_onboard.is_enabled(conn),
        "segments": auto_onboard.list_segments(conn),
        "recent": auto_onboard.recent_auto_audit(conn, 50),
    }


@app.patch("/api/auto-onboard/enabled")
def auto_onboard_set_enabled(
    body: AutoOnboardEnabledBody,
    session: sqlite3.Row = Depends(require_auth),
    conn: sqlite3.Connection = Depends(get_db),
):
    """總開關。關閉＝排程不會自動納管任何機器（預設關閉）。

    這只擋「無人在場的排程」；就算開著，也只碰授權網段內的主機。
    """
    import auto_onboard

    auto_onboard.set_enabled(conn, body.enabled)
    return {"ok": True, "enabled": auto_onboard.is_enabled(conn)}


@app.post("/api/auto-onboard/segments")
def auto_onboard_save_segment(
    body: SegmentBody,
    session: sqlite3.Row = Depends(require_auth),
    conn: sqlite3.Connection = Depends(get_db),
):
    """新增／更新一個授權網段（以前綴為唯一鍵）。加進來才代表「授權排程自動納管這段」。"""
    import auto_onboard

    try:
        auto_onboard.save_segment(conn, body.prefix, body.enabled, body.note)
    except ValueError as exc:
        raise HTTPException(400, str(exc))
    return {"ok": True, "segments": auto_onboard.list_segments(conn)}


@app.patch("/api/auto-onboard/segments/{seg_id}/enabled")
def auto_onboard_toggle_segment(
    seg_id: int,
    body: SegmentEnabledBody,
    session: sqlite3.Row = Depends(require_auth),
    conn: sqlite3.Connection = Depends(get_db),
):
    """啟用／停用一個授權網段。停用＝排程跳過但保留設定。"""
    import auto_onboard

    if not auto_onboard.set_segment_enabled(conn, seg_id, body.enabled):
        raise HTTPException(404, "查無此授權網段")
    return {"ok": True, "segments": auto_onboard.list_segments(conn)}


@app.delete("/api/auto-onboard/segments/{seg_id}")
def auto_onboard_delete_segment(
    seg_id: int,
    session: sqlite3.Row = Depends(require_auth),
    conn: sqlite3.Connection = Depends(get_db),
):
    import auto_onboard

    if not auto_onboard.delete_segment(conn, seg_id):
        raise HTTPException(404, "查無此授權網段")
    return {"ok": True, "segments": auto_onboard.list_segments(conn)}


@app.post("/api/auto-onboard/run")
def auto_onboard_run_now(
    session: sqlite3.Row = Depends(require_auth), conn: sqlite3.Connection = Depends(get_db)
):
    """手動「立即執行」一輪自動納管：試連 → 納管授權網段內未納管的 Linux → 收 facts。

    操作者在場的手動觸發，不受總開關約束，但仍受授權網段約束（沒授權就沒候選）。
    會花數秒到數十秒（每台一次試連／SSH），前端要顯示執行中三態。
    """
    import auto_onboard

    collector_ip = onboard_engine.resolve_collector_ip(conn)
    return auto_onboard.scheduled_cycle(collector_ip=collector_ip, conn=conn)


# ===== 收集入口收斂（決策 C4，2026-08-16）=====
# 使用者只做一個動作：貼網段或 IP 清單、按一下。選路（SSH／WinRM／Agent／只能匯入）
# 由系統依實際探測結果決定，不要人先自己判斷該點哪個按鈕——那個判斷本來就該系統做。

class CollectDispatchBody(BaseModel):
    targets: str          # 網段（10.99.1.0/24）、範圍（10.99.1.10-20）、單一 IP，可混用


@app.post("/api/collect/dispatch")
def collect_dispatch_run(
    body: CollectDispatchBody,
    session: sqlite3.Row = Depends(require_auth),
    conn: sqlite3.Connection = Depends(get_db),
):
    """一個入口跑完整收集流程，回一張分得出成敗的結果表。

    會花數秒到數分鐘（每台要探測、通的還要真的連進去收），前端必須顯示執行中三態。
    憑證（WinRM）由加密憑證庫提供、用完即丟，不進回應也不進結果表。
    """
    import collect_dispatch

    try:
        return collect_dispatch.run_dispatch(
            conn, body.targets, triggered_by=session["username"])
    except ValueError as exc:
        raise HTTPException(400, str(exc))


@app.get("/api/collect/dispatch/latest")
def collect_dispatch_latest(
    session: sqlite3.Row = Depends(require_auth), conn: sqlite3.Connection = Depends(get_db)
):
    """最近一次分派的結果。頁面重新整理／隔天回來還看得到上次跑到哪，不用重跑。"""
    import collect_dispatch

    return collect_dispatch.latest_run(conn) or {"run": None, "results": [], "total": 0,
                                                 "collected": 0, "needs_action": 0,
                                                 "by_status": {}, "by_route": {}}


class AgentPackageBody(BaseModel):
    ip: str


@app.post("/api/collect/agent-package")
def collect_agent_package(
    body: AgentPackageBody,
    session: sqlite3.Row = Depends(require_auth),
    conn: sqlite3.Connection = Depends(get_db),
):
    """替一台「系統進不去」的主機產 Push Agent 安裝包。

    key 綁在資產上，所以這台一定要先有資產記錄。未登記的在這裡補一筆最小資產——
    這是**人明確按下按鈕**才發生的（跟排程自動納管刻意不自動建資產不同，
    見 auto_onboard 的說明），欄位只填系統知道的，業務欄位留空給人補。
    """
    ip = (body.ip or "").strip()
    if not ip:
        raise HTTPException(400, "缺少 ip")
    row = conn.execute("SELECT asset_serial FROM hardware WHERE ip = ? LIMIT 1", (ip,)).fetchone()
    if row:
        serial = row["asset_serial"]
    else:
        scan_hn = conn.execute(
            "SELECT hostname FROM scan_history WHERE ip = ? AND scan_ok = 1 "
            "ORDER BY scan_time DESC LIMIT 1", (ip,)).fetchone()
        serial = f"AUTO-{ip}"
        insert_hardware(
            conn, asset_serial=serial, ip=ip,
            hostname=(scan_hn["hostname"] if scan_hn and scan_hn["hostname"] else None),
            environment="正式", asset_status="使用中",
        )

    scripts_dir = Path(__file__).parent / "agent_scripts"
    key = agent_auth.issue_host_key(conn, serial)
    collector_url = f"http://{onboard_engine.resolve_collector_ip(conn)}:8000"
    return {
        "ip": ip,
        "asset_serial": serial,
        "created_asset": row is None,
        "collector_url": collector_url,
        # 交給「要請人裝」的那位的東西：一支自帶全部內容的 bootstrap 腳本＋一行指令。
        # 腳本本身固定不變（好稽核／好比對版本），會變的 key 與 URL 另外帶。
        "install_command": "sudo bash /tmp/bootstrap_watcher.sh",
        "files": {
            "agent_key": key,
            "collector_url": collector_url,
            "push_agent.sh": (scripts_dir / "push_agent.sh").read_text(encoding="utf-8"),
            "install.sh": (scripts_dir / "install.sh").read_text(encoding="utf-8"),
            "bootstrap_watcher.sh": (scripts_dir / "bootstrap_watcher.sh").read_text(
                encoding="utf-8"),
        },
    }


@app.get("/api/pipeline")
def pipeline_endpoint(
    session: sqlite3.Row = Depends(require_auth), conn: sqlite3.Connection = Depends(get_db)
):
    """納管漏斗：每台機器走到哪一關、下一步要做什麼。

    使用者 2026-08-16：「300 台測試機要一台一台匯，我至少要知道哪些是我還需要處理的。」
    四態只分到「連不連得進去」；這裡把「已納管」再往後拆成事實／服務／帳號幾關，
    每台剛好落在它還沒完成的第一關（互斥窮盡，可對帳）。
    """
    import pipeline

    return pipeline.summarize(conn)


@app.get("/api/manage-state")
def manage_state_endpoint(
    session: sqlite3.Row = Depends(require_auth), conn: sqlite3.Connection = Depends(get_db)
):
    """納管四態：未登記／未納管／已納管／失聯。互斥且窮盡，加總＝你知道的所有機器。

    跟 asset_status 是兩條獨立的軸——一台可以同時「使用中」而且「連不進去」。
    儀表板四格與資產清單的狀態欄共用這一份，數字才不會跟清單對不上。
    """
    import manage_state

    return manage_state.summarize(conn)


@app.get("/api/diagnostics")
def diagnostics_endpoint(
    note: str = "",
    desensitize: bool = True,
    session: sqlite3.Row = Depends(require_auth),
    conn: sqlite3.Connection = Depends(get_db),
):
    """萬用診斷包：出問題時一個動作產出「足以判斷問題在哪」的去識別化資料。

    不綁任何單一功能——核心負責環境快照/去識別化/打包，各功能自己註冊一段
    （見 diagnostics.register）。新功能加一個函式就會自動被收進來。

    desensitize 預設 True：公司資料不出這台機器是硬規則，要關掉必須明確指定。
    """
    import diagnostics as dg
    # 確保各功能的診斷外掛都已註冊（模組載入時才會 register）
    import auto_onboard  # noqa: F401
    import identity  # noqa: F401
    import manage_state  # noqa: F401
    import normalize  # noqa: F401
    import vcenter_autoimport  # noqa: F401

    return dg.collect(conn, note=note, desensitize=desensitize)


@app.get("/api/dashboard/composition")
def dashboard_composition(
    session: sqlite3.Row = Depends(require_auth), conn: sqlite3.Connection = Depends(get_db)
):
    """全站組成統計：我的機器長什麼樣子（各平台幾台、虛實、環境、狀態）。

    這是儀表板該回答的問題——「有幾台 Windows」是統計；
    「兩邊相符／登記卻掃不到」是對帳細節，屬於小功能，不該佔戰情室頭條。
    """
    import manage_state

    return manage_state.composition(conn)


def _model_hint(r: sqlite3.Row) -> str | None:
    """給 normalize_model() 的救援線索：device_model 認不出來時，具體型號常常
    躲在資產名稱或資產用途欄位（使用者 2026-08-13 實際發現「EMC XT480」寫在
    資產用途，不是資產名稱——兩欄都不確定會落在哪個，索性都給）。"""
    parts = [r["asset_name"], r["asset_purpose"]]
    joined = " ".join(p for p in parts if p)
    return joined or None


def _identity_guess(r: sqlite3.Row, conn) -> tuple[str | None, str, bool]:
    """使用者 2026-08-13 重新設計：先前把「猜 OS」跟「猜型號」當兩件獨立的事，
    「未分類」清單只用了 OS 規則去猜，涵蓋率很差——實際上這批 os 查不到的資產，
    資產用途欄裡藏的多半根本是**型號**資訊（3PAR／EMC／IBM Storage 這些），
    不是 OS 版本，選錯規則庫，猜不到不是規則不夠多，是問錯規則。

    這裡統一成一支：os 規則猜不到，換型號規則（用同一套已經驗證過、涵蓋率高
    很多的 normalize_model() hint 機制）再猜一次。

    ⚠️ 使用者 2026-08-13 再次要求：一開始只在「靠 hint 猜到」時才顯示，
    device_model 本身規則就直接認得出來（method="rule"）時刻意不顯示，
    理由是「已經看得到了不用重複」——但使用者反應「不顯示就不知道系統到底
    有沒有判斷對，跟完全沒查是同一種空白」。改成不管是規則直接認出來的還是
    靠 hint 猜的都回傳，用第三個回傳值 confirmed 區分：True＝規則/別名字典
    直接對到（可信），False＝靠 hint 猜的（僅供參考）——由前端決定用不同
    樣式呈現，但兩種狀況都要看得到，不能因為「應該已經知道了」就悄悄跳過。

    回傳 (值, 種類, confirmed)，種類是 "os" 或 "model"，都沒有回 (None, "", False)。
    """
    import normalize

    os_val = r["os"]
    os_matched = bool(os_val) and normalize.normalize_os(os_val, conn, r["device_model"])["matched"]
    if os_matched:
        return None, "", False  # os 本身就查得到，不需要猜

    if r["asset_purpose"]:
        os_guess = normalize.suggest_os_canonical(r["asset_purpose"])
        if os_guess:
            return os_guess, "os", False  # suggest_os_canonical 定義上只做「猜」，不算確認

    if r["device_model"]:
        model_info = normalize.normalize_model(r["device_model"], conn, _model_hint(r))
        if model_info["method"] != "unmatched" \
                and model_info["canonical"] not in normalize._HW_VM_PRODUCTS:
            confirmed = model_info["method"] in ("rule", "alias")
            return model_info["canonical"], "model", confirmed

    return None, "", False


@app.get("/api/eos/summary")
def eos_summary(
    session: sqlite3.Row = Depends(require_auth), conn: sqlite3.Connection = Depends(get_db)
):
    """軟硬體 EOS 總覽：全站有幾台的 OS／硬體已過 EOS、一年內到期、還沒查到官方日期。

    只算有效資產（排除退役）——道理跟 composition() 一樣：停用/報廢的機器沒有
    「還要不要換」的問題，混進來會讓「還有幾台要處理」這個數字失真。
    """
    import eos
    import manage_state
    import normalize

    rows = conn.execute(
        "SELECT asset_serial, hostname, os, device_model, asset_status, asset_name, asset_purpose FROM hardware"
    ).fetchall()
    # 人工歸類覆寫（使用者 2026-08-12）：有些設備技術上偵測到的 OS 沒錯（例：Palo Alto
    # Panorama／Finika 底層真的跑 CentOS/Linux），但維護權責歸網路設備組，優先於自動分類。
    overrides = {
        r["canonical"]: r["category"]
        for r in conn.execute("SELECT canonical, category FROM eos_category_override")
    }

    def bump(d, k):
        d[k] = d.get(k, 0) + 1

    # 使用者 2026-08-12：伺服器/主機 OS 跟網路設備韌體要分開，因為要找不同的人維護；
    # 規則/字典都認不出來的（裸版本號、iDRAC 這類設備類型標籤）獨立一桶「未分類」——
    # 這桶正是使用者最需要人工補資料的清單，不能悄悄混進另外兩桶假裝有分類。
    # 使用者 2026-08-13：MySQL 這類資料庫軟體常被登打人員誤填進 OS 欄位，
    # 技術上偵測到的字串沒錯，但維護權責是「軟體」不是「作業系統」——跟 host_os/firmware
    # 一樣走人工覆寫（eos_category_override），不做自動關鍵字判斷，因為「哪些字串算軟體」
    # 沒有一致規則，交給人工補比較不會誤判。
    # 使用者 2026-08-13 追加「資訊不足」：像「網路設備」「儲存設備」「客製化系統」這種
    # 登打人員填的泛用設備類型標籤，不是規則認不出來（那是「未分類」的情境），是原始
    # 資料本身就沒有可辨識的產品資訊，兩種情況分開才不會讓「未分類」清單裡混進「查也沒用」
    # 的雜訊，稀釋掉真正還有機會補資料的項目。
    BUCKETS = ("host_os", "firmware", "software", "insufficient", "other")
    by_status: dict[str, dict[str, int]] = {b: {} for b in BUCKETS}
    items: dict[str, dict[str, dict]] = {b: {} for b in BUCKETS}
    hw_by_status: dict[str, int] = {}
    hw_items: dict[str, dict] = {}
    device_models_by_canonical: dict[str, dict[str, int]] = {}  # canonical os -> {device_model: 台數}

    for r in rows:
        if (r["asset_status"] or "").strip() in manage_state.RETIRED_STATUS:
            continue
        if r["os"]:
            os_info = normalize.normalize_os(r["os"], conn, r["device_model"])
            canonical = os_info["canonical"]
            if os_info["product"] in normalize.HW_ROUTED_PRODUCTS:
                # 使用者 2026-08-13：iDRAC／Unisphere Central 這類跟硬體綁死的管理
                # 軟韌體，維護權責跟隨它管理的硬體本身，直接併進硬體側對應分類，
                # 不落在 os 側五桶的任何一桶。
                hit = eos.lookup_os_eos(canonical)
                status = eos.eos_status(hit["eos_date"]) if hit else "unknown"
                bump(hw_by_status, status)
                hw_family = overrides.get(canonical, normalize.HW_ROUTED_PRODUCTS[os_info["product"]])
                entry = hw_items.setdefault(canonical, {
                    "name": canonical, "status": status,
                    "eos_date": hit["eos_date"] if hit else None,
                    "source_url": hit["source_url"] if hit else None,
                    "confidence": hit.get("confidence") if hit else None,
                    "note": hit.get("note") if hit else None,
                    "overridden": canonical in overrides,
                    "count": 0,
                    "family": hw_family, "vendor": os_info["vendor"],
                    "model_confirmed": True,  # iDRAC/Unisphere Central 這類是 os 值直接對到具體
                                               # 產品規則，不是靠 hint 猜的，永遠算 confirmed。
                    # 使用者 2026-08-13 實際發現：這批雖然顯示在硬體型號分頁，但 canonical
                    # 其實是 normalize_os() 算出來的，不是 normalize_model()——前端如果只憑
                    # 「顯示在硬體分頁」就假設是 device_model 種類去存改名覆寫，會存錯種類、
                    # 存了但查不到（跟「N台→」篩到0筆是同一個病灶）。這裡明講真正的種類，
                    # 前端改名時要用這個而不是自己用 _origin 猜。
                    "kind": "os",
                })
                entry["count"] += 1
                continue
            bucket = overrides.get(canonical) or normalize.os_category(os_info["product"], canonical)
            hit = eos.lookup_os_eos(canonical)
            status = eos.eos_status(hit["eos_date"]) if hit else "unknown"
            bump(by_status[bucket], status)
            family, linux_distro = normalize.os_family(os_info["product"], canonical)
            # 使用者 2026-08-13 要求：「未分類」清單裡，raw OS 值認不出來時，附上
            # 系統推測（見 _identity_guess()：OS 規則猜不到就換型號規則，不是只
            # 用 OS 規則）。只在第一次建這個 canonical 的項目時算一次——⚠️ 這對
            # 「N/A」這種很多不同資產共用同一個 raw 值的群組是已知限制：只反映
            # 群組裡第一台的猜測，不代表整組都一樣。要看每一台各自的推測，
            # 點「N 台 →」進資產清單，那邊逐列都有算（list_assets 的
            # identity_guess 欄位），不會被群組彙總擋住。
            suggested = None
            if bucket == "other":
                suggested, _, _ = _identity_guess(r, conn)
            entry = items[bucket].setdefault(canonical, {
                "name": canonical, "status": status,
                "eos_date": hit["eos_date"] if hit else None,
                "source_url": hit["source_url"] if hit else None,
                "confidence": hit.get("confidence") if hit else None,
                "note": hit.get("note") if hit else None,
                "count": 0,
                "family": family, "linux_distro": linux_distro, "vendor": os_info["vendor"],
                "overridden": canonical in overrides,
                "suggested": suggested,
            })
            entry["count"] += 1
            # 使用者 2026-08-12：光看「15.7(3)M8」這種裸版本號看不出是什麼設備，
            # 規則認不出來時尤其如此——附上這個版本掛在哪些機型上，人一看機型就知道
            # 這是什麼（例：機型是 Cisco Catalyst，就知道 15.7(3)M8 是 IOS 版本）。
            if r["device_model"]:
                dm_canonical = normalize.normalize_model(r["device_model"], conn, _model_hint(r))["canonical"]
                bump(device_models_by_canonical.setdefault(canonical, {}), dm_canonical)
        if r["device_model"]:
            model_info = normalize.normalize_model(r["device_model"], conn, _model_hint(r))
            canonical = model_info["canonical"]
            hw_family, hw_vendor = normalize.hardware_family(model_info["vendor"], canonical)
            if canonical in overrides:
                hw_family = overrides[canonical]  # 人工覆寫優先於自動判斷（跟 os 側同一張表）
            # 使用者 2026-08-13：「虛擬化」（VMware/KVM/Hyper-V 虛擬機、平台不明的裸
            # 「(VM)」標記）不是實體硬體，根本沒有 EOS 這回事可查——這台機器真正該追的
            # EOS 是它的作業系統，已經算在 host_os 那邊，所以不計入頭部四格統計，避免
            # 同一台機器的「查不到」被算兩次。但使用者後續要求清單本身還是要看得到
            # （純對照用途，「這批資產裡有哪些是虛擬機」），所以項目照樣建立，只是
            # 不 bump 統計數字，也不查 EOS（虛擬機沒有查的意義，直接標未公佈）。
            hit = None if hw_family == "虛擬化" else eos.lookup_hardware_eos(canonical)
            status = eos.eos_status(hit["eos_date"]) if hit else "unknown"
            if hw_family != "虛擬化":
                bump(hw_by_status, status)
            entry = hw_items.setdefault(canonical, {
                "name": canonical, "status": status,
                "eos_date": hit["eos_date"] if hit else None,
                "source_url": hit["source_url"] if hit else None,
                "confidence": hit.get("confidence") if hit else None,
                "note": hit.get("note") if hit else None,
                "overridden": canonical in overrides,
                "count": 0,
                "family": hw_family, "vendor": hw_vendor,
                # 使用者 2026-08-13 要求：硬體型號分頁也要看得出這個型號名是規則直接從
                # device_model 讀到的（confirmed）還是靠 hint（資產名稱/資產用途）
                # 猜出來的——跟 /api/assets 的 identity_guess_confirmed 同一個精神，
                # 只是這裡是彙總層級，只記第一筆建這個 canonical 的判斷方式。
                "model_confirmed": model_info["method"] in ("rule", "alias"),
                "kind": "device_model",  # 真的是從 normalize_model() 算出來的，跟上面
                                          # HW_ROUTED_PRODUCTS 分支（kind="os"）要分清楚。
            })
            entry["count"] += 1

    for bucket in BUCKETS:
        for canonical, entry in items[bucket].items():
            models = device_models_by_canonical.get(canonical, {})
            entry["device_models"] = [
                m for m, _ in sorted(models.items(), key=lambda kv: -kv[1])
            ]

    def sort_items(d: dict) -> list[dict]:
        order = {"expired": 0, "upcoming": 1, "unknown": 2, "ok": 3}
        return sorted(d.values(), key=lambda x: (order.get(x["status"], 9), -x["count"]))

    return {
        "host_os": {"by_status": by_status["host_os"], "items": sort_items(_merge_patch_versions(items["host_os"]))},
        "firmware": {"by_status": by_status["firmware"], "items": sort_items(_merge_patch_versions(items["firmware"]))},
        "software": {"by_status": by_status["software"], "items": sort_items(_merge_patch_versions(items["software"]))},
        "insufficient": {"by_status": by_status["insufficient"], "items": sort_items(_merge_patch_versions(items["insufficient"]))},
        "other": {"by_status": by_status["other"], "items": sort_items(_merge_patch_versions(items["other"]))},
        "hardware": {"by_status": hw_by_status, "items": sort_items(hw_items)},
    }


_VERSION_TAIL = re.compile(r"^(.*\d+\.\d+)\.\d+[A-Za-z0-9()]*$")


def _merge_patch_versions(items_by_canonical: dict) -> dict:
    """把只差小版本號、EOS 查詢結果完全一樣的項目合併成一列，純顯示用途——使用者
    2026-08-13 要求：畫面上一堆「Cisco Network OS 16.12」「16.12.07」看起來很雜，
    但查到的 EOS 日期/來源根本一樣，合併方便看統計。**不動資產本身**，只是把
    eos_summary() 回傳的顯示列合併，device_models/count 都照實加總。

    只有 eos_date／source_url／confidence 三個都相同才合併，避免小版本間 EOS
    其實不一樣時被誤蓋掉（例：某小版本剛好有查到延伸支援公告，其他沒有）。
    """
    groups: dict[tuple, list[dict]] = {}
    for entry in items_by_canonical.values():
        m = _VERSION_TAIL.match(entry["name"])
        base_name = m.group(1) if m else entry["name"]
        key = (base_name, entry["eos_date"], entry.get("source_url"), entry.get("confidence"))
        groups.setdefault(key, []).append(entry)

    merged: dict[str, dict] = {}
    for (base_name, _eos_date, _source_url, _confidence), group in groups.items():
        if len(group) == 1:
            merged[group[0]["name"]] = group[0]
            continue
        combined_models: list[str] = []
        seen_models: set[str] = set()
        for g in group:
            for model in g.get("device_models") or []:
                if model not in seen_models:
                    seen_models.add(model)
                    combined_models.append(model)
        merged[base_name] = {
            **group[0],
            "name": base_name,
            "count": sum(g["count"] for g in group),
            "device_models": combined_models,
            "merged_versions": sorted({g["name"] for g in group}),
        }
    return merged


@app.get("/api/eos/export")
def export_eos_summary(
    session: sqlite3.Row = Depends(require_auth), conn: sqlite3.Connection = Depends(get_db)
):
    """匯出 EOS 頁面清單，每個分類一個分頁，逐資產列出（使用者 2026-08-13 要求，
    給 OS 帳號盤點交 AP 單位統計用途，AP部門/AP負責人/SP負責人依 D27 對應：
    使用單位=AP部門、使用者=AP負責人（AP User）、保管者=SP負責人（SP保管者））。

    跟 /api/normalize/pending/export 不一樣：那支只匯出「查不到」待補的清單，用來
    回填別名。這支是逐資產列出「這台在哪個分類」，一台資產若 OS 跟設備機型分屬
    不同分類（例：OS 是「軟體」但設備機型是「硬體型號」），會各自出現在對應分頁
    ——這是判斷依據本來就不同的兩件事，不能只選一邊。
    """
    import eos
    import manage_state
    import normalize

    PAGE_ORDER = ["作業系統", "網路設備", "硬體型號", "軟體", "資訊不足", "未分類"]
    OS_BUCKET_TO_PAGE = {
        "host_os": "作業系統", "firmware": "網路設備", "software": "軟體",
        "insufficient": "資訊不足", "other": "未分類",
    }
    STATUS_LABEL = {"expired": "已過 EOS", "upcoming": "一年內到期", "ok": "尚在支援期", "unknown": "未公佈"}
    # 使用者 2026-08-13 要求：前面欄位要對齊公司既有的專案工作表單格式（項目/階段/
    # 工作項目/主機IP/負責單位/負責人/開始日/結束日/工期(天)/狀態/備註/聯繫單單號），
    # 方便直接複製貼上去用；沒有資料可填的規劃性欄位（項目/階段/開始日/結束日/
    # 工期(天)/狀態）留空給人工填，不亂塞值。「負責單位」「負責人」先帶 AP部門/
    # AP負責人（跟後面詳細欄位同一個來源，方便對照）。其他細節資訊放後面。
    TEMPLATE_COLUMNS = [
        "項目", "階段", "工作項目", "主機/IP", "負責單位", "負責人",
        "開始日", "結束日", "工期(天)", "狀態", "備註", "聯繫單單號",
    ]
    DETAIL_COLUMNS = [
        "資產狀態", "主機名稱", "IP", "作業系統", "設備機型",
        "EOS狀態", "EOS日期", "可信度", "來源", "AP部門", "AP負責人", "SP負責人",
    ]
    COLUMNS = TEMPLATE_COLUMNS + DETAIL_COLUMNS

    rows = conn.execute(
        "SELECT asset_serial, hostname, ip, os, device_model, asset_status, "
        "asset_name, asset_purpose, usage_unit, user_name, custodian FROM hardware"
    ).fetchall()
    overrides = {
        r["canonical"]: r["category"]
        for r in conn.execute("SELECT canonical, category FROM eos_category_override")
    }

    sheets: dict[str, list] = {p: [] for p in PAGE_ORDER}

    def base_row(r):
        host_ip = " / ".join(v for v in (r["hostname"], r["ip"]) if v)
        return [
            "", "", "", host_ip, r["usage_unit"] or "", r["user_name"] or "",
            "", "", "", "", "", "",
            r["asset_status"] or "", r["hostname"] or "", r["ip"] or "",
            r["os"] or "", r["device_model"] or "",
        ]

    def eos_cols(hit):
        status = eos.eos_status(hit["eos_date"]) if hit else "unknown"
        return [
            STATUS_LABEL.get(status, status),
            hit["eos_date"] if hit else "",
            hit.get("confidence") if hit else "",
            hit.get("source_url") if hit else "",
        ]

    def tail(r):
        return [r["usage_unit"] or "", r["user_name"] or "", r["custodian"] or ""]

    for r in rows:
        if (r["asset_status"] or "").strip() in manage_state.RETIRED_STATUS:
            continue
        # OS 側跟設備機型側各自獨立判斷分類跟 EOS 查詢結果——一台資產若兩邊分屬不同
        # 分類（例：OS 是「軟體」但設備機型是「硬體型號」），或兩邊 EOS 結果不同
        # （這正是這次要補的欄位重點），都不能只選一邊代表，各自輸出一列。
        if r["os"]:
            os_info = normalize.normalize_os(r["os"], conn, r["device_model"])
            canonical = os_info["canonical"]
            if os_info["product"] in normalize.HW_ROUTED_PRODUCTS:
                hw_family = overrides.get(canonical, normalize.HW_ROUTED_PRODUCTS[os_info["product"]])
                page = "網路設備" if hw_family == "網路設備" else "硬體型號"
            else:
                bucket = overrides.get(canonical) or normalize.os_category(os_info["product"], canonical)
                page = OS_BUCKET_TO_PAGE[bucket]
            hit = eos.lookup_os_eos(canonical)
            sheets[page].append(base_row(r) + eos_cols(hit) + tail(r))
        if r["device_model"]:
            model_info = normalize.normalize_model(r["device_model"], conn, _model_hint(r))
            hw_canonical = model_info["canonical"]
            hw_family, _ = normalize.hardware_family(model_info["vendor"], hw_canonical)
            if hw_canonical in overrides:
                hw_family = overrides[hw_canonical]
            page = "網路設備" if hw_family == "網路設備" else "硬體型號"
            hit = None if hw_family == "虛擬化" else eos.lookup_hardware_eos(hw_canonical)
            sheets[page].append(base_row(r) + eos_cols(hit) + tail(r))

    wb = Workbook()
    wb.remove(wb.active)
    for page in PAGE_ORDER:
        ws = wb.create_sheet(page)
        ws.append(COLUMNS)
        for row_out in sheets[page]:
            ws.append(row_out)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"eos_full_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return StreamingResponse(
        buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


class EosCategoryOverrideBody(BaseModel):
    canonical: str
    category: str


@app.post("/api/eos/category-override")
def upsert_eos_category_override(
    body: EosCategoryOverrideBody,
    session: sqlite3.Row = Depends(require_auth), conn: sqlite3.Connection = Depends(get_db),
):
    """人工把某個 canonical 名稱強制指定分類，優先於自動判斷。

    給 Palo Alto Panorama／Finika 這類「技術上偵測到的 OS 沒錯，但維護權責歸網路設備組」
    的情況用——這不是名稱認錯（normalize_alias 的情境），是分類權責跟技術判斷對不上。
    os 側跟 hardware 側共用這張表、共用這支端點，category 值兩邊字彙不同（os 側是
    host_os/firmware/other；hardware 側是網路設備/主機設備/儲存設備/虛擬化/其他），
    eos_summary() 分別在各自的迴圈用各自的字彙解讀，這裡只驗證值屬於兩邊之一即可。
    """
    valid = {
        "host_os", "firmware", "software", "insufficient", "other",
        "網路設備", "主機設備", "儲存設備", "虛擬化", "其他",
    }
    if body.category not in valid:
        raise HTTPException(400, f"category 不支援這個值：{body.category}")
    conn.execute(
        "INSERT INTO eos_category_override (canonical, category, overridden_by) VALUES (?, ?, ?) "
        "ON CONFLICT(canonical) DO UPDATE SET category = excluded.category, "
        "overridden_by = excluded.overridden_by, overridden_at = datetime('now','localtime')",
        (body.canonical, body.category, session["username"]),
    )
    conn.commit()
    return {"ok": True}


@app.delete("/api/eos/category-override/{canonical}")
def delete_eos_category_override(
    canonical: str,
    session: sqlite3.Row = Depends(require_auth), conn: sqlite3.Connection = Depends(get_db),
):
    """撤銷覆寫，改回系統自動判斷的分類。"""
    cur = conn.execute("DELETE FROM eos_category_override WHERE canonical = ?", (canonical,))
    conn.commit()
    if cur.rowcount == 0:
        raise HTTPException(404, "查無此筆覆寫")
    return {"ok": True}


@app.get("/api/normalize/pending")
def normalize_pending(
    session: sqlite3.Row = Depends(require_auth), conn: sqlite3.Connection = Depends(get_db)
):
    """列出規則跟別名字典都認不出來的原始值——EOS 頁「未分類」清單的資料來源，
    也是匯出待補 Excel 的資料來源。"""
    import normalize

    return normalize.pending_values(conn)


class NormalizeAliasBody(BaseModel):
    kind: str
    raw_value: str
    canonical: str
    note: str | None = None


@app.post("/api/normalize/alias")
def upsert_normalize_alias(
    body: NormalizeAliasBody,
    session: sqlite3.Row = Depends(require_auth), conn: sqlite3.Connection = Depends(get_db),
):
    """人工補一筆「原始值 → 標準名稱」的對應。補完立刻生效——normalize.py 的
    _load_aliases() 每次都重新查表，沒有快取，不用重啟服務。"""
    import normalize

    if body.kind not in (normalize.KIND_OS, normalize.KIND_MODEL):
        raise HTTPException(400, f"kind 只接受 {normalize.KIND_OS} 或 {normalize.KIND_MODEL}")
    if not body.canonical.strip():
        raise HTTPException(400, "標準名稱不能是空的")
    conn.execute(
        "INSERT INTO normalize_alias (kind, raw_value, canonical, note) VALUES (?, ?, ?, ?) "
        "ON CONFLICT(kind, raw_value) DO UPDATE SET canonical = excluded.canonical, note = excluded.note",
        (body.kind, body.raw_value, body.canonical.strip(), body.note),
    )
    conn.commit()
    return {"ok": True}


class CanonicalOverrideBody(BaseModel):
    kind: str
    old_canonical: str
    new_canonical: str


@app.post("/api/normalize/canonical-override")
def upsert_canonical_override(
    body: CanonicalOverrideBody,
    session: sqlite3.Row = Depends(require_auth), conn: sqlite3.Connection = Depends(get_db),
):
    """使用者 2026-08-13 要求：不管系統目前顯示的名稱是規則確認還是靠 hint 猜的，
    都能直接改成正確名稱，改完永遠照使用者的為準——跟 /api/normalize/alias
    不一樣：alias 是「原始字串 → 標準名」，只對還沒被規則收斂掉的原始值有用；
    這支是「系統目前算出來的 canonical → 使用者確認的正確 canonical」，對任何
    已經顯示出來的名稱都能用。補完立刻生效，normalize.py 每次都重新查表。
    """
    import normalize

    if body.kind not in (normalize.KIND_OS, normalize.KIND_MODEL):
        raise HTTPException(400, f"kind 只接受 {normalize.KIND_OS} 或 {normalize.KIND_MODEL}")
    if not body.new_canonical.strip():
        raise HTTPException(400, "新名稱不能是空的")
    conn.execute(
        "INSERT INTO normalize_canonical_override (kind, old_canonical, new_canonical, overridden_by) "
        "VALUES (?, ?, ?, ?) "
        "ON CONFLICT(kind, old_canonical) DO UPDATE SET new_canonical = excluded.new_canonical, "
        "overridden_by = excluded.overridden_by, overridden_at = datetime('now','localtime')",
        (body.kind, body.old_canonical, body.new_canonical.strip(), session["username"]),
    )
    conn.commit()
    return {"ok": True}


@app.delete("/api/normalize/canonical-override/{kind}/{old_canonical}")
def delete_canonical_override(
    kind: str,
    old_canonical: str,
    session: sqlite3.Row = Depends(require_auth), conn: sqlite3.Connection = Depends(get_db),
):
    """撤銷改名，改回系統自動判斷的名稱。"""
    cur = conn.execute(
        "DELETE FROM normalize_canonical_override WHERE kind = ? AND old_canonical = ?",
        (kind, old_canonical),
    )
    conn.commit()
    if cur.rowcount == 0:
        raise HTTPException(404, "查無此筆覆寫")
    return {"ok": True}


@app.get("/api/normalize/alias")
def list_normalize_alias(
    kind: str | None = None,
    session: sqlite3.Row = Depends(require_auth), conn: sqlite3.Connection = Depends(get_db),
):
    """列出目前已經補過的別名，給使用者對照/複查用。"""
    if kind:
        rows = conn.execute(
            "SELECT id, kind, raw_value, canonical, note, created_at FROM normalize_alias "
            "WHERE kind = ? ORDER BY created_at DESC", (kind,),
        ).fetchall()
    else:
        rows = conn.execute(
            "SELECT id, kind, raw_value, canonical, note, created_at FROM normalize_alias "
            "ORDER BY created_at DESC",
        ).fetchall()
    return [dict(r) for r in rows]


@app.delete("/api/normalize/alias/{alias_id}")
def delete_normalize_alias(
    alias_id: int,
    session: sqlite3.Row = Depends(require_auth), conn: sqlite3.Connection = Depends(get_db),
):
    """撤銷一筆補錯的別名。"""
    cur = conn.execute("DELETE FROM normalize_alias WHERE id = ?", (alias_id,))
    conn.commit()
    if cur.rowcount == 0:
        raise HTTPException(404, "查無此筆別名")
    return {"ok": True}


@app.get("/api/normalize/pending/export")
def export_normalize_pending(
    session: sqlite3.Row = Depends(require_auth), conn: sqlite3.Connection = Depends(get_db)
):
    """匯出「查不到 EOS 的待補清單」成 Excel，讓人（或 AI）離線查證後填「標準名稱」欄，
    再用 /api/normalize/pending/import 匯入回來。兩個工作表對應 normalize_alias 的
    kind='os'／'device_model'，欄位設計成填完直接可以匯入（不用改欄名/欄序）。
    """
    import normalize

    pending = normalize.pending_values(conn)

    # os 待補清單也附「掛在哪些機型上」——跟 EOS 頁「未分類」分頁同一個線索，
    # 光看裸版本號查不出是什麼，機型是唯一能推斷的依據。
    os_device_models: dict[str, dict[str, int]] = {}
    for r in conn.execute(
        "SELECT os AS v, device_model AS m, asset_name, asset_purpose FROM hardware "
        "WHERE os IS NOT NULL AND os != ''"
    ):
        if r["m"] and not normalize.normalize_os(r["v"], conn, r["m"])["matched"]:
            bucket = os_device_models.setdefault(r["v"], {})
            # 這裡之前漏傳 hint，同一輪其他呼叫點都已經補上（2026-08-13 全面稽核）：
            # 這欄純粹是給人看的提示文字，沒有比對用途，漏了不會造成 0 筆這種功能性
            # 錯誤，但補上讓提示文字跟其他地方顯示的名稱一致。
            dm = normalize.normalize_model(r["m"], conn, _model_hint(r))["canonical"]
            bucket[dm] = bucket.get(dm, 0) + 1

    wb = Workbook()
    ws = wb.active
    ws.title = "作業系統韌體"
    ws.append(["kind", "原始值", "掛在哪些機型上", "出現次數", "標準名稱（填這欄）", "備註"])
    for item in pending[normalize.KIND_OS]:
        models = os_device_models.get(item["raw_value"], {})
        model_str = "、".join(m for m, _ in sorted(models.items(), key=lambda kv: -kv[1]))
        ws.append([normalize.KIND_OS, item["raw_value"], model_str, item["count"], "", ""])

    ws2 = wb.create_sheet("硬體型號")
    ws2.append(["kind", "原始值", "出現次數", "標準名稱（填這欄）", "備註"])
    for item in pending[normalize.KIND_MODEL]:
        ws2.append([normalize.KIND_MODEL, item["raw_value"], item["count"], "", ""])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"eos_pending_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return StreamingResponse(
        buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


@app.post("/api/normalize/pending/import")
def import_normalize_pending(
    file: UploadFile = File(...),
    session: sqlite3.Row = Depends(require_auth), conn: sqlite3.Connection = Depends(get_db),
):
    """匯入補好的待補清單 Excel：逐列讀「標準名稱」欄，非空就補一筆別名；
    空白代表使用者還沒查到，正常跳過，不算錯誤（不是每一列都必須這次補完）。
    """
    import normalize

    if not file.filename or not file.filename.lower().endswith(".xlsx"):
        raise HTTPException(400, "只接受 .xlsx 檔案")

    contents = file.file.read()
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        tmp.write(contents)
        tmp_path = Path(tmp.name)

    applied = 0
    skipped = 0
    errors: list[str] = []
    try:
        try:
            wb = load_workbook(tmp_path, read_only=True, data_only=True)
        except Exception as exc:  # noqa: BLE001 - openpyxl對壞檔案的錯誤型態不一，統一攔截如實回報
            raise HTTPException(400, f"匯入失敗，請確認檔案格式：{exc}") from exc

        for ws in wb.worksheets:
            rows = ws.iter_rows(values_only=True)
            header = next(rows, None)
            if not header:
                continue
            col = {str(h).strip(): i for i, h in enumerate(header) if h is not None}
            required = {"kind", "原始值", "標準名稱（填這欄）"}
            if not required.issubset(col.keys()):
                continue  # 不是這個匯出格式產生的工作表，跳過不當錯誤
            for row in rows:
                if row is None or all(c is None for c in row):
                    continue
                kind = row[col["kind"]]
                raw_value = row[col["原始值"]]
                canonical = row[col["標準名稱（填這欄）"]]
                note = row[col["備註"]] if "備註" in col else None
                if not canonical or not str(canonical).strip():
                    skipped += 1
                    continue
                if kind not in (normalize.KIND_OS, normalize.KIND_MODEL) or not raw_value:
                    errors.append(f"這列 kind/原始值格式不對，略過：{row}")
                    continue
                conn.execute(
                    "INSERT INTO normalize_alias (kind, raw_value, canonical, note) "
                    "VALUES (?, ?, ?, ?) ON CONFLICT(kind, raw_value) DO UPDATE "
                    "SET canonical = excluded.canonical, note = excluded.note",
                    (str(kind), str(raw_value), str(canonical).strip(),
                     str(note).strip() if note else None),
                )
                applied += 1
        wb.close()  # read_only 模式的 workbook 會一直握著檔案控制代碼，Windows 上
        # 不先關掉就刪暫存檔會噴 PermissionError（檔案仍被本行程佔用）——2026-08-12 實測踩到。
        conn.commit()
    finally:
        tmp_path.unlink(missing_ok=True)

    return {"applied": applied, "skipped": skipped, "errors": errors}


@app.post("/api/manage-state/collect")
def manage_state_collect(
    session: sqlite3.Row = Depends(require_auth), conn: sqlite3.Connection = Depends(get_db)
):
    """對已納管的機器收 facts 並寫回資產（真 OS／序號／機型）。"""
    import manage_state

    return manage_state.collect_facts_into_assets(conn)


@app.post("/api/manage-state/refresh")
def manage_state_refresh(
    session: sqlite3.Row = Depends(require_auth), conn: sqlite3.Connection = Depends(get_db)
):
    """手動觸發「試連所有已登記主機」。會花數秒到數十秒（每台一次 SSH）。"""
    import manage_state

    return manage_state.refresh_collect_status(conn)


class AssetUpdateBody(BaseModel):
    fields: dict


@app.put("/api/assets/{asset_serial}")
def update_asset(
    asset_serial: str,
    body: AssetUpdateBody,
    session: sqlite3.Row = Depends(require_auth),
    conn: sqlite3.Connection = Depends(get_db),
):
    """修改一台資產的欄位。

    在這之前系統**完全沒有編輯功能**——資料一旦進來就只能靠重新匯入 Excel 覆蓋，
    連打錯一個字都要重跑匯入。盤點系統的資料本來就會被持續修正，這是必要缺口。

    只接受 hardware 真實存在的欄位（擋掉亂塞）；asset_serial 是主鍵不給改
    （改序號等於換一台，會弄丟 personnel/software 的關聯）。空字串一律存成 NULL，
    否則「空」會有 '' 和 NULL 兩種寫法，篩選和排序就會分岔。
    """
    if conn.execute("SELECT 1 FROM hardware WHERE asset_serial = ?", (asset_serial,)).fetchone() is None:
        raise HTTPException(404, "查無此資產")

    cols = _sortable_columns(conn, "hardware")
    unknown = [k for k in body.fields if k not in cols]
    if unknown:
        raise HTTPException(400, f"不支援的欄位：{', '.join(sorted(unknown))}")

    cleaned = {k: (None if v == "" else v) for k, v in body.fields.items()}
    changed = update_hardware(conn, asset_serial, cleaned)
    # 只有「人動過」才更新 manual_updated_at。updated_at 每次自動匯入都會被刷新
    # （dynassets_import 會寫），拿它當「有沒有人在維護」的指標，跑一次匯入就全部變新鮮，
    # 那個數字是假的（2026-08-15 自我檢查發現）。這欄是資料品質新鮮度的唯一依據。
    if changed:
        conn.execute(
            "UPDATE hardware SET manual_updated_at = ? WHERE asset_serial = ?",
            (datetime.now().strftime("%Y-%m-%d %H:%M:%S"), asset_serial),
        )
        conn.commit()
    row = conn.execute("SELECT * FROM hardware WHERE asset_serial = ?", (asset_serial,)).fetchone()
    return {"updated_fields": changed and len(cleaned) or 0, "hardware": dict(row)}


class ProvisionInfo(BaseModel):
    """申請單來源資訊。source=direct（IT 自建）時整包可以不給。"""
    source: str = "direct"
    request_no: str | None = None
    applicant_unit: str | None = None
    applicant: str | None = None
    unit_manager: str | None = None
    form_date: str | None = None
    change_kind: str | None = None
    raw_fields: dict | None = None


class AssetCreateBody(BaseModel):
    fields: dict
    provision: ProvisionInfo | None = None


# 新資產的預設狀態：還沒過上線檢查前不是「使用中」。
# 刻意選一個不在 manage_state.RETIRED_STATUS 裡的新值——那組是「退役」語意（停用/報廢/
# 閒置），待上線是「還沒開始用」，兩者混在一起會讓報表把新機器算成退役。
PENDING_GOLIVE_STATUS = "待上線"


@app.post("/api/assets/manual")
def create_asset_manual(
    body: AssetCreateBody,
    session: sqlite3.Row = Depends(require_auth),
    conn: sqlite3.Connection = Depends(get_db),
):
    """手動新增一筆資產（CIA 資產清單欄位為主，見 field_mapping.json「硬體」）。

    在這之前系統唯一的建立資產方式是批次匯入（Excel/RVTools/dynassets）——臨時要
    單獨掛一台機器（例如先裝 Push Agent 才發現主機還沒登記），沒有比重跑整份匯入
    更輕的路可走。欄位規則跟 update_asset 共用（白名單、空字串轉 NULL）；
    asset_serial 必填且要唯一（跟批次匯入的 upsert 不同：手動新增遇到重複序號直接
    擋下，不靜默覆蓋既有資產）。

    2026-08-15 起這裡同時是「主機及網路異動需求單」的轉錄入口（provision）。兩種來源
    合併成同一支 API、同一份欄位：申請單位還沒有系統帳號，短期只能填 Word 交給 IT
    承辦人轉錄；等哪天開放讓他們自己上系統填，也只是「同一頁換一個人來填」，不用重做。
    送出後資產是「待上線」，要過上線前檢查表才會變成「使用中」（見 golive.py）。
    """
    asset_serial = str(body.fields.get("asset_serial") or "").strip()
    if not asset_serial:
        raise HTTPException(400, "資產序號必填")
    if conn.execute("SELECT 1 FROM hardware WHERE asset_serial = ?", (asset_serial,)).fetchone():
        raise HTTPException(409, f"資產序號 {asset_serial} 已存在")

    cols = _sortable_columns(conn, "hardware")
    unknown = [k for k in body.fields if k not in cols]
    if unknown:
        raise HTTPException(400, f"不支援的欄位：{', '.join(sorted(unknown))}")

    # 只擋「使用中」的撞號——已停用/報廢/閒置的資產，IP 合法可能被重新分配，
    # 連這種正常情境都擋死只會逼使用者硬改資料才過得了關，防呆變成擋路。
    ip = str(body.fields.get("ip") or "").strip()
    if ip:
        conflict = conn.execute(
            "SELECT asset_serial, hostname FROM hardware WHERE ip = ? AND asset_status = '使用中'",
            (ip,),
        ).fetchone()
        if conflict:
            raise HTTPException(
                409,
                f"IP {ip} 已經被使用中的資產 {conflict['asset_serial']}"
                f"（{conflict['hostname'] or '未填主機名稱'}）使用，請確認是不是同一台",
            )

    cleaned = {k: (None if v == "" else v) for k, v in body.fields.items() if k != "id"}
    cleaned["asset_serial"] = asset_serial
    # 沒指定狀態就是「待上線」；使用者自己選了狀態就尊重他（例如補登記一台早就在跑的機器）
    if not cleaned.get("asset_status"):
        cleaned["asset_status"] = PENDING_GOLIVE_STATUS
    insert_hardware(conn, **cleaned)

    p = body.provision or ProvisionInfo()
    if p.source == "form" and not (p.request_no or "").strip():
        raise HTTPException(400, "依申請單轉錄時，單據編號必填")
    conn.execute(
        "INSERT INTO provision_request "
        "(asset_serial, source, request_no, applicant_unit, applicant, unit_manager, "
        " form_date, change_kind, raw_fields, created_by, created_at) "
        "VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
        (
            asset_serial, p.source, p.request_no, p.applicant_unit, p.applicant,
            p.unit_manager, p.form_date, p.change_kind,
            json.dumps(p.raw_fields, ensure_ascii=False) if p.raw_fields else None,
            session["username"],
            datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        ),
    )
    conn.commit()

    # 一併開一份上線檢查表，並先跑一次 auto 判定（機器測得到的先填好，人只處理剩下的）
    golive.ensure_check(conn, asset_serial)
    golive.refresh_auto_results(conn, asset_serial)

    row = conn.execute("SELECT * FROM hardware WHERE asset_serial = ?", (asset_serial,)).fetchone()
    return {"hardware": dict(row), "golive": golive.get_check_detail(conn, asset_serial)}


# 選單值上限：欄位現有值的「種類數」在這個範圍內才當成選單（避免選錯字），超過代表
# 這欄本質上是自由填寫（型號/用途/機櫃號這類），硬做成選單只會塞爆下拉選單。
_MANUAL_OPTIONS_MAX_DISTINCT = 30


@app.get("/api/assets/manual/field-options")
def manual_field_options(
    session: sqlite3.Row = Depends(require_auth),
    conn: sqlite3.Connection = Depends(get_db),
):
    """手動新增表單用：哪些欄位該做成選單，選項是什麼——從現有資料算，不寫死。

    判斷依據：這欄「現有資料的種類數」夠少（<=30）就代表是類別型欄位（處別/部門/
    環境別這種），做成選單讓使用者選、不能亂打；種類數太多（型號/用途/IP 這種近乎
    每筆都不同）維持自由輸入，選單塞不下也沒意義。門檻是資料驅動的，欄位新增/資料
    累積了自然會跟著變，不用另外維護一份清單。

    保管者／使用者是人名，例外處理：不用 hardware 表自己的歷史值當選項（那份本來就
    可能有錯字——同一個人被打成兩三種寫法，選單只是把錯字原封不動搬進選項，防呆
    等於沒防），改用 personnel 表（人員維護，Excel「人員」分頁匯入的那份）的
    person_name 當唯一真相來源，且不受種類數上限限制（人員名冊本來就可能超過 30 人）。
    """
    _PERSON_NAME_FIELDS = ("custodian", "user_name")

    fields = [
        f for f in _table_columns(conn, "hardware")
        if f not in ("id", "asset_serial", *_PERSON_NAME_FIELDS)
    ]
    options: dict[str, list[str]] = {}
    for field in fields:
        rows = conn.execute(
            f"SELECT DISTINCT {field} FROM hardware "
            f"WHERE {field} IS NOT NULL AND TRIM({field}) != '' "
            f"ORDER BY {field} LIMIT {_MANUAL_OPTIONS_MAX_DISTINCT + 1}"
        ).fetchall()
        values = [str(r[0]) for r in rows]
        if 1 <= len(values) <= _MANUAL_OPTIONS_MAX_DISTINCT:
            options[field] = values

    person_rows = conn.execute(
        "SELECT DISTINCT person_name FROM personnel "
        "WHERE person_name IS NOT NULL AND TRIM(person_name) != '' ORDER BY person_name"
    ).fetchall()
    person_names = [str(r[0]) for r in person_rows]
    if person_names:
        for field in _PERSON_NAME_FIELDS:
            options[field] = person_names
    return options


# ===== 單據檔案室（既有 Word 表單的歸檔與索引）=====

def _doc_archive_dir() -> Path:
    d = get_db_path().parent / "doc_archive"
    d.mkdir(parents=True, exist_ok=True)
    return d


@app.post("/api/documents/import")
def documents_import(
    files: list[UploadFile] = File(...),
    session: sqlite3.Row = Depends(require_auth),
    conn: sqlite3.Connection = Depends(get_db),
):
    """批次歸檔既有的 Word 單據（.doc／.docx 都吃）。

    只抽識別欄位建索引、把原檔存起來，**不解析表格內容**（見 doc_import.py 檔頭）。
    一份檔案失敗不能讓整批中斷——那會逼使用者一次次試錯找出是哪一份有問題。
    """
    import doc_import

    ok, failed = [], []
    for f in files:
        name = Path(f.filename or "doc").name
        try:
            with tempfile.NamedTemporaryFile(delete=False, suffix=Path(name).suffix) as tmp:
                tmp.write(f.file.read())
                tmp_path = Path(tmp.name)
            try:
                ok.append(doc_import.import_document(
                    conn, tmp_path, _doc_archive_dir(), session["username"], original_name=name,
                ))
            finally:
                tmp_path.unlink(missing_ok=True)
        except Exception as e:  # noqa: BLE001 - 逐份回報，不讓一顆壞蘋果毀掉整批
            failed.append({"file_name": name, "error": str(e)})
    return {
        "imported": len(ok), "failed": len(failed), "results": ok, "errors": failed,
        "auto_bound": sum(1 for r in ok if r["bind_confidence"] == "auto"),
        "need_review": sum(1 for r in ok if r["bind_confidence"] == "review"),
    }


@app.get("/api/documents")
def documents_list(
    confidence: str | None = Query(None),
    q: str | None = Query(None),
    session: sqlite3.Row = Depends(require_auth),
    conn: sqlite3.Connection = Depends(get_db),
):
    """單據清單／搜尋。q 連 Word 內文一起搜——使用者以前要找一筆資料得把每份檔案
    打開來看，只搜檔名和單號等於沒解決那個問題。命中會附上片段。"""
    import doc_import

    return {
        "documents": doc_import.list_documents(conn, confidence, q),
        "summary": doc_import.summary(conn),
    }


@app.get("/api/documents/{doc_id}/download")
def documents_download(
    doc_id: int,
    session: sqlite3.Row = Depends(require_auth),
    conn: sqlite3.Connection = Depends(get_db),
):
    """下載原始 Word。索引查到了卻打不開原檔，等於沒歸檔。"""
    row = conn.execute("SELECT * FROM doc_archive WHERE id = ?", (doc_id,)).fetchone()
    if row is None:
        raise HTTPException(404, "找不到這份單據")
    path = Path(row["file_path"])
    if not path.exists():
        raise HTTPException(410, f"索引還在，但原始檔不見了：{row['file_name']}")
    # 全文檢索那份已遮掉帳密，但原始 Word 沒有（那是稽核證據，不能改）。
    # 既然一定要能下載，至少留下「誰在什麼時候看了哪一份」。
    conn.execute(
        "INSERT INTO doc_download_audit (doc_id, file_name, username, downloaded_at) "
        "VALUES (?, ?, ?, ?)",
        (doc_id, row["file_name"], session["username"],
         datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
    )
    conn.commit()
    return FileResponse(path, filename=row["file_name"])


class DocReviewBody(BaseModel):
    values: dict[str, str] | None = None


@app.post("/api/documents/{doc_id}/review")
def documents_review(
    doc_id: int,
    body: DocReviewBody,
    session: sqlite3.Row = Depends(require_auth),
    conn: sqlite3.Connection = Depends(get_db),
):
    """人工審核：確認或修正抽出來的規格值（CPU／記憶體／硬碟／OS 版本）。

    使用者 2026-08-15 要求的關卡。自由填寫欄抓歪不會報錯，唯一有效的防線就是有人看過；
    沒審過的值標成 pending，不參與任何比對，避免「沒人看過的數字被當成事實」。
    """
    import doc_import

    try:
        return doc_import.review_document(conn, doc_id, body.values, session["username"])
    except ValueError as e:
        raise HTTPException(404, str(e))


class DocBindBody(BaseModel):
    asset_serial: str | None = None


@app.post("/api/documents/{doc_id}/bind")
def documents_bind(
    doc_id: int,
    body: DocBindBody,
    session: sqlite3.Row = Depends(require_auth),
    conn: sqlite3.Connection = Depends(get_db),
):
    import doc_import

    try:
        doc_import.bind_document(conn, doc_id, body.asset_serial)
    except ValueError as e:
        raise HTTPException(400, str(e))
    return {"ok": True}


@app.get("/api/assets/{asset_serial}/documents")
def asset_documents(
    asset_serial: str,
    session: sqlite3.Row = Depends(require_auth),
    conn: sqlite3.Connection = Depends(get_db),
):
    """這台資產的單據史（申請單、上線檢查表），含兩種單互相對應的關聯。"""
    import doc_import

    return {
        "documents": doc_import.documents_of_asset(conn, asset_serial),
        # 一台機器常有多張單（新增→異動→異動）：「當初申請多少」跟「現在應該多少」
        # 是兩個問題，時間軸兩個都答得出來
        "timeline": doc_import.asset_timeline(conn, asset_serial),
    }


# ===== 網段配置表 =====
# 來源是公司的「總分公司網段配置表」Excel。三個用途：IP 配置的三層選單、
# 掃描範圍的依據（弱掃說明的排除註記）、資料品質涵蓋率的分母。

@app.get("/api/segments")
def segments_list(
    session: sqlite3.Row = Depends(require_auth),
    conn: sqlite3.Connection = Depends(get_db),
):
    import segments

    return {"segments": segments.list_segments(conn)}


@app.get("/api/segments/tree")
def segments_tree(
    session: sqlite3.Row = Depends(require_auth),
    conn: sqlite3.Connection = Depends(get_db),
):
    """機房 → 環境 → 網段，給新增資產的 IP 選單用。"""
    import segments

    return {"tree": segments.tree(conn)}


@app.get("/api/segments/scan-candidates")
def segments_scan_candidates(
    session: sqlite3.Row = Depends(require_auth),
    conn: sqlite3.Connection = Depends(get_db),
):
    """該掃哪些網段／哪些被註記排除。資料品質頁「涵蓋率 0%」的下一步就是看這個。"""
    import segments

    return segments.scan_candidates(conn)


@app.get("/api/segments/ips")
def segments_ips(
    cidr: str = Query(...),
    session: sqlite3.Row = Depends(require_auth),
    conn: sqlite3.Connection = Depends(get_db),
):
    """這段裡已登記的 IP 與建議可用的下一個。"""
    import segments

    try:
        return segments.segment_ips(conn, cidr)
    except ValueError as e:
        raise HTTPException(400, str(e))


@app.post("/api/segments/import")
def segments_import(
    file: UploadFile = File(...),
    session: sqlite3.Row = Depends(require_auth),
    conn: sqlite3.Connection = Depends(get_db),
):
    """匯入網段配置表（Excel 為主，也吃 Tab 分隔文字檔）。

    **整批取代**：Excel 是這份清單的唯一真相，段被刪掉就該從系統消失。
    解析失敗（一格兩段、IP 範圍）的列照樣入庫並列進警告——靜默丟掉網段，
    之後不會有人發現，而「系統裡沒有這段」跟「這段不存在」是兩件事。
    """
    import segments

    name = Path(file.filename or "segments.xlsx").name
    if Path(name).suffix.lower() not in (".xlsx", ".xlsm", ".txt", ".tsv", ".csv"):
        raise HTTPException(400, "只接受 Excel（.xlsx）或 Tab 分隔文字檔")
    with tempfile.NamedTemporaryFile(delete=False, suffix=Path(name).suffix) as tmp:
        tmp.write(file.file.read())
        tmp_path = Path(tmp.name)
    try:
        summary = segments.import_segments(conn, tmp_path)
    except ValueError as e:
        raise HTTPException(400, str(e))
    finally:
        tmp_path.unlink(missing_ok=True)
    return {"file": name, **summary}


@app.get("/api/data-quality")
def data_quality_summary(
    session: sqlite3.Row = Depends(require_auth),
    conn: sqlite3.Connection = Depends(get_db),
):
    """資料品質量測：盤點清單到底準不準，用查得到來源的數字講。

    只對「有機器事實可以對照」的維度給正確率；保管者/用途/CIA 這種人為判斷欄位
    給填寫率與新鮮度，並在畫面上講清楚衡量的不是對錯——硬給正確率是在編數字。
    """
    import data_quality

    return data_quality.measure(conn)


@app.get("/api/data-quality/{dim_key}")
def data_quality_offenders(
    dim_key: str,
    session: sqlite3.Row = Depends(require_auth),
    conn: sqlite3.Connection = Depends(get_db),
):
    """某維度不合格的是哪幾台。數字一定要能下鑽，不然沒有行動意義。"""
    import data_quality

    return {"items": data_quality.list_offenders(conn, dim_key)}


# ===== 資產生命週期：申請單 → 上線前檢查 → 基線回檢 =====
# 模型見 golive.py 檔頭。這區的端點都圍著「一台資產一份檢查表」轉。

def _attachment_dir() -> Path:
    """簽核用的申請單掃描檔落地位置。放 DATA 底下跟 DB 同一層，
    才會被既有的每日備份（backup.py 備 DATA）一起帶走。"""
    d = get_db_path().parent / "provision_attachments"
    d.mkdir(parents=True, exist_ok=True)
    return d


@app.get("/api/assets/{asset_serial}/provision")
def get_provision(
    asset_serial: str,
    session: sqlite3.Row = Depends(require_auth),
    conn: sqlite3.Connection = Depends(get_db),
):
    """這台資產的申請單資訊（沒有就回 null，直接新增的資產也算沒有紙本單）。"""
    row = conn.execute(
        "SELECT * FROM provision_request WHERE asset_serial = ? ORDER BY id DESC LIMIT 1",
        (asset_serial,),
    ).fetchone()
    if row is None:
        return {"provision": None}
    d = dict(row)
    if d.get("raw_fields"):
        try:
            d["raw_fields"] = json.loads(d["raw_fields"])
        except ValueError:
            pass  # 存進去的不是合法 JSON 就原字串回去，不要讓整頁掛掉
    return {"provision": d}


@app.post("/api/assets/{asset_serial}/provision-attachment")
def upload_provision_attachment(
    asset_serial: str,
    file: UploadFile = File(...),
    session: sqlite3.Row = Depends(require_auth),
    conn: sqlite3.Connection = Depends(get_db),
):
    """上傳簽核用的申請單（Word/PDF 掃描檔）。存原始檔，不解析。

    刻意不做 Word 解析自動建資產：那種表是巢狀表格＋□/☑ 符號，版面一動就抓歪，
    而且抓錯不會報錯，是「安靜把錯資料寫進盤點清單」——比沒有還糟。
    """
    row = conn.execute(
        "SELECT id FROM provision_request WHERE asset_serial = ? ORDER BY id DESC LIMIT 1",
        (asset_serial,),
    ).fetchone()
    if row is None:
        raise HTTPException(404, f"{asset_serial} 沒有申請單紀錄")

    name = Path(file.filename or "attachment").name
    if Path(name).suffix.lower() not in (".doc", ".docx", ".pdf", ".png", ".jpg", ".jpeg"):
        raise HTTPException(400, "只接受 Word／PDF／圖片掃描檔")
    # 檔名帶 asset_serial 與流水號，避免不同資產的同名檔互相覆蓋
    safe = f"{asset_serial}_{row['id']}_{name}"
    path = _attachment_dir() / safe
    path.write_bytes(file.file.read())

    conn.execute(
        "UPDATE provision_request SET attachment_name = ?, attachment_path = ? WHERE id = ?",
        (name, str(path), row["id"]),
    )
    conn.commit()
    return {"attachment_name": name}


@app.get("/api/assets/{asset_serial}/provision-attachment-file")
def download_provision_attachment(
    asset_serial: str,
    session: sqlite3.Row = Depends(require_auth),
    conn: sqlite3.Connection = Depends(get_db),
):
    """下載申請單的簽核掃描檔。存了卻打不開等於沒存（2026-08-15 自我檢查補上）。"""
    row = conn.execute(
        "SELECT attachment_name, attachment_path FROM provision_request "
        "WHERE asset_serial = ? AND attachment_path IS NOT NULL ORDER BY id DESC LIMIT 1",
        (asset_serial,),
    ).fetchone()
    if row is None:
        raise HTTPException(404, "這台資產沒有申請單附件")
    path = Path(row["attachment_path"])
    if not path.exists():
        raise HTTPException(410, f"紀錄還在，但原始檔不見了：{row['attachment_name']}")
    return FileResponse(path, filename=row["attachment_name"])


@app.get("/api/golive")
def golive_list(
    status: str = Query("open"),
    session: sqlite3.Row = Depends(require_auth),
    conn: sqlite3.Connection = Depends(get_db),
):
    """上線檢查表清單（預設看還沒過的——那才是待辦）。"""
    rows = conn.execute(
        "SELECT g.*, h.hostname, h.ip, h.os, h.asset_status, h.asset_name "
        "FROM golive_check g LEFT JOIN hardware h ON h.asset_serial = g.asset_serial "
        "WHERE g.status = ? ORDER BY g.started_at DESC",
        (status,),
    ).fetchall()
    out = []
    for r in rows:
        detail = golive.get_check_detail(conn, r["asset_serial"])
        out.append({**dict(r), "total": detail["total"], "done": detail["done"]})
    return {"checks": out}


@app.get("/api/golive/{asset_serial}")
def golive_detail(
    asset_serial: str,
    session: sqlite3.Row = Depends(require_auth),
    conn: sqlite3.Connection = Depends(get_db),
):
    if conn.execute(
        "SELECT 1 FROM hardware WHERE asset_serial = ?", (asset_serial,)
    ).fetchone() is None:
        raise HTTPException(404, f"找不到資產 {asset_serial}")
    golive.refresh_auto_results(conn, asset_serial)
    return golive.get_check_detail(conn, asset_serial)


class GoliveItemBody(BaseModel):
    item_key: str
    verdict: str


@app.post("/api/golive/{asset_serial}/item")
def golive_set_item(
    asset_serial: str,
    body: GoliveItemBody,
    session: sqlite3.Row = Depends(require_auth),
    conn: sqlite3.Connection = Depends(get_db),
):
    try:
        golive.set_item_verdict(conn, asset_serial, body.item_key, body.verdict,
                                session["username"])
    except ValueError as e:
        raise HTTPException(400, str(e))
    return golive.get_check_detail(conn, asset_serial)


@app.post("/api/golive/{asset_serial}/pass")
def golive_pass(
    asset_serial: str,
    session: sqlite3.Row = Depends(require_auth),
    conn: sqlite3.Connection = Depends(get_db),
):
    """全部項目處理完才准通過；通過同時把資產轉「使用中」並把 auto 項存成基線。"""
    try:
        return golive.pass_check(conn, asset_serial, session["username"])
    except ValueError as e:
        raise HTTPException(400, str(e))


@app.get("/api/drift")
def drift_list(
    status: str = Query("open"),
    session: sqlite3.Row = Depends(require_auth),
    conn: sqlite3.Connection = Depends(get_db),
):
    """基線失效清單：上線時刻意設定成這樣、現在不一樣了的項目。"""
    return {"drifts": golive.list_drift(conn, status if status != "all" else None)}


class DriftDispositionBody(BaseModel):
    status: str
    note: str | None = None


@app.post("/api/drift/{drift_id}/disposition")
def drift_disposition(
    drift_id: int,
    body: DriftDispositionBody,
    session: sqlite3.Row = Depends(require_auth),
    conn: sqlite3.Connection = Depends(get_db),
):
    """標記處置。沿用帳號盤點 finding_disposition 的精神：標過的下次回檢還記得，
    不要每天重新亮一次同一條紅燈；但 ack 只是「我知道了」，不代表基線改了——
    機器恢復成基線時仍然會自動轉 fixed。"""
    if body.status not in ("open", "ack", "fixed"):
        raise HTTPException(400, f"不支援的處置狀態：{body.status}")
    n = conn.execute(
        "UPDATE baseline_drift SET status = ?, note = ?, decided_by = ?, decided_at = ? "
        "WHERE id = ?",
        (body.status, body.note, session["username"],
         datetime.now().strftime("%Y-%m-%d %H:%M:%S"), drift_id),
    ).rowcount
    conn.commit()
    if not n:
        raise HTTPException(404, "找不到這筆基線失效紀錄")
    return {"ok": True}


@app.post("/api/drift/recheck")
def drift_recheck(
    session: sqlite3.Row = Depends(require_auth),
    conn: sqlite3.Connection = Depends(get_db),
):
    """手動觸發回檢。平常掛在每日掃描之後自動跑（見 scan_service）。"""
    return golive.run_drift_check(conn)


@app.get("/api/assets/manual/field-groups")
def manual_field_groups(session: sqlite3.Row = Depends(require_auth)):
    """手動新增表單的欄位分組（見 manual_form_groups.json）——29 個欄位攤平成一片沒人填得下去。

    注意跟 field_groups.json 不是同一回事：那份是「資產查詢頁先顯示哪些欄位」的常用/進階
    分層，這份是「新增表單怎麼分段排版」。兩者的分法本來就不同，不要合併。

    也獨立於 field_mapping.json：那份是「Excel 標題 → 欄位名」，匯入/匯出都靠它，
    分組硬塞進去要改 value 的型別，會一次打壞 excel_import、匯出表頭與匯入設定頁。
    這裡回傳的 key 是資料庫欄位名，前端拿去對 field_mapping 算出來的欄位清單即可。

    設定檔壞掉/讀不到不讓整頁掛掉：回空清單，前端會退回原本的單一格子排版。
    """
    path = Path(__file__).parent / "manual_form_groups.json"
    try:
        return {"groups": json.loads(path.read_text(encoding="utf-8"))["groups"]}
    except (OSError, ValueError, KeyError):
        return {"groups": []}


@app.get("/api/os-catalog")
def os_catalog(
    session: sqlite3.Row = Depends(require_auth),
    conn: sqlite3.Connection = Depends(get_db),
):
    """手動新增表單的作業系統欄位：分層選單（大類 → 發行版／產品 → 版本），不能亂打。

    OS 原始字串種類太多（正式環境 250+ 種），沒辦法像其他欄位一樣直接攤平成一個選單；
    但也不用另外維護一份分類表——直接復用既有、已經測過的 normalize_os()／os_family()
    （EOS 頁、儀表板平台統計都在用同一套），對現有資料裡每個不同的 os 字串分類一次，
    照 (大類, 發行版) 分組，組內選項是算好的 canonical 字串（使用者選到的就是標準名，
    不會把舊資料的雜亂寫法帶進新資產）。認不出來的字串歸進「其他／未分類」，不會憑空消失。
    """
    import normalize

    rows = conn.execute(
        "SELECT DISTINCT os FROM hardware WHERE os IS NOT NULL AND TRIM(os) != ''"
    ).fetchall()

    tree: dict[str, dict[str, set[str]]] = {}
    for r in rows:
        raw = r["os"]
        result = normalize.normalize_os(raw, conn)
        canonical = result["canonical"] or str(raw).strip()
        product = result["product"]
        family, linux_distro = normalize.os_family(product, canonical)
        distro_key = linux_distro if family == "Linux" else (product or "未分類")
        tree.setdefault(family, {}).setdefault(distro_key, set()).add(canonical)

    return {
        family: {distro: sorted(versions) for distro, versions in sorted(distros.items())}
        for family, distros in sorted(tree.items())
    }


NO_IMPORT_SENTINEL = "不匯入"


def _table_columns(conn: sqlite3.Connection, table: str) -> set[str]:
    return {row["name"] for row in conn.execute(f"PRAGMA table_info({table})").fetchall()}


@app.get("/api/import/last")
def import_last(
    session: sqlite3.Row = Depends(require_auth),
    conn: sqlite3.Connection = Depends(get_db),
):
    row = get_latest_import_log(conn)
    return dict(row) if row else None


@app.get("/api/import/field-mapping")
def import_field_mapping(session: sqlite3.Row = Depends(require_auth)):
    """D14精神：欄位對應可調整、不寫死。回傳目前的field_mapping.json內容。"""
    return load_mapping()


class FieldMappingBody(BaseModel):
    mapping: dict[str, dict[str, str]]


@app.put("/api/import/field-mapping")
def update_import_field_mapping(
    body: FieldMappingBody,
    session: sqlite3.Row = Depends(require_auth),
    conn: sqlite3.Connection = Depends(get_db),
):
    """value為「不匯入」的項目不寫入設定檔（等於移除對應）；其餘value要對得上目標資料表
    的實際欄位名稱（PRAGMA table_info動態查，不寫死欄位清單，schema改了這裡自動跟上）。

    _comment說明欄位：load_mapping()讀取時會把它拿掉（不是真的分頁對應資料），這裡存檔
    要把原本檔案裡的_comment原封放回去，不然存一次設定檔裡給後人看的說明文字就消失了。
    """
    cleaned: dict[str, dict[str, str]] = {}
    for sheet, columns in body.mapping.items():
        if sheet not in SHEET_CONFIG:
            raise HTTPException(400, f"不支援的分頁：{sheet}")
        valid_columns = _table_columns(conn, SHEET_CONFIG[sheet]["table"])
        cleaned[sheet] = {}
        for excel_header, system_field in columns.items():
            if system_field == NO_IMPORT_SENTINEL:
                continue
            if system_field not in valid_columns:
                raise HTTPException(
                    400, f"{sheet}分頁「{excel_header}」對應到不存在的欄位：{system_field}"
                )
            cleaned[sheet][excel_header] = system_field

    existing_raw = json.loads(MAPPING_PATH.read_text(encoding="utf-8"))
    to_write = {}
    if "_comment" in existing_raw:
        to_write["_comment"] = existing_raw["_comment"]
    to_write.update(cleaned)

    MAPPING_PATH.write_text(
        json.dumps(to_write, ensure_ascii=False, indent=2) + "\n", encoding="utf-8"
    )
    return cleaned


@app.post("/api/import/excel")
def import_excel_upload(
    file: UploadFile = File(...),
    session: sqlite3.Row = Depends(require_auth),
    conn: sqlite3.Connection = Depends(get_db),
):
    """故意用sync def（不是async def）：UploadFile.file是同步可讀的SpooledTemporaryFile，
    sync路由可以直接讀，不需要await。

    ⚠️ 原本這裡寫著「FastAPI 會把 get_db 依賴跟 sync 路由排進同一個 threadpool worker
    執行緒」——那個假設是錯的，並行時兩者可能落在不同 worker，會炸
    sqlite3「不同執行緒不能共用連線」。真正的修正在 db.get_connection()
    （check_same_thread=False），不是靠 sync/async 的寫法去賭執行緒。
    """
    if not file.filename or not file.filename.lower().endswith(".xlsx"):
        raise HTTPException(400, "只接受 .xlsx 檔案")

    contents = file.file.read()
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        tmp.write(contents)
        tmp_path = Path(tmp.name)

    try:
        try:
            summary = import_excel(tmp_path, conn)
        except Exception as exc:  # noqa: BLE001 - openpyxl對壞檔案的錯誤型態不一，統一攔截如實回報
            raise HTTPException(400, f"匯入失敗，請確認檔案格式：{exc}") from exc
    finally:
        tmp_path.unlink(missing_ok=True)

    create_import_log(
        conn,
        imported_by=session["username"],
        hardware_count=summary["sheets"].get("硬體", {}).get("inserted", 0)
        + summary["sheets"].get("硬體", {}).get("updated", 0),
        personnel_count=summary["sheets"].get("人員", {}).get("inserted", 0)
        + summary["sheets"].get("人員", {}).get("updated", 0),
        software_count=summary["sheets"].get("軟體", {}).get("inserted", 0)
        + summary["sheets"].get("軟體", {}).get("updated", 0),
        error_count=len(summary["errors"]),
        source="cia_excel", file_name=file.filename,
    )
    return summary


RVTOOLS_LAST_DATA_AT = "rvtools_last_data_at"   # 上次成功匯入的那份檔案的「資料時間」


@app.post("/api/import/rvtools")
def import_rvtools_upload(
    file: UploadFile = File(...),
    data_at: str | None = Form(None),
    force: bool = Form(False),
    session: sqlite3.Row = Depends(require_auth),
    conn: sqlite3.Connection = Depends(get_db),
):
    """S19 VC 採集器：吃一份 RVTools 匯出的 vCenter 盤點（vInfo 分頁）。

    每台 VM 走身分解析：對到既有資產就更新機器事實（不碰業務欄位），新的建成 VC- 資產，
    判不準的進人工審核佇列（不自動合併）。真實 vCenter 資料，非假資料。

    ## 時序保護（data_at / force）

    匯入是「後蓋前」，沒有時間概念——不小心傳了一份舊的匯出檔，裡面的 os／is_vm
    就會直接蓋掉比較新的值，而且系統沒有欄位變更歷史，蓋掉就救不回來
    （只能翻每日備份，超過保留天數就沒了）。

    所以這裡收 data_at（前端帶上傳檔案的最後修改時間），跟上次成功匯入的時間比：
    比上次舊就先擋下來、回報兩邊時間讓人確認，確定要覆蓋才帶 force=true 再送一次。
    沒帶 data_at 就照舊直接匯（例如 curl 手動呼叫），不強制。
    """
    import rvtools_import
    from db import get_setting, set_setting

    if not file.filename or not file.filename.lower().endswith(".xlsx"):
        raise HTTPException(400, "只接受 RVTools 匯出的 .xlsx 檔案")

    prev_at = get_setting(conn, RVTOOLS_LAST_DATA_AT, "") or ""
    if data_at and prev_at and not force and data_at < prev_at:
        # 用 409 而不是 400：這不是「檔案有問題」，是「請你確認一下」，
        # 前端要據此跳確認框而不是當成錯誤紅字。
        raise HTTPException(
            409,
            {
                "code": "stale_import",
                "message": "這份檔案的資料時間比系統現有的舊，直接匯入會用舊資料覆蓋新的。",
                "incoming_data_at": data_at,
                "current_data_at": prev_at,
                "hint": "確定要覆蓋就再送一次並帶 force=true。",
            },
        )

    contents = file.file.read()
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        tmp.write(contents)
        tmp_path = Path(tmp.name)
    try:
        try:
            summary = rvtools_import.import_rvtools(tmp_path, conn)
        except ValueError as exc:
            # 「這份檔不對」——使用者自己就能判斷、能修（例如根本沒有 vInfo 分頁）
            raise HTTPException(400, f"匯入失敗，這份檔看起來不是 RVTools 匯出：{exc}") from exc
        except Exception as exc:  # noqa: BLE001
            # 其他例外是**我們自己的 bug**，不是使用者的檔案有問題。
            # 2026-08-20 踩過：五個貨真價實的 RVTools 匯出檔全部收到
            # 「請確認是 RVTools 匯出的檔：Object of type datetime is not JSON serializable」
            # ——訊息把責任推給使用者，害人去反覆檢查根本沒問題的檔案。
            # 分不出來就不要亂猜原因，如實說是系統的錯，並把錯誤原文給出來。
            raise HTTPException(
                500,
                f"匯入時系統發生錯誤（不是你的檔案有問題，請把這段訊息回報）："
                f"{type(exc).__name__}: {exc}",
            ) from exc
    finally:
        tmp_path.unlink(missing_ok=True)

    create_import_log(
        conn, imported_by=session["username"],
        hardware_count=summary["inserted"] + summary["updated"],
        personnel_count=0, software_count=0, error_count=len(summary["errors"]),
        source="rvtools", file_name=file.filename,
        # 這份是「哪天從 vCenter 匯出的」，不是「哪天匯進系統」。取不到就存 None，
        # 不拿匯入時間頂替——「不知道多舊」跟「是今天的」是完全不同的兩句話。
        exported_at=rvtools_import.export_time_from_filename(file.filename),
    )

    # 記下這份的資料時間，供下次比對。強制覆蓋舊資料時不要往回退，
    # 否則之後每一份新檔都會被拿來跟這個舊時間比，保護就失效了。
    if data_at and (not prev_at or data_at > prev_at):
        set_setting(conn, RVTOOLS_LAST_DATA_AT, data_at)
    summary["data_at"] = data_at or None
    summary["previous_data_at"] = prev_at or None
    return summary


@app.get("/api/import/log")
def list_import_log_endpoint(
    source: str | None = None, limit: int = 20,
    session: sqlite3.Row = Depends(require_auth),
    conn: sqlite3.Connection = Depends(get_db),
):
    """最近的匯入紀錄——2026-08-19 使用者原話「要有紀錄表讓我查」：一次要匯好幾份
    RVTools檔案時，需要看得出哪些已經匯過（檔名+時間），不用憑記憶判斷。
    source 篩選：rvtools/cia_excel/dynassets/scan_import；不帶就全部混著回。
    """
    if limit < 1 or limit > 100:
        raise HTTPException(400, "limit 只接受 1–100")
    return [dict(r) for r in list_import_log(conn, source=source, limit=limit)]


@app.post("/api/import/dynassets")
def import_dynassets_upload(
    file: UploadFile = File(...),
    session: sqlite3.Row = Depends(require_auth),
    conn: sqlite3.Connection = Depends(get_db),
):
    """dynassets 匯入：吃「存活掃描＋CMDB」清單（.csv／.xlsx），跟 CIA 登記對帳。

    每列走身分解析：對到既有資產就更新機器事實（不碰業務欄位）、掃到但沒登記的建成
    DYN- 資產（＝漏登記／帳外資產）、判不準的進人工審核佇列（不自動合併）。
    dynassets 是每次重新產生的存活快照，不做 RVTools 那種時序保護。
    """
    import dynassets_import

    fn = (file.filename or "").lower()
    if not (fn.endswith(".csv") or fn.endswith(".xlsx")):
        raise HTTPException(400, "只接受 dynassets 的 .csv 或 .xlsx 檔案")
    suffix = ".xlsx" if fn.endswith(".xlsx") else ".csv"

    contents = file.file.read()
    with tempfile.NamedTemporaryFile(suffix=suffix, delete=False) as tmp:
        tmp.write(contents)
        tmp_path = Path(tmp.name)
    try:
        try:
            summary = dynassets_import.import_dynassets(tmp_path, conn)
        except Exception as exc:  # noqa: BLE001 - 壞檔/非dynassets格式統一如實回報
            raise HTTPException(400, f"匯入失敗，請確認是 dynassets 存活清單：{exc}") from exc
    finally:
        tmp_path.unlink(missing_ok=True)

    create_import_log(
        conn, imported_by=session["username"],
        hardware_count=summary["inserted"] + summary["updated"],
        personnel_count=0, software_count=0, error_count=len(summary["errors"]),
        source="dynassets", file_name=file.filename,
    )
    return summary


# ===== 網段存活掃描（手動新增資產的第二種入口：整段掃、勾選要納入哪些）=====
# 只認 nmap TCP connect（-sT，不用 root）打 22/445——這兩個 port 對得上「這台機器
# 收得到 SSH 或 SMB」，是我們實際會去操作的存活定義，不是任意 ping 存活就算。
# 只掃內網私有網段，避免被拿去對外部位址做連接埠掃描。

_SCAN_PORTS = "22,445"
_SCAN_MAX_HOSTS = 1024


def _assert_private_network(network: "ipaddress.IPv4Network | ipaddress.IPv6Network") -> None:
    if not network.is_private:
        raise HTTPException(400, "只能掃內網私有網段（10/8、172.16/12、192.168/16）")


class ScanDiscoverBody(BaseModel):
    cidr: str


@app.post("/api/assets/scan/discover")
def scan_discover(
    body: ScanDiscoverBody,
    session: sqlite3.Row = Depends(require_auth),
    conn: sqlite3.Connection = Depends(get_db),
):
    """網段存活掃描：回傳掃到的主機清單（機器事實而已），不寫資料庫。
    使用者從結果裡勾選要納入的主機，再呼叫 /api/assets/scan/import 才真的建資產。
    """
    try:
        network = ipaddress.ip_network(body.cidr.strip(), strict=False)
    except ValueError:
        raise HTTPException(400, "不是合法的網段格式（例：192.168.1.0/24 或單一 IP）")
    _assert_private_network(network)
    if network.num_addresses > _SCAN_MAX_HOSTS:
        raise HTTPException(400, f"網段太大（上限 {_SCAN_MAX_HOSTS} 個位址），請縮小範圍")

    try:
        proc = subprocess.run(
            ["nmap", "-sT", "-p", _SCAN_PORTS, "--open", "-oG", "-", str(network)],
            capture_output=True, text=True, timeout=120,
        )
    except FileNotFoundError:
        raise HTTPException(500, "系統未安裝 nmap，無法掃描")
    except subprocess.TimeoutExpired:
        raise HTTPException(504, "掃描逾時，請縮小網段範圍再試")

    existing = {
        r["ip"]: r["asset_serial"]
        for r in conn.execute("SELECT ip, asset_serial FROM hardware WHERE ip IS NOT NULL")
    }
    found = []
    for line in proc.stdout.splitlines():
        m = re.match(r"Host:\s+(\S+)\s+\(([^)]*)\)\s+Ports:\s+(.*?)(?:\s+Ignored State:|$)", line)
        if not m:
            continue
        ip, rdns, ports_str = m.groups()
        open_ports = [p.split("/")[0] for p in ports_str.split(",") if "/open/" in p]
        if not open_ports:
            continue
        hostname = rdns or None
        if not hostname:
            try:
                hostname = socket.gethostbyaddr(ip)[0]
            except (socket.herror, socket.gaierror, OSError):
                hostname = None
        found.append({
            "ip": ip, "hostname": hostname, "open_ports": open_ports,
            "already_registered": ip in existing,
            "existing_asset_serial": existing.get(ip),
        })
    return {"cidr": str(network), "found": found}


class ScanHostItem(BaseModel):
    ip: str
    hostname: str | None = None


class ScanImportBody(BaseModel):
    hosts: list[ScanHostItem]


@app.post("/api/assets/scan/import")
def scan_import(
    body: ScanImportBody,
    session: sqlite3.Row = Depends(require_auth),
    conn: sqlite3.Connection = Depends(get_db),
):
    """把使用者勾選的掃描結果納入資產清單。只有機器事實（IP/Hostname），沒有業務
    欄位——借道既有 dynassets 匯入管道（identity 解析／DYN- 序號／source_record
    都直接沿用），不重造一套規則。業務欄位事後用 PUT /api/assets/{serial} 補。
    """
    if not body.hosts:
        raise HTTPException(400, "沒有選擇任何主機")

    with tempfile.NamedTemporaryFile(
        mode="w", suffix=".csv", delete=False, newline="", encoding="utf-8"
    ) as tmp:
        writer = csv.writer(tmp)
        writer.writerow(["IP", "Hostname"])
        for h in body.hosts:
            writer.writerow([h.ip, h.hostname or ""])
        tmp_path = Path(tmp.name)

    import dynassets_import
    try:
        summary = dynassets_import.import_dynassets(tmp_path, conn)
    finally:
        tmp_path.unlink(missing_ok=True)

    create_import_log(
        conn, imported_by=session["username"],
        hardware_count=summary["inserted"] + summary["updated"],
        personnel_count=0, software_count=0, error_count=len(summary["errors"]),
        source="scan_import",
    )
    return summary


# ===== 清空盤點資料（危險操作，打 ClearALL 才執行）=====

# 像 GitHub 刪 repo 那樣，要手動打這個字串才會清——防手滑。
_RESET_CONFIRM = "ClearALL"

# 只清「盤點／採集資料」。明確白名單，絕不誤清帳號/憑證/連線/設定。
# 用 defer_foreign_keys 把外鍵檢查延到 commit，全清後就一致，刪除順序不必講究。
_RESET_TABLES = (
    "merge_review", "source_record",
    "finding_disposition", "account_finding", "account_collect_runs", "host_account",
    "service_collect_runs", "host_service",
    "comparison_result", "scan_history", "scan_runs", "onboard_audit",
    "system_deps", "systems",
    "duplicate_dismiss",
    "software", "personnel", "hardware",
    "import_log",
)


class ResetConfirm(BaseModel):
    confirm: str


@app.post("/api/admin/reset-inventory")
def reset_inventory(
    body: ResetConfirm,
    session: sqlite3.Row = Depends(require_auth),
    conn: sqlite3.Connection = Depends(get_db),
):
    """清空所有盤點／採集資料，回到「有設定、無資料」的空白，供重新匯入。

    ⚠️ 破壞性且不可逆（只能靠備份還原）。必須 confirm == "ClearALL" 才執行。
    清：資產/人員/軟體、來源紀錄、待審、帳號盤點、服務、掃描、業務系統、匯入紀錄。
    不動：登入帳號(users)、憑證庫(collect_credential)、連線設定、授權網段、
         功能開關、系統設定(app_settings)、正規化別名。
    """
    if body.confirm != _RESET_CONFIRM:
        raise HTTPException(400, f'要清空必須輸入「{_RESET_CONFIRM}」確認（防手滑），已取消。')
    conn.execute("PRAGMA defer_foreign_keys=ON")  # 外鍵檢查延到 commit → 刪除順序不重要
    deleted: dict[str, int] = {}
    for t in _RESET_TABLES:
        deleted[t] = conn.execute(f"DELETE FROM {t}").rowcount
    conn.commit()
    return {"ok": True, "cleared": deleted, "total_rows": sum(deleted.values()),
            "by": session["username"]}


@app.get("/api/merge-reviews")
def list_merge_reviews(
    session: sqlite3.Row = Depends(require_auth), conn: sqlite3.Connection = Depends(get_db)
):
    """待人工判定的合併案（身分解析判不準的）。空＝沒有懸而未決的。

    這些是「同 IP 但不同 UUID」這種不敢自動合併的案子——攤開來讓人決定，
    不要藏著，否則等於資料默默對不上卻沒人知道。
    """
    rows = conn.execute(
        "SELECT mr.id, mr.reason, mr.candidates, mr.status, mr.created_at, "
        "sr.source, sr.payload FROM merge_review mr "
        "JOIN source_record sr ON sr.id = mr.source_record_id "
        "WHERE mr.status = 'open' ORDER BY mr.id DESC LIMIT 200").fetchall()
    out = []
    for r in rows:
        d = dict(r)
        try:
            d["candidates"] = json.loads(d["candidates"]) if d["candidates"] else []
            d["payload"] = json.loads(d["payload"]) if d["payload"] else {}
        except (ValueError, TypeError):
            pass
        out.append(d)
    return out


def _short_host(name: str | None) -> str:
    """FQDN 取第一段並轉大寫。

    vCenter 回報的是 FQDN（srv-abc-1234.corp.example.com），CIA 的 Excel 是人填的短名
    且慣例全大寫（SRV-ABC-1234）。兩邊指同一台機器，卻因為一個有網域後綴、一個沒有，
    在身分解析時對不上而全部落到人工審核。切掉網域再統一大小寫才比得出來。
    """
    return (name or "").strip().split(".")[0].upper()


# 可自動合併的判定規則：短名與 IP 同時相符。
# 為什麼這個組合敢自動合併：
#   · 主機名相符 → 指向同一台的可能性高，但主機名可能重複（dev/prod 同名）
#   · IP 相符   → 也高，但 IP 會被回收再指派
#   兩者「同時」相符，還要落在同一筆既有資產上，才算數。任一項不符就不碰，留給人看。
# 這不是最強識別碼（vm_uuid），所以仍然要人按一次確認鈕，只是不必逐筆按 494 次。
_AUTO_RULE = "short_name_and_ip"


def _scan_auto_matchable(conn: sqlite3.Connection) -> list[dict]:
    """掃出「切掉網域後主機名與 IP 同時相符」的待審案，回傳配對結果。

    每次都重新從資料庫算，不信任前端傳來的 id 清單——避免畫面上看到的候選
    與實際執行時的資料已經不同（例如中間有人改了資產）。
    """
    hw = conn.execute(
        "SELECT id, asset_serial, hostname, ip FROM hardware "
        "WHERE hostname IS NOT NULL AND length(trim(hostname)) > 0 "
        "AND ip IS NOT NULL AND length(trim(ip)) > 0"
    ).fetchall()
    # (短名, IP) -> 既有資產。同鍵有多筆就整組排除：那代表既有資料本身重複，
    # 亂挑一筆合併會把資料釘在錯的那台上。
    index: dict[tuple[str, str], list[sqlite3.Row]] = {}
    for r in hw:
        index.setdefault((_short_host(r["hostname"]), (r["ip"] or "").strip()), []).append(r)

    rows = conn.execute(
        "SELECT mr.id, mr.reason, sr.id AS sr_id, sr.payload FROM merge_review mr "
        "JOIN source_record sr ON sr.id = mr.source_record_id "
        "WHERE mr.status = 'open'"
    ).fetchall()

    out = []
    for r in rows:
        try:
            d = json.loads(r["payload"]) if r["payload"] else {}
        except (ValueError, TypeError):
            continue
        h = (d.get("hostname") or "").strip()
        ip = (d.get("ip") or "").strip()
        if not h or not ip:
            continue
        hits = index.get((_short_host(h), ip)) or []
        if len(hits) != 1:      # 0 筆＝對不上；>1 筆＝既有資料重複，都不自動處理
            continue
        target = hits[0]
        out.append({
            "review_id": r["id"],
            "source_record_id": r["sr_id"],
            "incoming_hostname": h,
            "incoming_ip": ip,
            "target_id": target["id"],
            "target_serial": target["asset_serial"],
            "target_hostname": target["hostname"],
            "payload": d,
        })
    return out


@app.get("/api/merge-reviews/auto-matchable")
def merge_reviews_auto_matchable(
    session: sqlite3.Row = Depends(require_auth), conn: sqlite3.Connection = Depends(get_db)
):
    """待審案裡有多少可以用「短名＋IP 同時相符」自動配對。

    1,004 筆待審一筆筆看是做不完的，而其中大半只是「vCenter 給 FQDN、Excel 填短名」
    這一個原因。先算出可安全批次處理的數量，剩下的才值得花人力逐筆判斷。
    """
    total_open = conn.execute(
        "SELECT COUNT(*) FROM merge_review WHERE status = 'open'"
    ).fetchone()[0]
    m = _scan_auto_matchable(conn)
    return {
        "total_open": total_open,
        "matchable": len(m),
        "remaining": total_open - len(m),
        "rule": _AUTO_RULE,
        "rule_label": "切掉網域後主機名與 IP 同時相符",
        # 給畫面看幾筆長什麼樣，確認規則沒抓錯再按執行
        "samples": [
            {
                "incoming_hostname": x["incoming_hostname"],
                "incoming_ip": x["incoming_ip"],
                "target_serial": x["target_serial"],
                "target_hostname": x["target_hostname"],
            }
            for x in m[:10]
        ],
    }


class BatchMergeBody(BaseModel):
    rule: str


@app.post("/api/merge-reviews/batch-merge")
def merge_reviews_batch_merge(
    body: BatchMergeBody,
    session: sqlite3.Row = Depends(require_auth),
    conn: sqlite3.Connection = Depends(get_db),
):
    """把可自動配對的待審案一次併進既有資產。

    只接受預先定義好的規則名稱，不接受前端傳 id 清單——後者等於把「要合併哪些」
    的決定權交給畫面，一旦畫面算錯就會把資料併到錯的機器上，而合併錯不會噴錯、很難救。

    合併時**只寫機器事實**（os／vm_uuid／is_vm／hostname），業務欄位（用途、保管者、
    機房、盤點單位）一律不動——那些是人維護的，vCenter 不知道也無權覆蓋。
    """
    if body.rule != _AUTO_RULE:
        raise HTTPException(400, f"不支援的規則：{body.rule}")

    matches = _scan_auto_matchable(conn)
    FACTS = ("os", "vm_uuid", "is_vm")
    merged = 0
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    for m in matches:
        d = m["payload"]
        setters = {k: d[k] for k in FACTS if d.get(k) not in (None, "")}
        if setters:
            assigns = ", ".join(f"{k} = ?" for k in setters)
            conn.execute(
                f"UPDATE hardware SET {assigns}, updated_at = ? WHERE id = ?",
                (*setters.values(), now, m["target_id"]),
            )
        # 來源紀錄回填解析結果，之後查得出這筆是靠哪條規則併上去的
        conn.execute(
            "UPDATE source_record SET resolved_status = 'matched', resolved_hardware_id = ?, "
            "resolved_rule = ?, resolved_confidence = ? WHERE id = ?",
            (m["target_id"], f"batch:{_AUTO_RULE}", 0.9, m["source_record_id"]),
        )
        conn.execute(
            "UPDATE merge_review SET status = 'merged', decided_by = ?, decided_at = ? "
            "WHERE id = ? AND status = 'open'",
            (session["username"], now, m["review_id"]),
        )
        merged += 1
    conn.commit()

    remaining = conn.execute(
        "SELECT COUNT(*) FROM merge_review WHERE status = 'open'"
    ).fetchone()[0]
    return {"merged": merged, "remaining_open": remaining, "rule": _AUTO_RULE}


class VcAutoImportBody(BaseModel):
    enabled: bool
    dir: str = ""
    max_age_hours: int = 36


@app.get("/api/vcenter-autoimport")
def vcenter_autoimport_status(
    session: sqlite3.Row = Depends(require_auth), conn: sqlite3.Connection = Depends(get_db)
):
    """VC 自動匯入（方案 B）現況：設定＋鮮度燈（今晚的匯出到底有沒有進來）。"""
    import vcenter_autoimport

    return vcenter_autoimport.health(conn)


@app.put("/api/vcenter-autoimport")
def vcenter_autoimport_save(
    body: VcAutoImportBody,
    session: sqlite3.Row = Depends(require_auth),
    conn: sqlite3.Connection = Depends(get_db),
):
    """設定監看資料夾＋開關＋逾時門檻。開了之後排程器每輪會看資料夾有沒有新檔。"""
    import vcenter_autoimport

    vcenter_autoimport.set_config(conn, body.enabled, body.dir, body.max_age_hours)
    return vcenter_autoimport.health(conn)


@app.post("/api/vcenter-autoimport/run")
def vcenter_autoimport_run(
    session: sqlite3.Row = Depends(require_auth), conn: sqlite3.Connection = Depends(get_db)
):
    """手動「立即抓一次」：找監看資料夾最新且寫完的 RVTools 檔，比上次新才匯。"""
    import vcenter_autoimport

    return vcenter_autoimport.pickup(conn)


def _serialize_connection(row: sqlite3.Row) -> dict:
    """密碼write-only：HTTP回應一律不帶password欄位本身，只給has_password布林值讓畫面
    知道「這筆有沒有設定過密碼」，不洩漏內容（S10 done_when明訂）。
    """
    d = dict(row)
    has_password = bool(d.pop("password", None))
    d["has_password"] = has_password
    return d


class ConnectionBody(BaseModel):
    name: str
    connection_type: str | None = None
    target: str
    port: int | None = None
    username: str | None = None
    password: str | None = None


@app.get("/api/connections")
def list_connections_endpoint(
    session: sqlite3.Row = Depends(require_auth), conn: sqlite3.Connection = Depends(get_db)
):
    return [_serialize_connection(r) for r in list_connections(conn)]


@app.get("/api/connections/suggest-segments")
def suggest_segments(
    fmt: str = Query("json", pattern="^(json|txt)$"),
    session: sqlite3.Row = Depends(require_auth),
    conn: sqlite3.Connection = Depends(get_db),
):
    """從既有資產的 IP 反推該掃描哪些網段，並對上機房。

    fmt=txt 回純文字的 segments.txt，直接餵給外部掃描機的 scan_segments.py。
    那台連不到這裡的資料庫（正是要外部掃描的原因），清單只能由人搬過去。

    為什麼要有這支：資產清單是匯入進來的（幾千筆），但掃描目標得一個一個手動建，
    只要漏建，那個網段的機器就全部變成「失聯」——看起來像機器出事，其實只是沒去掃。
    網段資訊本來就藏在資產的 IP 裡，不該再要人另外整理一份網段表。

    每個網段會標出資產分布在哪些機房。同一網段對到多個機房時一併回報：
    那要嘛是真的跨機房共用，要嘛是資產的機房欄位填錯——後者正是盤點系統該抓的問題。
    """
    import manage_state

    existing = set()
    for c in conn.execute("SELECT target FROM connections").fetchall():
        t = (c["target"] or "").strip()
        if t:
            existing.add(t)

    segs: dict[str, dict] = {}
    bad_ip = 0
    for r in conn.execute(
        "SELECT ip, physical_location FROM hardware WHERE ip IS NOT NULL AND ip != ''"
    ).fetchall():
        parts = (r["ip"] or "").strip().split(".")
        if len(parts) != 4 or not all(p.isdigit() and 0 <= int(p) <= 255 for p in parts):
            bad_ip += 1
            continue
        cidr = ".".join(parts[:3]) + ".0/24"
        s = segs.setdefault(cidr, {"cidr": cidr, "hosts": 0, "locations": {}})
        s["hosts"] += 1
        loc = manage_state.group_location(r["physical_location"])
        s["locations"][loc] = s["locations"].get(loc, 0) + 1

    out = []
    for s in segs.values():
        locs = sorted(s["locations"].items(), key=lambda kv: -kv[1])
        main_loc = locs[0][0] if locs else "未填"
        out.append({
            "cidr": s["cidr"],
            "hosts": s["hosts"],
            "main_location": main_loc,
            "locations": dict(locs),
            "mixed": len(locs) > 1,
            "already": s["cidr"] in existing,
            # 建議的連線名稱：帶上機房，之後在清單裡一眼看得出這條是掃哪裡的
            "suggest_name": f"{main_loc} {s['cidr']}",
        })
    out.sort(key=lambda x: -x["hosts"])

    if fmt == "txt":
        # 只放整行註解、不放行尾註解：scan_segments.py 的 load_segments 只跳過整行 #，
        # 行尾註解會讓整行解析失敗被略過——那會變成「檔案看起來有 129 段、實際掃 0 段」。
        # 機房資訊是給人看的，UI 上就有，不必塞進要餵給機器的檔。
        covered = sum(s["hosts"] for s in out)
        lines = [
            "# 網段掃描清單 —— 由戰情室依現有資產 IP 反推（每筆取前三段成 /24）",
            f"# 產生時間：{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
            f"# 網段數：{len(out)}　涵蓋資產：{covered} 台",
        ]
        if bad_ip:
            # 這些資產不屬於任何網段，掃描永遠看不到它們。不講的話，
            # 掃完發現「還有機器沒出現」會回頭懷疑掃描器，其實是資料本身有問題。
            lines.append(
                f"# ⚠ 另有 {bad_ip} 筆資產的 IP 格式不正確，未納入任何網段——"
                "掃描不會涵蓋它們，請在系統裡修正後重新下載"
            )
        lines += [
            "#",
            "# 一行一個網段，# 開頭為註解。要增減網段直接改這個檔即可。",
            "",
        ]
        lines += [s["cidr"] for s in out]
        # 檔名固定 segments.txt，不加時間戳：mmap.sh 只認這個檔名，
        # 帶戳記就得人工改名，漏改會變成「找不到網段清單」而中止。
        # 產生時間寫在檔頭註解裡，要區分版本看那裡。
        return Response(
            content="\n".join(lines) + "\n",
            media_type="text/plain; charset=utf-8",
            headers={"Content-Disposition": 'attachment; filename="segments.txt"'},
        )

    return {"segments": out, "bad_ip_count": bad_ip, "existing_count": len(existing)}


class SegmentBatchBody(BaseModel):
    cidrs: list[str]


@app.post("/api/connections/batch-segments")
def batch_create_segments(
    body: SegmentBatchBody,
    session: sqlite3.Row = Depends(require_auth),
    conn: sqlite3.Connection = Depends(get_db),
):
    """把選定的網段一次建成掃描來源。已存在的略過，重複點不會長出一堆重複設定。"""
    import manage_state

    existing = {
        (c["target"] or "").strip()
        for c in conn.execute("SELECT target FROM connections").fetchall()
    }

    # 先算好每個網段的主要機房，名稱才帶得出「板橋 10.92.198.0/24」
    loc_of: dict[str, str] = {}
    counter: dict[str, dict[str, int]] = {}
    for r in conn.execute(
        "SELECT ip, physical_location FROM hardware WHERE ip IS NOT NULL AND ip != ''"
    ).fetchall():
        parts = (r["ip"] or "").strip().split(".")
        if len(parts) != 4 or not all(p.isdigit() for p in parts):
            continue
        cidr = ".".join(parts[:3]) + ".0/24"
        loc = manage_state.group_location(r["physical_location"])
        counter.setdefault(cidr, {})
        counter[cidr][loc] = counter[cidr].get(loc, 0) + 1
    for cidr, d in counter.items():
        loc_of[cidr] = max(d.items(), key=lambda kv: kv[1])[0]

    created, skipped = [], []
    for cidr in body.cidrs:
        cidr = (cidr or "").strip()
        if not cidr:
            continue
        if cidr in existing:
            skipped.append(cidr)
            continue
        name = f"{loc_of.get(cidr, '未知')} {cidr}"
        create_connection_record(conn, name, "網路掃描", cidr, None, None, None)
        existing.add(cidr)
        created.append({"cidr": cidr, "name": name})
    return {"created": created, "skipped": skipped}


@app.post("/api/connections")
def create_connection_endpoint(
    body: ConnectionBody,
    session: sqlite3.Row = Depends(require_auth),
    conn: sqlite3.Connection = Depends(get_db),
):
    new_id = create_connection_record(
        conn, body.name, body.connection_type, body.target, body.port, body.username, body.password
    )
    return _serialize_connection(get_connection_by_id(conn, new_id))


@app.put("/api/connections/{connection_id}")
def update_connection_endpoint(
    connection_id: int,
    body: ConnectionBody,
    session: sqlite3.Row = Depends(require_auth),
    conn: sqlite3.Connection = Depends(get_db),
):
    if get_connection_by_id(conn, connection_id) is None:
        raise HTTPException(404, "查無此連線設定")
    # 密碼write-only的另一半：留空（None或空字串）代表「這次不變更密碼」，不是「清空密碼」。
    password_to_set = body.password if body.password else None
    update_connection_record(
        conn,
        connection_id,
        body.name,
        body.connection_type,
        body.target,
        body.port,
        body.username,
        password_to_set,
    )
    return _serialize_connection(get_connection_by_id(conn, connection_id))


class ConnectionEnabledBody(BaseModel):
    enabled: bool


@app.patch("/api/connections/{connection_id}/enabled")
def set_connection_enabled_endpoint(
    connection_id: int,
    body: ConnectionEnabledBody,
    session: sqlite3.Row = Depends(require_auth),
    conn: sqlite3.Connection = Depends(get_db),
):
    """來源啟用／停用。停用的來源排程掃描會直接跳過，不會被算成「掃描失敗」。

    存在的理由：有些來源現階段本來就連不到（例如 CMDB Gateway 在家裡碰不到公司內網）。
    沒有開關的話它每次掃描都失敗、每次都把「掃描不完整」點亮——常態性的假警報
    會讓真正的掃描問題沒人看見。
    """
    if get_connection_by_id(conn, connection_id) is None:
        raise HTTPException(404, "查無此連線設定")
    set_connection_enabled(conn, connection_id, body.enabled)
    return _serialize_connection(get_connection_by_id(conn, connection_id))


@app.delete("/api/connections/{connection_id}")
def delete_connection_endpoint(
    connection_id: int,
    session: sqlite3.Row = Depends(require_auth),
    conn: sqlite3.Connection = Depends(get_db),
):
    if get_connection_by_id(conn, connection_id) is None:
        raise HTTPException(404, "查無此連線設定")
    delete_connection_record(conn, connection_id)
    return {"ok": True}


@app.post("/api/connections/{connection_id}/test")
def test_connection_endpoint(
    connection_id: int,
    session: sqlite3.Row = Depends(require_auth),
    conn: sqlite3.Connection = Depends(get_db),
):
    """D15「重新檢查」按鈕：真的做一次TCP連線測試，不是硬編一個假狀態。這個階段還沒有
    真實vCenter/SNMP協定整合（S3的ScanSource還是mock，接上真實協定是之後的工作），但
    port通不通得起來是真訊號，比起顯示假的「已連線」更誠實，符合D31精神。
    """
    row = get_connection_by_id(conn, connection_id)
    if row is None:
        raise HTTPException(404, "查無此連線設定")
    if not row["port"]:
        raise HTTPException(400, "這筆連線沒有設定Port，無法測試")
    try:
        with socket.create_connection((row["target"], row["port"]), timeout=3):
            status = "綠"
    except OSError:
        status = "紅"
    update_connection_status(conn, connection_id, status)
    return _serialize_connection(get_connection_by_id(conn, connection_id))


# ===== S14 備份健康儀表 =====
# 目的是讓工程師/助理不用進命令列就能看狀態、手動備份。備份這件事最糟的失敗模式是
# 「以為有備份，還原時才發現壞的」，所以這裡不只顯示「有沒有跑過」，而是實際去驗證。

@app.get("/api/backup/status")
def backup_status_endpoint(
    session: sqlite3.Row = Depends(require_auth), conn: sqlite3.Connection = Depends(get_db)
):
    """燈號 + 明細。會實際開檔跑 integrity_check，不是只看檔案存不存在。"""
    return backup.health(conn=conn)


class OffsiteBody(BaseModel):
    dir: str = ""


@app.put("/api/backup/offsite")
def backup_set_offsite(
    body: OffsiteBody,
    session: sqlite3.Row = Depends(require_auth),
    conn: sqlite3.Connection = Depends(get_db),
):
    """設定異地備份目錄（畫面可設，不用改 systemd 環境變數再重啟）。

    這裡只存路徑；路徑本身要指到「真正獨立的儲存」才算異地——例如掛載另一台機器
    （222）的 /ai_backup。留空＝清掉異地（回黃燈）。存完立刻回最新健康狀態讓畫面看變化。
    """
    from db import set_setting

    set_setting(conn, backup.OFFSITE_SETTING_KEY, (body.dir or "").strip())
    return backup.health(conn=conn)


@app.post("/api/backup/run")
def backup_run_endpoint(
    session: sqlite3.Row = Depends(require_auth), conn: sqlite3.Connection = Depends(get_db)
):
    """手動觸發一次備份（安全快照 + 完整性驗證 + 異地複製 + 清理過期）。

    刻意用 sync def：跟 import_excel_upload 同理，備份是 I/O 密集的阻塞操作，
    交給 FastAPI 的 threadpool 跑，不要卡住 event loop。
    """
    db_path = get_db_path()
    result = backup.run_backup(
        db_path, backup.get_backup_dir(db_path), datetime.now(), backup.get_offsite_dir(conn)
    )
    payload = {
        "ok": result.ok,
        "path": str(result.path) if result.path else None,
        "size_bytes": result.size_bytes,
        "integrity_ok": result.integrity_ok,
        "integrity_detail": result.integrity_detail,
        "offsite_path": str(result.offsite_path) if result.offsite_path else None,
        "offsite_error": result.offsite_error,
        "pruned_count": len(result.pruned),
        "took_seconds": result.took_seconds,
        "error": result.error,
    }
    if not result.ok:
        # 用 200 回失敗細節而不是丟 500：前端要能把「為什麼失敗」原樣顯示給使用者，
        # 500 只會變成一句「伺服器錯誤」，等於把診斷資訊丟掉。
        payload["message"] = f"備份失敗：{result.error}"
    return payload


@app.get("/api/backup/list")
def backup_list_endpoint(
    session: sqlite3.Row = Depends(require_auth), conn: sqlite3.Connection = Depends(get_db)
):
    """本地與異地的備份檔清單（新到舊）。"""
    db_path = get_db_path()
    offsite = backup.get_offsite_dir(conn)
    return {
        "local": backup.list_backups(backup.get_backup_dir(db_path)),
        "offsite": backup.list_backups(offsite) if offsite else [],
    }


@app.get("/api/backup/dump")
def backup_dump_endpoint(
    fmt: str = "binary",
    session: sqlite3.Row = Depends(require_auth),
    conn: sqlite3.Connection = Depends(get_db),
):
    """匯出全部資料庫，供備份／搬機使用。跟 /api/backup/run（存到伺服器本機/異地）不同，
    這支是**直接下載到使用者瀏覽器**，兩種格式讓使用者自己選：

    fmt=binary（預設）：跟 backup.snapshot() 同一套 VACUUM INTO，二進位 .db 檔，
      跟既有系統/任何 SQLite 工具都能直接讀，不會被誤貼到聊天室/網頁（2026-08-18
      實際發生過 RVTools 資料被貼到公開分享碼網站，這支刻意提供二進位選項）。
    fmt=sql：純文字 SQL dump（用 sqlite3 內建的 iterdump()，不шell out），
      體積小、人可讀，但也因為是文字，複製貼上風險比二進位檔案高，使用者自己權衡。

    這支只匯出。還原是獨立端點 /api/backup/restore（2026-08-19 拍板方案B新增），
    刻意不跟這支共用——風險層級不同，還原會覆蓋正式資料。
    """
    import tempfile

    db_path = get_db_path()
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")

    if fmt == "sql":
        def sql_lines():
            for line in conn.iterdump():
                yield line + "\n"
        return StreamingResponse(
            sql_lines(), media_type="application/sql",
            headers={"Content-Disposition": f'attachment; filename="asset_dump_{stamp}.sql"'},
        )

    if fmt != "binary":
        raise HTTPException(400, "fmt 只接受 binary 或 sql")

    with tempfile.TemporaryDirectory() as tmp:
        dest = Path(tmp) / f"asset_dump_{stamp}.db"
        backup.snapshot(db_path, dest)
        ok, msg = backup.verify_integrity(dest)
        if not ok:
            raise HTTPException(500, f"匯出後完整性檢查失敗，拒絕提供下載：{msg}")
        # 讀進記憶體再回傳（不是回傳暫存檔路徑）——FileResponse 是回應送出「後」才真的去
        # 讀檔案，但那時 with 區塊已經把暫存目錄清掉了，會踩到檔案不存在。資料庫只有幾MB，
        # 讀進記憶體沒有負擔。
        content = dest.read_bytes()
    return Response(
        content, media_type="application/octet-stream",
        headers={"Content-Disposition": f'attachment; filename="asset_dump_{stamp}.db"'},
    )


@app.post("/api/backup/restore")
def backup_restore_endpoint(
    file: UploadFile = File(...),
    session: sqlite3.Row = Depends(require_auth),
):
    """整庫覆蓋還原（2026-08-19 拍板方案B：簡單覆蓋＋前端單一確認框，開關預設關）。

    刻意不用 `Depends(get_db)`——這支要整檔替換掉資料庫本體，不能借用一條可能還開著
    讀取交易的連線；讓 backup.restore() 自己開短命連線做驗證，本體替換用檔案系統層級的
    os.replace()。前端是否顯示這個功能由 feature_flags 的 'restore' 開關決定
    （查 GET /api/feature-flags），這支本身不重複檢查開關——已登入使用者手動點開開關、
    走到這支端點，就是已經確認過要做這件事。
    """
    if not file.filename or not file.filename.lower().endswith(".db"):
        raise HTTPException(400, "只接受 .db 二進位資料庫檔（用「匯出全部資料」的二進位格式匯出的那份）")

    contents = file.file.read()
    if not contents:
        raise HTTPException(400, "上傳的檔案是空的（0 bytes）")
    try:
        return backup.restore(contents, get_db_path())
    except ValueError as exc:
        # 「你給的檔案不對」——使用者自己就能修
        raise HTTPException(400, str(exc)) from exc
    except backup.RestoreFailed as exc:
        # 伺服器端出錯，但已經寫成人看得懂、且講明資料有沒有被動到的訊息
        raise HTTPException(500, str(exc)) from exc
    except Exception as exc:  # noqa: BLE001
        # 最後一道網。沒有這段，任何沒預期到的例外都會變成 FastAPI 的裸 500
        # `{"detail":"Internal Server Error"}`，使用者在一個**不可逆**的操作上
        # 只看得到「失敗」兩個字——2026-08-19 回報的就是這個。
        raise HTTPException(
            500,
            f"還原時發生未預期的錯誤：{type(exc).__name__}: {exc}。"
            f"若失敗發生在覆蓋開始前，正式資料庫未被更動；"
            f"請檢查伺服器紀錄，或用 backups/ 目錄裡的存證備份還原。",
        ) from exc


# ===== 系統組月報（2026-08-21）=====
# 薄 HTTP 層，邏輯都在 system_report.py。使用者每月要把三張表貼進部門報告，
# 原本是人工統計的——他的原話：「以後我就 COPY 畫面，不用再自己統計」。

@app.get("/api/reports/system-group")
def report_system_group(
    period: str | None = None,
    session: sqlite3.Row = Depends(require_auth),
    conn: sqlite3.Connection = Depends(get_db),
):
    """當下即時算的整份報告。帶 period 只影響備註的取用範圍，數字一律是即時的
    ——要看某個月的定稿數字請走 /snapshots/{period}，兩者刻意分開，
    免得有人以為畫面上的數字就是當月存檔的那份。"""
    import system_report

    return system_report.build(conn, period)


class ReportNoteBody(BaseModel):
    row_key: str
    note: str
    period: str | None = None


@app.put("/api/reports/system-group/note")
def report_system_group_note(
    body: ReportNoteBody,
    session: sqlite3.Row = Depends(require_auth),
    conn: sqlite3.Connection = Depends(get_db),
):
    import system_report

    system_report.set_note(conn, body.row_key, body.note, session["username"], body.period)
    return {"ok": True}


@app.get("/api/reports/system-group/drill")
def report_drill(
    table: str,
    platform: str | None = None,
    status: str | None = None,
    bucket: str | None = None,
    retired: bool = False,
    location: str | None = None,
    service: str | None = None,
    vcenter: str | None = None,
    cluster: str | None = None,
    version: str | None = None,
    os_canonical: str | None = None,
    session: sqlite3.Row = Depends(require_auth),
    conn: sqlite3.Connection = Depends(get_db),
):
    """任何一格數字點下去要看的清單。

    使用者 2026-08-21：「你每個數字我都要可以追」。這支跟加總走**同一份**計算
    （system_report 的 classify_assets / _vhosts），所以格子上的數字跟這裡回的
    筆數必然一致——不一致的話那張報表就不能用了。

    每一列都帶 `reason`：不只列得出是哪幾台，還講得出為什麼是這一台。
    """
    import system_report

    if table == "platform":
        return system_report.drill_platform(
            conn, platform=platform, status=status, bucket=bucket, retired=retired,
            os_canonical=os_canonical)
    if table in ("cluster", "virtualization"):
        return system_report.drill_cluster(
            conn, location=location, service=service, vcenter=vcenter, cluster=cluster,
            version=version)
    raise HTTPException(400, "table 只接受 platform / cluster / virtualization")


@app.get("/api/reports/system-group/snapshots")
def report_snapshots(
    session: sqlite3.Row = Depends(require_auth),
    conn: sqlite3.Connection = Depends(get_db),
):
    import system_report

    return system_report.list_snapshots(conn)


class ReportSnapshotBody(BaseModel):
    period: str


@app.post("/api/reports/system-group/snapshots")
def report_snapshot_save(
    body: ReportSnapshotBody,
    session: sqlite3.Row = Depends(require_auth),
    conn: sqlite3.Connection = Depends(get_db),
):
    """存當月定稿。同月重存＝覆蓋（同月本來就只該有一份定稿）。"""
    import re as _re

    import system_report

    if not _re.fullmatch(r"\d{4}-\d{2}", body.period or ""):
        raise HTTPException(400, "period 格式須為 YYYY-MM")
    sid = system_report.save_snapshot(conn, body.period, session["username"])
    return {"id": sid, "period": body.period}


@app.get("/api/reports/system-group/snapshots/{period}")
def report_snapshot_get(
    period: str,
    session: sqlite3.Row = Depends(require_auth),
    conn: sqlite3.Connection = Depends(get_db),
):
    import system_report

    snap = system_report.get_snapshot(conn, period)
    if snap is None:
        raise HTTPException(404, f"沒有 {period} 的存檔")
    return snap


# ===== CI 圖譜（MICS 切片1：影響範圍查詢的地基）=====
# 薄 HTTP 層，邏輯都在 ci_graph.py。rebuild() 同步跑（資料量數百~數千節點，純 DB
# 運算不碰網路，不用像 scan_service 那樣起背景執行緒）——但仍照 scan_runs 的慣例留
# running 狀態擋同時觸發兩次，避免兩個請求並發時互相踩。
#
# run 記錄（開 running 列、成功寫 counts、失敗寫 error）在 ci_graph.run_rebuild()，
# 不寫在這裡：每天凌晨的排程要做完全一樣的事，兩邊各寫一份必然漂走。

@app.post("/api/ci/rebuild")
def ci_rebuild_endpoint(
    session: sqlite3.Row = Depends(require_auth),
    conn: sqlite3.Connection = Depends(get_db),
):
    import ci_graph

    try:
        return ci_graph.run_rebuild(conn, "manual", session["username"])
    except ci_graph.RebuildInProgress as exc:
        raise HTTPException(409, "已有一次重建正在進行，請稍候") from exc
    except Exception as exc:  # noqa: BLE001 - 失敗已由 run_rebuild 記進 ci_graph_runs
        raise HTTPException(500, f"重建失敗：{exc}") from exc


class CiScheduleBody(BaseModel):
    enabled: bool
    time: str


@app.get("/api/ci/schedule")
def ci_schedule_get_endpoint(
    session: sqlite3.Row = Depends(require_auth),
    conn: sqlite3.Connection = Depends(get_db),
):
    import ci_graph

    return ci_graph.get_schedule(conn)


@app.put("/api/ci/schedule")
def ci_schedule_put_endpoint(
    body: CiScheduleBody,
    session: sqlite3.Row = Depends(require_auth),
    conn: sqlite3.Connection = Depends(get_db),
):
    import ci_graph

    try:
        ci_graph.set_schedule(conn, body.enabled, body.time)
    except ValueError as exc:
        raise HTTPException(400, str(exc)) from exc
    return ci_graph.get_schedule(conn)


@app.get("/api/ci/rebuild/status")
def ci_rebuild_status_endpoint(
    session: sqlite3.Row = Depends(require_auth),
    conn: sqlite3.Connection = Depends(get_db),
):
    row = conn.execute("SELECT * FROM ci_graph_runs ORDER BY id DESC LIMIT 1").fetchone()
    return dict(row) if row else None


# ===== 影響範圍查詢（MICS 切片2：Blast Radius，殺手級功能）=====
# 薄 HTTP 層，邏輯都在 blast_radius.py。三種問法（陌生IP研判/事故爆炸半徑/計畫性停機）
# 共用同一顆引擎，只差 mode 參數，見 blast_radius.py 模組開頭說明。

@app.get("/api/blast/systems")
def blast_systems_endpoint(
    session: sqlite3.Row = Depends(require_auth),
    conn: sqlite3.Connection = Depends(get_db),
):
    """瀏覽清單：全庫5148節點畫成一張圖是看不出結構的毛球，2026-08-19 使用者拍板
    用瀏覽清單當「不知道要查什麼名字」時的入口，不做全部關聯圖。"""
    import blast_radius
    return blast_radius.list_business_systems(conn)


@app.get("/api/blast/resolve")
def blast_resolve_endpoint(
    q: str, port: int | None = None,
    session: sqlite3.Row = Depends(require_auth),
    conn: sqlite3.Connection = Depends(get_db),
):
    import blast_radius
    return blast_radius.resolve(conn, q, port)


@app.get("/api/blast/impact")
def blast_impact_endpoint(
    node_id: str, depth: int = 6, mode: str = "incident", only_evidence: bool = False,
    session: sqlite3.Row = Depends(require_auth),
    conn: sqlite3.Connection = Depends(get_db),
):
    import blast_radius
    try:
        return blast_radius.impact(conn, node_id, depth=depth, mode=mode, only_evidence=only_evidence)
    except ValueError as exc:
        raise HTTPException(404, str(exc)) from exc


@app.get("/api/blast/graph")
def blast_graph_endpoint(
    node_id: str, depth: int = 3, direction: str = "dependents",
    session: sqlite3.Row = Depends(require_auth),
    conn: sqlite3.Connection = Depends(get_db),
):
    import blast_radius
    return blast_radius.graph_elements(conn, node_id, depth, direction)


# ===== 計畫性停機評估存證快照（MICS 切片3）=====
# mode=planned 版面按下「存快照」時把當下 impact() 整包結果存證——見 blast_radius.py
# save_snapshot() docstring：圖會變，要拿得出「當初評估說不影響」是哪次算出來的。

class BlastSnapshotBody(BaseModel):
    node_id: str
    reason: str | None = None
    depth: int = 6
    # 2026-08-20 拍板：檢查清單（方案A）要能從事故模式存證，不是只有計畫性停機——
    # mode 只是存證動機的標籤（畫面上顯示用），底層一律算 incident 那份 dependents，
    # 見 save_snapshot()。
    mode: str = "incident"


@app.post("/api/blast/snapshot")
def blast_snapshot_create_endpoint(
    body: BlastSnapshotBody,
    session: sqlite3.Row = Depends(require_auth),
    conn: sqlite3.Connection = Depends(get_db),
):
    import blast_radius
    try:
        return blast_radius.save_snapshot(
            conn, body.node_id, body.mode, body.reason, session["username"], depth=body.depth,
        )
    except ValueError as exc:
        raise HTTPException(404, str(exc)) from exc


@app.get("/api/blast/snapshots")
def blast_snapshot_list_endpoint(
    node_id: str | None = None,
    session: sqlite3.Row = Depends(require_auth),
    conn: sqlite3.Connection = Depends(get_db),
):
    import blast_radius
    return blast_radius.list_snapshots(conn, node_id)


@app.get("/api/blast/snapshot/{snapshot_id}")
def blast_snapshot_get_endpoint(
    snapshot_id: int,
    session: sqlite3.Row = Depends(require_auth),
    conn: sqlite3.Connection = Depends(get_db),
):
    import blast_radius
    snap = blast_radius.get_snapshot(conn, snapshot_id)
    if not snap:
        raise HTTPException(404, "查無此快照")
    return snap


def _blast_result_to_csv(conn: sqlite3.Connection, node_id: str, label: str, result: dict, header_rows: list) -> str:
    """impact() 結果轉 CSV，快照跟即時查詢共用這支。2026-08-19 使用者原話：
    「拿到這張圖第一件事就是把聯絡資料匯出來，散給每個人去盤點」——所以：
    - 受影響節點要顯示人看得懂的名字（label），不是內部 node_id
    - 應通知要帶部門，不然收到清單的人不知道找哪個單位對口
    - 查不到負責人的資產另外列一段，不能讓「不知道」悄悄從匯出檔裡消失
    """
    import blast_radius

    labels = {node_id: label}
    for h in result["dependents"]:
        labels[h["node_id"]] = blast_radius.label_for_node(conn, h["node_id"])

    buf = io.StringIO()
    writer = csv.writer(buf)
    for row in header_rows:
        writer.writerow(row)
    writer.writerow(["查詢節點", node_id, label])
    writer.writerow([])
    writer.writerow(["受影響節點", "距離", "關係", "可信度"])
    for h in result["dependents"]:
        writer.writerow([labels.get(h["node_id"], h["node_id"]), h["depth"], h["edge_type"] or "", h["confidence"]])
    writer.writerow([])
    writer.writerow(["應通知", "部門", "電話", "角色/代理人"])
    for n in result["summary"]["notify"]:
        writer.writerow([
            n.get("name"), n.get("department") or "", n.get("phone") or "",
            n.get("role") or n.get("proxy") or "",
        ])
    if result["summary"]["unknown_owner"]:
        writer.writerow([])
        writer.writerow(["查不到負責人的資產", "資產序號"])
        for u in result["summary"]["unknown_owner"]:
            writer.writerow([u["label"], u["asset_serial"]])
    return "﻿" + buf.getvalue()  # BOM：Excel 開 UTF-8 CSV 中文不亂碼


@app.get("/api/blast/snapshot/{snapshot_id}/csv")
def blast_snapshot_csv_endpoint(
    snapshot_id: int,
    session: sqlite3.Row = Depends(require_auth),
    conn: sqlite3.Connection = Depends(get_db),
):
    import blast_radius
    snap = blast_radius.get_snapshot(conn, snapshot_id)
    if not snap:
        raise HTTPException(404, "查無此快照")

    content = _blast_result_to_csv(
        conn, snap["node_id"], snap["result"]["label"], snap["result"],
        header_rows=[
            ["快照ID", snap["id"]], ["原因", snap["reason"] or ""],
            ["查詢人", snap["asked_by"]], ["查詢時間", snap["asked_at"]],
        ],
    )
    from urllib.parse import quote
    fname = f"停機影響評估_{snap['id']}_{snap['asked_at'][:10]}.csv"
    return Response(
        content=content, media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{quote(fname)}"},
    )


@app.get("/api/blast/impact/csv")
def blast_impact_csv_endpoint(
    node_id: str, depth: int = 6, mode: str = "incident", only_evidence: bool = False,
    session: sqlite3.Row = Depends(require_auth),
    conn: sqlite3.Connection = Depends(get_db),
):
    """即時查詢直接匯出 CSV，不用先存快照——事故當下要的是「馬上把清單發出去」，
    不是「先存證再匯出」（存證是 mode=planned 才需要的流程，事故爆炸半徑不用）。
    """
    import blast_radius
    try:
        result = blast_radius.impact(conn, node_id, depth=depth, mode=mode, only_evidence=only_evidence)
    except ValueError as exc:
        raise HTTPException(404, str(exc)) from exc

    content = _blast_result_to_csv(
        conn, node_id, result["label"], result,
        header_rows=[
            ["查詢時間", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
            ["查詢人", session["username"]],
        ],
    )
    from urllib.parse import quote
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    fname = f"影響範圍_{result['label']}_{stamp}.csv"
    return Response(
        content=content, media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{quote(fname)}"},
    )


# ===== 檢查清單（2026-08-20 拍板方案A）=====
# 事故當下要的是派工清單，不是查詢結果——攤平成「一列一個（主機,聯絡人）配對」，
# 給一個連結，全隊登入都看同一份，改狀態/寫備註就地存。見 blast_radius.py 說明。

class ChecklistCreateBody(BaseModel):
    snapshot_id: int


@app.post("/api/blast/checklist")
def blast_checklist_create_endpoint(
    body: ChecklistCreateBody,
    session: sqlite3.Row = Depends(require_auth),
    conn: sqlite3.Connection = Depends(get_db),
):
    import blast_radius
    try:
        return blast_radius.create_checklist(conn, body.snapshot_id)
    except ValueError as exc:
        raise HTTPException(404, str(exc)) from exc


@app.get("/api/blast/checklist/{snapshot_id}")
def blast_checklist_list_endpoint(
    snapshot_id: int,
    session: sqlite3.Row = Depends(require_auth),
    conn: sqlite3.Connection = Depends(get_db),
):
    import blast_radius
    return blast_radius.list_checklist(conn, snapshot_id)


class ChecklistItemUpdateBody(BaseModel):
    status: str | None = None
    note: str | None = None


@app.put("/api/blast/checklist/item/{item_id}")
def blast_checklist_item_update_endpoint(
    item_id: int, body: ChecklistItemUpdateBody,
    session: sqlite3.Row = Depends(require_auth),
    conn: sqlite3.Connection = Depends(get_db),
):
    import blast_radius
    try:
        return blast_radius.update_checklist_item(conn, item_id, body.status, body.note, session["username"])
    except ValueError as exc:
        raise HTTPException(404, str(exc)) from exc


@app.get("/api/feature-flags")
def list_feature_flags_endpoint(
    session: sqlite3.Row = Depends(require_auth), conn: sqlite3.Connection = Depends(get_db)
):
    """D28：只列出 schema 種好的可切換模組（不含「系統設定」本身，見 schema.sql 註解）。"""
    return [dict(r) for r in list_feature_flags(conn)]


class FeatureFlagBody(BaseModel):
    enabled: bool


@app.put("/api/feature-flags/{module_key}")
def update_feature_flag_endpoint(
    module_key: str,
    body: FeatureFlagBody,
    session: sqlite3.Row = Depends(require_auth),
    conn: sqlite3.Connection = Depends(get_db),
):
    if get_feature_flag(conn, module_key) is None:
        raise HTTPException(404, "查無此功能模組")
    set_feature_flag(conn, module_key, body.enabled)
    return dict(get_feature_flag(conn, module_key))


# ============ 掃描：手動重掃 + 排程設定 ============
@app.post("/api/scan/run")
def scan_run_endpoint(session: sqlite3.Row = Depends(require_auth)):
    """手動重掃：背景執行，馬上回覆，前端輪詢 /api/scan/status 取四態。"""
    if not scan_service.start_scan("manual"):
        raise HTTPException(409, "已有一次掃描正在進行，請稍候")
    return {"started": True}


@app.get("/api/scan/status")
def scan_status_endpoint(
    session: sqlite3.Row = Depends(require_auth),
    conn: sqlite3.Connection = Depends(get_db),
):
    return scan_service.latest_status(conn)


@app.get("/api/scan/schedule")
def scan_schedule_get_endpoint(
    session: sqlite3.Row = Depends(require_auth), conn: sqlite3.Connection = Depends(get_db)
):
    return scan_service.get_schedule(conn)


class ScheduleBody(BaseModel):
    enabled: bool
    mode: str            # daily / interval
    time: str            # HH:MM（daily 用）
    interval_hours: int  # interval 用


@app.put("/api/scan/schedule")
def scan_schedule_put_endpoint(
    body: ScheduleBody,
    session: sqlite3.Row = Depends(require_auth),
    conn: sqlite3.Connection = Depends(get_db),
):
    try:
        scan_service.set_schedule(conn, body.enabled, body.mode, body.time, body.interval_hours)
    except ValueError as exc:
        raise HTTPException(400, str(exc)) from exc
    return scan_service.get_schedule(conn)


# ============ 納入管理（把掃到的主機登記成資產）============
@app.get("/api/cmdb/pull")
def cmdb_pull_endpoint(
    group: str | None = None,
    session: sqlite3.Row = Depends(require_auth),
    conn: sqlite3.Connection = Depends(get_db),
):
    """從 CMDB Gateway 拉資料：回筆數、看到的欄位名（用來對應到我們的欄位）、樣本。
    連不到/認證失敗回 502，錯誤原因如實帶出（不假裝成功）。"""
    try:
        items = cmdb_gateway.fetch_group(conn, group)
    except ConnectionError as exc:
        raise HTTPException(502, str(exc)) from exc
    return {
        "group": group,
        "count": len(items),
        "fields": cmdb_gateway.seen_fields(items),
        "sample": items[:20],
    }


@app.get("/api/export")
def export_assets_endpoint(
    session: sqlite3.Row = Depends(require_auth), conn: sqlite3.Connection = Depends(get_db)
):
    """匯出資產（硬體/人員/軟體三分頁 xlsx）——對應資料匯入，有進就有出。

    表頭優先用「匯入對應表(field_mapping.json)的標題」，保證匯出的檔案能原封不動再匯入
    （round-trip）；欄位依 field_meta 分類排序，讓同一類欄位在表上相鄰、好填。
    """
    base = Path(__file__).parent
    meta_all = json.loads((base / "field_meta.json").read_text(encoding="utf-8"))
    fields_meta = meta_all.get("fields", {})
    cat_order = [c["key"] for c in meta_all.get("categories", [])]
    mapping = json.loads((base / "field_mapping.json").read_text(encoding="utf-8"))

    def _cat_rank(col: str) -> int:
        cat = fields_meta.get(col, {}).get("category")
        return cat_order.index(cat) if cat in cat_order else len(cat_order)

    wb = Workbook()
    first = True
    for table, sheet in (("hardware", "硬體"), ("personnel", "人員"), ("software", "軟體")):
        ws = wb.active if first else wb.create_sheet(sheet)
        if first:
            ws.title = sheet
            first = False
        # 反轉匯入對應表：db欄位 -> 匯入期待的標題（確保 round-trip）。
        # 用 setdefault 保留「第一個出現的標題」當正式標題：field_mapping 允許一個db欄位
        # 掛多個標題別名（例如 api_id 同時吃 "API ID" 與 "AP ID"），若用 {v:k} 推導式會
        # last-wins，匯出表頭會被別名蓋掉、隨設定檔行序漂移。
        rev: dict[str, str] = {}
        for header, column in mapping.get(sheet, {}).items():
            if header != "_comment":
                rev.setdefault(column, header)
        orig = [r[1] for r in conn.execute(f"PRAGMA table_info({table})")]
        skip = {"id", "created_at", "updated_at"}
        orig = [c for c in orig if c not in skip]
        cols = sorted(orig, key=lambda c: (_cat_rank(c), orig.index(c)))
        # 表頭：先用匯入標題(可 round-trip)，退回 field_meta 中文 label，再退回欄名
        ws.append([rev.get(c) or fields_meta.get(c, {}).get("label", c) for c in cols])
        for row in conn.execute(f"SELECT {', '.join(cols)} FROM {table}"):
            ws.append([row[c] for c in cols])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"assets_export_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


@app.get("/api/field-meta")
def field_meta_endpoint(session: sqlite3.Row = Depends(require_auth)):
    """欄位中繼資料（分類×兩層），前端據此把納入管理表單分區、標 tech/biz。
    路徑刻意不放 /api/assets/ 底下——會被 /api/assets/{asset_serial} 這條 path 參數路由吃掉。"""
    return json.loads((Path(__file__).parent / "field_meta.json").read_text(encoding="utf-8"))


@app.get("/api/scan/unregistered")
def unregistered_endpoint(
    session: sqlite3.Row = Depends(require_auth), conn: sqlite3.Connection = Depends(get_db)
):
    """最新一次掃到、但 hardware(CIA) 未登記的主機——納入管理的候選清單。"""
    row = conn.execute("SELECT MAX(scan_time) AS t FROM scan_history WHERE scan_ok = 1").fetchone()
    t = row["t"] if row else None
    if not t:
        return []
    scanned = conn.execute(
        "SELECT ip, hostname, device_model, is_vm, segment, "
        "mac, mac_vendor, open_ports, ttl, os_guess "
        "FROM scan_history WHERE scan_time = ? AND scan_ok = 1",
        (t,),
    ).fetchall()
    out = []
    for r in scanned:
        if not r["ip"] and not r["hostname"]:
            continue
        matched = conn.execute(
            "SELECT 1 FROM hardware WHERE (ip IS NOT NULL AND ip = ?) OR (hostname IS NOT NULL AND hostname = ?)",
            (r["ip"], r["hostname"]),
        ).fetchone()
        if matched is None:
            out.append(dict(r))
    return out


@app.get("/api/scan/results")
def scan_results_endpoint(
    registered: str | None = None,
    session: sqlite3.Row = Depends(require_auth),
    conn: sqlite3.Connection = Depends(get_db),
):
    """最新一次掃描「掃得到（存活）」的主機，每筆標註是否已在 CIA 登記。

    給儀表板的「本次掃描 / 本次掃到存活」磚塊下鑽用。/api/scan/unregistered 只回未登記的
    那一半（納入管理專用），這支回全部並附 registered 旗標，才能對得上磚塊上的總數。

    registered：不帶=全部；yes=已登記；no=未登記（等同 unregistered 那份）。
    """
    if registered is not None and registered not in ("yes", "no"):
        raise HTTPException(400, "registered 只接受 yes 或 no")

    scan_time = _latest_scan_time(conn)
    scanned = _scanned_alive_rows(conn, scan_time)
    if not scanned:
        return {"scan_time": scan_time, "items": []}

    ica_rows = conn.execute("SELECT ip, hostname FROM hardware").fetchall()
    ica_ips = {r["ip"] for r in ica_rows if r["ip"]}
    ica_hostnames = {r["hostname"] for r in ica_rows if r["hostname"]}

    items = []
    for r in scanned:
        if not r["ip"] and not r["hostname"]:
            continue
        is_registered = _row_in_keys(r, ica_ips, ica_hostnames)
        if registered == "yes" and not is_registered:
            continue
        if registered == "no" and is_registered:
            continue
        item = dict(r)
        item["registered"] = is_registered
        # 已登記的話一併帶出資產序號，前端才能直接連到那台的詳細頁
        if is_registered:
            hw = conn.execute(
                "SELECT asset_serial FROM hardware WHERE (ip IS NOT NULL AND ip = ?) "
                "OR (hostname IS NOT NULL AND hostname = ?) LIMIT 1",
                (r["ip"], r["hostname"]),
            ).fetchone()
            item["asset_serial"] = hw["asset_serial"] if hw else None
        items.append(item)

    return {"scan_time": scan_time, "items": items}


class AdoptBody(BaseModel):
    fields: dict


@app.post("/api/assets/adopt")
def adopt_endpoint(
    body: AdoptBody,
    session: sqlite3.Row = Depends(require_auth),
    conn: sqlite3.Connection = Depends(get_db),
):
    """把一台掃到的主機登記成受管資產（technical 前端已帶入、business 使用者填）。"""
    f = {k: v for k, v in body.fields.items() if v not in (None, "")}
    ip = f.get("ip")
    if not ip:
        raise HTTPException(400, "缺少 IP，無法納入管理")
    if conn.execute("SELECT 1 FROM hardware WHERE ip = ?", (ip,)).fetchone():
        raise HTTPException(409, f"{ip} 已納入管理")
    if not f.get("asset_serial"):
        f["asset_serial"] = f"ADOPT-{ip}"   # 沒公司資產序號時先給暫用值，之後可改
    # 資產狀態沒填就預設「使用中」，不要留空。
    # 理由：這台會出現在納管候選清單，正是因為**掃描掃到它活著**——我們明明知道它在跑，
    # 畫面卻顯示「未知」是錯的資訊，不是保守。留空還會讓狀態排序整片沉到最後。
    # （使用者實際反映：納管進來的 6 台全顯示「未知」。）
    if not f.get("asset_status"):
        f["asset_status"] = "使用中"
    cols = {r[1] for r in conn.execute("PRAGMA table_info(hardware)")}
    f = {k: v for k, v in f.items() if k in cols}
    new_id = insert_hardware(conn, **f)

    # 納管當下就撤銷這台的「漏登記」，不要等下一次掃描。
    # 實際踩到（2026-07-19）：使用者納管 5 台之後，儀表板仍顯示「漏登記 6 筆」，
    # 但實際只有 1 台沒登記——因為撤銷只發生在掃描後的重比對。中間這段時間畫面在說謊，
    # 而「剛做完的動作沒有反映在畫面上」是最容易讓人失去信任的一種。
    retracted = conn.execute(
        "UPDATE comparison_result SET is_read = 1, handled_at = datetime('now','localtime') "
        "WHERE issue_type = '漏登記' AND is_read = 0 AND ("
        "  (ip IS NOT NULL AND ip = ?) OR (hostname IS NOT NULL AND hostname != '' AND hostname = ?)"
        ")",
        (ip, f.get("hostname") or ""),
    ).rowcount
    conn.commit()
    return {"id": new_id, "asset_serial": f["asset_serial"], "retracted_issues": retracted}


@app.get("/api/version")
def version_endpoint():
    """版本 + git commit + 服務啟動時間——讓使用者一眼確認「跑的是不是新版」。

    git_commit 直接問 git（服務就跑在 git 工作區上），不再依賴部署時 stamp 的
    build_info.json。理由是實際踩過的坑：stamp 檔由部署腳本寫、版號是每次請求即時讀檔，
    所以「檔案換了但服務沒重啟」時版號照樣跳成新版，看起來部署成功，實際跑的還是舊程式
    （2026-07-18 換 0.6.0 時就是這樣，最後靠 systemd 啟動時間才抓到）。

    started_at 是「本行程的啟動時間」——只有真的重啟過才會變，才是新程式碼有沒有生效的
    可信證據。dirty=true 代表工作區有未提交的改動（就地開發時很常見，不是錯誤）。
    """
    info = {"version": "?", "git_commit": None, "dirty": None,
            "started_at": _PROCESS_STARTED_AT, "built_at": None}
    here = Path(__file__).parent
    try:
        info["version"] = json.loads(
            (here / "version.json").read_text(encoding="utf-8")
        ).get("version", "?")
    except Exception:  # noqa: BLE001
        pass
    try:
        r = subprocess.run(["git", "rev-parse", "--short", "HEAD"], cwd=here,
                           capture_output=True, encoding="utf-8", timeout=5)
        if r.returncode == 0:
            info["git_commit"] = r.stdout.strip()
            d = subprocess.run(["git", "status", "--porcelain"], cwd=here,
                               capture_output=True, encoding="utf-8", timeout=5)
            info["dirty"] = bool(d.stdout.strip()) if d.returncode == 0 else None
    except Exception:  # noqa: BLE001
        pass
    if info["git_commit"] is None:
        # 沒有 git 的環境（例如打包部署到別台）才退回 stamp 檔
        try:
            info.update(json.loads((here / "build_info.json").read_text(encoding="utf-8")))
        except Exception:  # noqa: BLE001
            pass
    return info


# ===== M2 系統聯通圖 =====
class SystemBody(BaseModel):
    id: str
    label: str
    category: str | None = None
    domain: str | None = None
    health: str = "ok"
    is_spof: bool = False
    note: str | None = None


class SystemUpdateBody(BaseModel):
    label: str | None = None
    category: str | None = None
    domain: str | None = None
    health: str | None = None
    is_spof: bool | None = None
    note: str | None = None


class DepBody(BaseModel):
    source: str
    target: str
    dep_type: str | None = None


_HEALTH = {"ok", "warn", "err"}
_ID_RE = re.compile(r"^[A-Za-z0-9_-]{1,32}$")


@app.get("/api/topology")
def get_topology(
    session: sqlite3.Row = Depends(require_auth),
    conn: sqlite3.Connection = Depends(get_db),
):
    """M2 圖資料：一次回 systems + deps，前端餵 Cytoscape。全新建、人維護（戰情室重寫模型）。

    改為需登入（原本讀免登入）：這張圖含系統依賴與 SPOF 標記，等於直接告訴人「打哪一台
    全部會倒」，是最不該公開的資料；頁面本來就在登入牆後，公開讀沒有帶來任何好處。
    健康檢查請改用 /api/version（deploy.sh 用的就是它）。
    """
    import manage_state

    systems = [dict(r) for r in conn.execute("SELECT * FROM systems ORDER BY category, label").fetchall()]
    deps = [dict(r) for r in conn.execute("SELECT * FROM system_deps ORDER BY id").fetchall()]

    # M1↔M2 接起來：系統健康度由關聯主機的實際納管狀態推導，取代人手動標。
    # 沒有關聯主機的系統維持人工值，但明確標成 manual——把「某人半年前填的 ok」
    # 跟「剛剛確認過是 ok」混在一起，等於讓人相信一個沒有根據的綠燈。
    health = manage_state.system_health(conn)
    for s in systems:
        h = health.get(s["id"], {})
        s["health"] = h.get("health", s["health"])
        s["health_source"] = h.get("health_source", "manual")
        s["hosts"] = h.get("hosts", [])
    return {"systems": systems, "deps": deps}


@app.post("/api/systems")
def create_system(
    body: SystemBody,
    session: sqlite3.Row = Depends(require_auth),
    conn: sqlite3.Connection = Depends(get_db),
):
    if not _ID_RE.match(body.id):
        raise HTTPException(400, "系統代碼只能是英數字/底線/減號，1-32 字")
    if body.health not in _HEALTH:
        raise HTTPException(400, "健康度必須是 ok/warn/err")
    if conn.execute("SELECT 1 FROM systems WHERE id = ?", (body.id,)).fetchone():
        raise HTTPException(409, f"系統代碼「{body.id}」已存在")
    conn.execute(
        "INSERT INTO systems (id, label, category, domain, health, is_spof, note) VALUES (?,?,?,?,?,?,?)",
        (body.id, body.label, body.category, body.domain, body.health, int(body.is_spof), body.note),
    )
    conn.commit()
    return {"id": body.id}


@app.put("/api/systems/{system_id}")
def update_system(
    system_id: str,
    body: SystemUpdateBody,
    session: sqlite3.Row = Depends(require_auth),
    conn: sqlite3.Connection = Depends(get_db),
):
    if conn.execute("SELECT 1 FROM systems WHERE id = ?", (system_id,)).fetchone() is None:
        raise HTTPException(404, "查無此系統")
    fields = body.model_dump(exclude_none=True)
    if fields.get("health") and fields["health"] not in _HEALTH:
        raise HTTPException(400, "健康度必須是 ok/warn/err")
    if "is_spof" in fields:
        fields["is_spof"] = int(fields["is_spof"])
    if not fields:
        return {"id": system_id, "updated": 0}
    sets = ", ".join(f"{k} = ?" for k in fields)
    conn.execute(
        f"UPDATE systems SET {sets}, updated_at = datetime('now','localtime') WHERE id = ?",
        [*fields.values(), system_id],
    )
    conn.commit()
    return {"id": system_id, "updated": len(fields)}


@app.delete("/api/systems/{system_id}")
def delete_system(
    system_id: str,
    session: sqlite3.Row = Depends(require_auth),
    conn: sqlite3.Connection = Depends(get_db),
):
    if conn.execute("SELECT 1 FROM systems WHERE id = ?", (system_id,)).fetchone() is None:
        raise HTTPException(404, "查無此系統")
    # 連帶刪掉相關依賴（schema 有 ON DELETE CASCADE，但 SQLite 預設不強制外鍵，這裡明確刪）
    conn.execute("DELETE FROM system_deps WHERE source = ? OR target = ?", (system_id, system_id))
    conn.execute("DELETE FROM systems WHERE id = ?", (system_id,))
    conn.commit()
    return {"id": system_id, "deleted": True}


@app.post("/api/deps")
def create_dep(
    body: DepBody,
    session: sqlite3.Row = Depends(require_auth),
    conn: sqlite3.Connection = Depends(get_db),
):
    if body.source == body.target:
        raise HTTPException(400, "系統不能依賴自己")
    for sid in (body.source, body.target):
        if conn.execute("SELECT 1 FROM systems WHERE id = ?", (sid,)).fetchone() is None:
            raise HTTPException(400, f"系統「{sid}」不存在")
    if conn.execute(
        "SELECT 1 FROM system_deps WHERE source = ? AND target = ?", (body.source, body.target)
    ).fetchone():
        raise HTTPException(409, "這條依賴已存在")
    cur = conn.execute(
        "INSERT INTO system_deps (source, target, dep_type) VALUES (?,?,?)",
        (body.source, body.target, body.dep_type),
    )
    conn.commit()
    return {"id": cur.lastrowid}


@app.delete("/api/deps/{dep_id}")
def delete_dep(
    dep_id: int,
    session: sqlite3.Row = Depends(require_auth),
    conn: sqlite3.Connection = Depends(get_db),
):
    if conn.execute("SELECT 1 FROM system_deps WHERE id = ?", (dep_id,)).fetchone() is None:
        raise HTTPException(404, "查無此依賴")
    conn.execute("DELETE FROM system_deps WHERE id = ?", (dep_id,))
    conn.commit()
    return {"id": dep_id, "deleted": True}


@app.get("/api/search")
def global_search(
    q: str,
    limit: int = 8,
    session: sqlite3.Row = Depends(require_auth),
    conn: sqlite3.Connection = Depends(get_db),
):
    """全域搜尋：一個框找遍資產／服務／業務系統／未登記的掃描結果。

    設計取捨：
    - **分組回傳而不是混成一串**：使用者搜「3306」時，「哪台在聽 3306」跟
      「哪台資產叫 3306」是完全不同的東西，混在一起只會讓人多按一次才找到。
    - **每組限量**：搜尋框是導航用的，不是報表。要看全部就點該組的「看全部」，
      帶著同一個關鍵字跳到對應的清單頁（連結由前端組，後端只回關鍵字命中）。
    - **掃到但未登記的主機也要出現**：那正是「這個 IP 是什麼」最常被問的時候，
      如果只搜已登記資產，答案會是「查無此項」——但它明明就在網路上。
    """
    kw = (q or "").strip()
    if not kw:
        return {"query": "", "groups": []}
    if limit < 1 or limit > 50:
        raise HTTPException(400, "limit 只接受 1–50")
    like = f"%{kw}%"

    groups = []

    assets = conn.execute(
        "SELECT asset_serial, hostname, ip, device_model, environment, asset_purpose "
        "FROM hardware WHERE hostname LIKE ? OR ip LIKE ? OR asset_serial LIKE ? "
        "OR device_model LIKE ? OR asset_purpose LIKE ? OR custodian LIKE ? "
        "ORDER BY hostname LIMIT ?",
        (like, like, like, like, like, like, limit + 1),
    ).fetchall()
    if assets:
        groups.append({
            "key": "assets", "label": "資產",
            "items": [{
                "title": r["hostname"] or r["asset_serial"],
                "subtitle": " · ".join(x for x in (r["ip"], r["device_model"],
                                                   r["environment"]) if x),
                "to": f"/assets/{r['asset_serial']}",
            } for r in assets[:limit]],
            "more": len(assets) > limit,
            "more_to": f"/assets?q={kw}",
        })

    # 服務：搜埠號、行程名、服務名都要中。數字就當埠號精準比，
    # 不然搜 "22" 會把 8022、2222 全撈進來，最想要的那筆反而被淹掉。
    if kw.isdigit():
        svc_rows = conn.execute(
            "SELECT hs.ip, hs.port, hs.process, hs.service_guess, hs.exposure, "
            "hs.asset_serial, h.hostname FROM host_service hs "
            "LEFT JOIN hardware h ON h.asset_serial = hs.asset_serial "
            "WHERE hs.gone_at IS NULL AND hs.port = ? ORDER BY hs.ip LIMIT ?",
            (int(kw), limit + 1),
        ).fetchall()
    else:
        svc_rows = conn.execute(
            "SELECT hs.ip, hs.port, hs.process, hs.service_guess, hs.exposure, "
            "hs.asset_serial, h.hostname FROM host_service hs "
            "LEFT JOIN hardware h ON h.asset_serial = hs.asset_serial "
            "WHERE hs.gone_at IS NULL AND (hs.process LIKE ? OR hs.service_guess LIKE ?) "
            "ORDER BY hs.ip LIMIT ?",
            (like, like, limit + 1),
        ).fetchall()
    if svc_rows:
        groups.append({
            "key": "services", "label": "服務",
            "items": [{
                "title": f"{r['service_guess'] or '未知服務'} · {r['port']}",
                "subtitle": f"{r['hostname'] or r['ip']}"
                            f"{'（僅本機）' if r['exposure'] == 'localhost' else ''}",
                "to": f"/services?port={r['port']}",
            } for r in svc_rows[:limit]],
            "more": len(svc_rows) > limit,
            "more_to": f"/services?port={kw}" if kw.isdigit() else "/services",
        })

    # 「業務系統」合併兩個來源，使用者不需要知道底下分開存：
    # (a) topology.vue 的 systems 表——人工維護的拓樸圖，舊資料
    # (b) hardware.api_id——MICS/blast 用的那份真相（177個真實業務系統代碼），
    #     2026-08-19 使用者反映搜代碼（如「N-218」）或系統名（如「STO 交易管理系統」）
    #     都搜不到，因為這裡原本只查 (a)，根本沒查 hardware.api_id。
    systems = conn.execute(
        "SELECT id, label, category, domain FROM systems "
        "WHERE id LIKE ? OR label LIKE ? OR category LIKE ? OR domain LIKE ? "
        "ORDER BY label LIMIT ?",
        (like, like, like, like, limit + 1),
    ).fetchall()
    biz_rows = conn.execute(
        "SELECT api_id, MIN(asset_name) AS name, COUNT(*) AS cnt FROM hardware "
        "WHERE api_id IS NOT NULL AND api_id != '' AND (api_id LIKE ? OR asset_name LIKE ?) "
        "GROUP BY api_id ORDER BY api_id LIMIT ?",
        (like, like, limit + 1),
    ).fetchall()
    system_items = [{
        "title": r["label"],
        "subtitle": " · ".join(x for x in (r["category"], r["domain"]) if x),
        # 帶 system id：topology 頁會自動選中這個節點，不然搜到系統名字進去
        # 只看到一整張沒有重點標記的圖，等於白搜。
        "to": f"/topology?system={r['id']}",
    } for r in systems]
    biz_items = [{
        "title": r["name"] or r["api_id"],
        "subtitle": f"{r['api_id']} · {r['cnt']} 台主機",
        "to": f"/assets?filter_field=api_id&filter_value={r['api_id']}",
    } for r in biz_rows]
    all_system_items = system_items + biz_items
    if all_system_items:
        groups.append({
            "key": "systems", "label": "業務系統",
            "items": all_system_items[:limit],
            "more": len(all_system_items) > limit,
            "more_to": "/topology",
        })

    # 機櫃：CI圖譜有 rack:{機房}#{櫃號} 節點（真的有依賴邊，不只是位置標籤）——
    # 2026-08-19 使用者反映連機櫃都要找得到，搜到直接帶去 /blast 查「這櫃掉電
    # 會波及誰」，比只列出同機櫃資產清單更有用。
    rack_rows = conn.execute(
        "SELECT physical_location, rack_no, COUNT(*) AS cnt FROM hardware "
        "WHERE physical_location IS NOT NULL AND physical_location != '' "
        "AND rack_no IS NOT NULL AND rack_no != '' "
        "AND (physical_location LIKE ? OR rack_no LIKE ? "
        "OR (physical_location || '/' || rack_no) LIKE ?) "
        "GROUP BY physical_location, rack_no ORDER BY physical_location, rack_no LIMIT ?",
        (like, like, like, limit + 1),
    ).fetchall()
    if rack_rows:
        from urllib.parse import quote as _quote
        groups.append({
            "key": "racks", "label": "機櫃",
            "items": [{
                "title": f"{r['physical_location']} / {r['rack_no']}",
                "subtitle": f"{r['cnt']} 台主機 · 查這櫃掉電會波及誰",
                "to": "/blast?mode=incident&q="
                      + _quote(f"rack:{r['physical_location']}#{r['rack_no']}"),
            } for r in rack_rows[:limit]],
            "more": len(rack_rows) > limit,
            "more_to": "/segments",
        })

    # 人員：2026-08-19 使用者反映「不只業務系統跟機櫃，是全部都要能全文搜」——
    # personnel 是獨立表，搜資產時的 custodian 只涵蓋 hardware.custodian 這個
    # 欄位，找不到用 personnel.person_name 登記的人。點進去帶去他名下的資產清單。
    people = conn.execute(
        "SELECT DISTINCT person_name, phone, belong_division, belong_department "
        "FROM personnel WHERE person_name IS NOT NULL AND person_name != '' "
        "AND (person_name LIKE ? OR phone LIKE ? OR job_desc LIKE ?) "
        "ORDER BY person_name LIMIT ?",
        (like, like, like, limit + 1),
    ).fetchall()
    if people:
        groups.append({
            "key": "people", "label": "人員",
            "items": [{
                "title": r["person_name"],
                "subtitle": " · ".join(
                    x for x in (r["belong_division"], r["belong_department"], r["phone"]) if x
                ),
                "to": f"/assets?filter_field=custodian&filter_value={r['person_name']}",
            } for r in people[:limit]],
            "more": len(people) > limit,
            "more_to": f"/assets?q={kw}",
        })

    # 軟體/資料庫盤點：跟資產是分開的表，搜資產名稱找不到裝在上面的軟體/DB名稱。
    soft = conn.execute(
        "SELECT asset_name, hostname, ip, db_software, asset_serial FROM software "
        "WHERE asset_name LIKE ? OR db_software LIKE ? OR hostname LIKE ? "
        "ORDER BY asset_name LIMIT ?",
        (like, like, like, limit + 1),
    ).fetchall()
    if soft:
        groups.append({
            "key": "software", "label": "軟體／資料庫",
            "items": [{
                "title": r["asset_name"] or r["db_software"] or "（未命名）",
                "subtitle": " · ".join(x for x in (r["db_software"], r["hostname"] or r["ip"]) if x),
                "to": f"/assets/{r['asset_serial']}" if r["asset_serial"] else f"/assets?q={kw}",
            } for r in soft[:limit]],
            "more": len(soft) > limit,
            "more_to": f"/assets?q={kw}",
        })

    # 掃到但沒登記的主機——「這個 IP 是什麼」最常就是在問這種
    scan_time = _latest_scan_time(conn)
    if scan_time:
        known_ips = {r["ip"] for r in conn.execute(
            "SELECT ip FROM hardware WHERE ip IS NOT NULL AND ip != ''")}
        unreg = [r for r in conn.execute(
            "SELECT ip, hostname, os_guess, mac_vendor FROM scan_history "
            "WHERE scan_time = ? AND scan_ok = 1 AND (ip LIKE ? OR hostname LIKE ?) "
            "ORDER BY ip LIMIT ?", (scan_time, like, like, limit + 10))
            if r["ip"] not in known_ips]
        if unreg:
            groups.append({
                "key": "unregistered", "label": "掃到但未登記",
                "items": [{
                    "title": r["ip"],
                    "subtitle": " · ".join(x for x in (r["hostname"], r["os_guess"],
                                                       r["mac_vendor"]) if x) or "無其他線索",
                    "to": "/adopt",
                } for r in unreg[:limit]],
                "more": len(unreg) > limit,
                "more_to": "/adopt",
            })

    # 單據檔案室：連 Word 內文一起搜。使用者原話是「以前要找相關資料，我要翻所有的
    # Word 檔」——全域搜尋不涵蓋單據的話，那個痛點只解決了一半。
    if conn.execute(
        "SELECT name FROM sqlite_master WHERE type='table' AND name='doc_archive'"
    ).fetchone():
        import doc_import

        docs = conn.execute(
            "SELECT id, file_name, doc_type, request_no, ref_request_no, form_date, "
            "hostname, ip, asset_serial, full_text FROM doc_archive "
            "WHERE full_text LIKE ? OR file_name LIKE ? OR request_no LIKE ? "
            "OR ref_request_no LIKE ? OR hostname LIKE ? OR ip LIKE ? "
            "ORDER BY form_date DESC LIMIT ?",
            (like, like, like, like, like, like, limit + 1),
        ).fetchall()
        if docs:
            groups.append({
                "key": "documents", "label": "單據（Word 內文）",
                "items": [{
                    "title": (r["request_no"] or r["ref_request_no"] or r["file_name"])
                             + f"　{r['hostname'] or ''}".rstrip(),
                    # 命中片段直接顯示，使用者不用開檔就知道是不是要找的那份
                    "subtitle": doc_import.snippet(r["full_text"] or "", kw)
                                or " · ".join(x for x in (r["form_date"], r["ip"]) if x),
                    "to": f"/documents?q={kw}",
                } for r in docs[:limit]],
                "more": len(docs) > limit,
                "more_to": f"/documents?q={kw}",
            })

    return {"query": kw, "groups": groups}


# ===== M2 第一片：服務盤點（一台主機看得出什麼服務）=====

ALLOWED_SERVICE_SORT = {
    "ip", "hostname", "port", "proto", "exposure", "process", "service_guess",
    "guess_source", "last_seen", "first_seen", "asset_serial",
}


@app.get("/api/services")
def list_services_endpoint(
    ip: str | None = None,
    asset_serial: str | None = None,
    port: int | None = None,
    exposure: str | None = None,
    include_gone: bool = False,
    include_infra: bool = True,
    sort_by: str = "ip",
    order: str = "asc",
    session: sqlite3.Row = Depends(require_auth),
    conn: sqlite3.Connection = Depends(get_db),
):
    """服務清單。預設不含已消失的，含基礎服務（SSH/NTP…，前端可一鍵收合）。

    排序在後端做（欄位走白名單、方向只認 asc/desc），跟 /api/assets 同一套；
    這張表可能上千列，前端排序只會排到目前拿到的那批。
    """
    import service_inventory

    rows = service_inventory.list_services(
        conn, ip=ip, asset_serial=asset_serial,
        include_gone=include_gone, include_infra=include_infra,
    )
    if port is not None:
        rows = [r for r in rows if r["port"] == port]
    if exposure:
        rows = [r for r in rows if r["exposure"] == exposure]

    key = sort_by if sort_by in ALLOWED_SERVICE_SORT else "ip"
    reverse = order == "desc"

    def sort_val(r):
        v = r.get(key)
        if v is None or v == "":
            return (1, "")           # 空值一律排最後，不隨 asc/desc 翻面
        if key == "ip" and isinstance(v, str) and v.count(".") == 3:
            try:
                return (0, tuple(int(p) for p in v.split(".")))
            except ValueError:
                return (0, v)
        return (0, v if not isinstance(v, str) else v.lower())

    try:
        rows.sort(key=sort_val, reverse=reverse)
    except TypeError:
        rows.sort(key=lambda r: (r.get(key) is None, str(r.get(key) or "")), reverse=reverse)

    return {"items": rows, "summary": service_inventory.service_summary(conn)}


@app.get("/api/services/summary")
def services_summary_endpoint(
    session: sqlite3.Row = Depends(require_auth),
    conn: sqlite3.Connection = Depends(get_db),
):
    import service_inventory

    return service_inventory.service_summary(conn)


@app.post("/api/services/collect")
def collect_services_endpoint(
    asset_serial: str | None = None,
    session: sqlite3.Row = Depends(require_auth),
    conn: sqlite3.Connection = Depends(get_db),
):
    """對已納管主機收一輪服務（可指定單台）。

    ⚠️ 這是同步執行、會真的連出去——大批主機時前端要顯示執行中三態（async-feedback 規範）。
    收不到行程名（沒 root）不算失敗：埠是拿得到的，那已經是主要價值。
    """
    import service_inventory

    try:
        return service_inventory.collect_services(
            conn, only_serial=asset_serial, trigger="manual"
        )
    except Exception as exc:  # noqa: BLE001 - 失敗原因原樣回給畫面，不吞成一句「失敗」
        raise HTTPException(500, f"服務採集失敗：{exc}") from exc


# ===== 帳號盤點（稽核導向）=====

ALLOWED_ACCOUNT_SORT = {
    "ip", "hostname", "username", "uid", "kind", "gecos", "last_login", "pw_last_change",
    "pw_max_days", "pw_status", "is_sudoer", "authorized_keys", "asset_serial",
}


@app.get("/api/accounts")
def list_accounts_endpoint(
    ip: str | None = None,
    asset_serial: str | None = None,
    kind: str | None = None,
    sudoer_only: bool = False,
    hide_builtin: bool = False,
    include_gone: bool = False,
    sort_by: str = "ip",
    order: str = "asc",
    session: sqlite3.Row = Depends(require_auth),
    conn: sqlite3.Connection = Depends(get_db),
):
    import account_inventory

    rows = account_inventory.list_accounts(
        conn, ip=ip, asset_serial=asset_serial, kind=kind,
        include_gone=include_gone, sudoer_only=sudoer_only,
    )

    # 「拉掉內建帳號」：只藏乾淨的系統/內建帳號，被稽核點名的一律留著——
    # 把有 finding 的內建帳號藏起來，正好會蓋掉 UID 0 後門這種最該看的東西。
    # 藏了幾個會回報，不靜默丟（靜默截斷會讓人以為「就這些帳號」）。
    hidden_builtin = 0
    if hide_builtin:
        # 只有「實質違規(fail)」才讓內建帳號留下。不能用全部 finding——
        # 沒 root 時每個帳號都有一堆 unknown(查不到)，那會讓每個帳號都被保護、
        # 篩選器一個都藏不掉（真機 306 條 unknown 就是這情形）。
        # unknown 是「這欄沒查到」不是「這帳號有問題」，不該撐住整列。
        flagged = {(f["ip"], f["username"])
                   for f in account_inventory.latest_findings(conn)
                   if f["verdict"] == "fail"}
        kept = []
        for r in rows:
            if r.get("is_builtin") and (r["ip"], r["username"]) not in flagged:
                hidden_builtin += 1
                continue
            kept.append(r)
        rows = kept

    key = sort_by if sort_by in ALLOWED_ACCOUNT_SORT else "ip"
    reverse = order == "desc"

    def sv(r):
        v = r.get(key)
        if v is None or v == "":
            return (1, "")
        if key == "ip" and isinstance(v, str) and v.count(".") == 3:
            try:
                return (0, tuple(int(p) for p in v.split(".")))
            except ValueError:
                return (0, v)
        return (0, v if not isinstance(v, str) else v.lower())

    try:
        rows.sort(key=sv, reverse=reverse)
    except TypeError:
        rows.sort(key=lambda r: (r.get(key) is None, str(r.get(key) or "")), reverse=reverse)
    return {"items": rows, "summary": account_inventory.audit_summary(conn),
            "hidden_builtin": hidden_builtin}


@app.get("/api/accounts/findings")
def account_findings_endpoint(
    severity: str | None = None,
    rule_id: str | None = None,
    session: sqlite3.Row = Depends(require_auth),
    conn: sqlite3.Connection = Depends(get_db),
):
    """稽核發現。這是這個模組真正的產出——清單只是中間產物。"""
    import account_inventory
    import account_rules

    return {
        "items": account_inventory.latest_findings(conn, severity=severity, rule_id=rule_id),
        "summary": account_inventory.audit_summary(conn),
        "rules": [{"id": r["id"], "label": r["label"], "severity": r["severity"],
                   "law": r["law"]} for r in account_rules.RULES],
        "thresholds": account_rules.get_thresholds(conn),
    }


@app.get("/api/accounts/findings/export")
def export_findings(
    session: sqlite3.Row = Depends(require_auth),
    conn: sqlite3.Connection = Depends(get_db),
):
    """把稽核發現匯出成 Excel——稽核當天要交的是可存檔的證據表，不是叫人看畫面。

    每條含：風險/項目/帳號/帳號備註/主機/IP/判定/處置狀態/例外到期/覆核人/覆核時間/
    手動備註/法規依據。含處置狀態＝這份報告本身就是稽核軌跡。
    """
    import account_inventory

    findings = account_inventory.latest_findings(conn)
    sev_label = {"high": "高", "medium": "中", "low": "低"}
    status_label = {"open": "待處理", "ack": "已確認", "exception": "核准例外", "fixed": "已修復"}

    wb = Workbook()
    ws = wb.active
    ws.title = "帳號稽核發現"
    ws.append(["風險", "項目", "帳號", "帳號備註", "主機", "IP", "判定",
               "處置狀態", "例外到期", "覆核人", "覆核時間", "手動備註", "法規依據"])
    for f in findings:
        ws.append([
            "查不到" if f.get("verdict") == "unknown" else sev_label.get(f.get("severity"), ""),
            f.get("label"), f.get("username"), f.get("gecos"),
            f.get("hostname") or "", f.get("ip"), f.get("detail"),
            status_label.get(f.get("status"), f.get("status")),
            f.get("exempt_until") or "", f.get("decided_by") or "", f.get("decided_at") or "",
            f.get("note") or "", f.get("law"),
        ])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"帳號稽核發現_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    from urllib.parse import quote
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{quote(fname)}"},
    )


@app.get("/api/accounts/matrix/export")
def export_matrix(
    cols: list[str] = Query(default=[]),
    kind: str | None = None,
    hide_builtin: bool = False,
    session: sqlite3.Row = Depends(require_auth),
    conn: sqlite3.Connection = Depends(get_db),
):
    """合規表匯出：使用者自選欄位，資料最小化給稽核看。

    只吐勾選的欄位——稽核不需要看到 UID、shell、金鑰把數這些內部細節，
    給越少越好。狀態欄的計算走 account_inventory.MATRIX_EXPORT_COLS（與前端同源）。
    """
    import account_inventory

    valid = account_inventory.MATRIX_EXPORT_COLS
    chosen = [c for c in cols if c in valid]
    if not chosen:
        raise HTTPException(400, "至少要選一個有效欄位")

    rows = account_inventory.list_accounts(conn, kind=kind or None)
    if hide_builtin:
        flagged = {(f["ip"], f["username"])
                   for f in account_inventory.latest_findings(conn) if f["verdict"] == "fail"}
        rows = [r for r in rows
                if not (r.get("is_builtin") and (r["ip"], r["username"]) not in flagged)]

    wb = Workbook()
    ws = wb.active
    ws.title = "帳號合規表"
    ws.append([valid[c][0] for c in chosen])
    for r in rows:
        ws.append([valid[c][1](r) for c in chosen])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"帳號合規表_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    from urllib.parse import quote
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{quote(fname)}"},
    )


@app.post("/api/accounts/collect")
def collect_accounts_endpoint(
    asset_serial: str | None = None,
    session: sqlite3.Row = Depends(require_auth),
    conn: sqlite3.Connection = Depends(get_db),
):
    import account_inventory

    try:
        return account_inventory.collect_accounts(
            conn, only_serial=asset_serial, trigger="manual")
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(500, f"帳號採集失敗：{exc}") from exc


class CollectAccountBody(BaseModel):
    account: str
    collector_ip: str | None = None      # 空字串＝清掉設定、退回自動偵測


# 允許的收集身分：唯讀預設 + 已知標準管理帳號。不開放任意字串——
# 收集身分決定拿多大權限，不該讓它變成打字就能指定任意帳號的欄位。
def _allowed_collect_accounts() -> set[str]:
    import account_collector

    return {"webit3scan"} | set(account_collector.STD_MGMT_ACCOUNTS.keys())


@app.get("/api/accounts/collect-config")
def get_collect_config(
    session: sqlite3.Row = Depends(require_auth),
    conn: sqlite3.Connection = Depends(get_db),
):
    """目前的遠端收集身分，以及切成管理帳號時要先做的授權動作。

    webit3 用自己的金鑰、但以設定的身分登入——所以切成 sysinfra 前，要先把
    webit3 收集公鑰授權進各機的 sysinfra/.ssh/authorized_keys（一次性、可撤）。
    這裡回傳那把公鑰與現成指令，讓使用者用既有管道（巡檢 ansible）佈下去。
    """
    import manage_state

    current = manage_state.get_collect_account(conn)
    # 走同一支解析函式：沒有金鑰就當場產一把，不要叫人去命令列跑腳本
    # （使用者 2026-08-16 指正：這是畫面功能的前提，不該依賴一次人工動作）
    try:
        pubkey = onboard_engine.collector_pubkey()
    except ValueError as exc:
        pubkey = f"（{exc}）"
    collector_ip = onboard_engine.resolve_collector_ip(conn)
    from db import get_setting

    ip_setting = (get_setting(conn, onboard_engine.COLLECTOR_IP_SETTING, "") or "").strip()
    return {
        "account": current,
        # 收集器自己的位址：會被寫進目標主機 authorized_keys 的 from= 來源限制。
        # 填錯不會報錯，只會讓金鑰永遠被拒而納管顯示成功——所以畫面要看得到它是誰、
        # 從哪來的（畫面設定／環境變數／自動偵測）。
        "collector_ip": collector_ip,
        "collector_ip_source": ("畫面設定" if ip_setting
                                else "環境變數" if os.environ.get("ASSET_COLLECTOR_IP")
                                else "自動偵測"),
        "collector_ip_detected": onboard_engine.detect_collector_ip(),
        "options": [
            {"value": "webit3scan", "label": "webit3scan（唯讀最小權限）",
             "note": "看不到需 root 的欄位（密碼效期、sudo 明細、authorized_keys）"},
            {"value": "sysinfra", "label": "sysinfra（標準管理帳號，完整資料）",
             "note": "NOPASSWD:ALL，sudo -n 全通；需先授權收集公鑰進 sysinfra"},
        ],
        "pubkey": pubkey,
        "provision_hint": (
            "切成 sysinfra 前，先在各主機把上面這把公鑰加進 "
            "~sysinfra/.ssh/authorized_keys（用巡檢 ansible 一次佈完）。"
            "webit3 不持有 sysinfra 私鑰，授權隨時可撤（移掉該行即失效）。"
        ),
    }


@app.put("/api/accounts/collect-config")
def set_collect_config(
    body: CollectAccountBody,
    session: sqlite3.Row = Depends(require_auth),
    conn: sqlite3.Connection = Depends(get_db),
):
    from db import set_setting

    if body.account not in _allowed_collect_accounts():
        raise HTTPException(
            400, f"不支援的收集身分：{body.account}（只接受 "
                 f"{', '.join(sorted(_allowed_collect_accounts()))}）")
    set_setting(conn, "collect_ssh_account", body.account)

    if body.collector_ip is not None:
        val = body.collector_ip.strip()
        if val:
            try:
                onboard_engine.validate_collector_ip(val)
            except ValueError as exc:
                raise HTTPException(400, str(exc))
        # 空字串＝清掉，退回自動偵測（不是把空值當成「設成空的」）
        set_setting(conn, onboard_engine.COLLECTOR_IP_SETTING, val)
    return {"account": body.account,
            "collector_ip": onboard_engine.resolve_collector_ip(conn)}


@app.get("/api/accounts/hosts")
def list_account_hosts(
    session: sqlite3.Row = Depends(require_auth),
    conn: sqlite3.Connection = Depends(get_db),
):
    """可收集的主機清單＋是否已排除。給排除管理與健檢選單用。"""
    import account_inventory

    return {"hosts": account_inventory.list_collectable_hosts(conn)}


class DispositionBody(BaseModel):
    ip: str
    username: str
    rule_id: str
    status: str
    note: str | None = None
    exempt_until: str | None = None


@app.put("/api/accounts/findings/disposition")
def set_finding_disposition_endpoint(
    body: DispositionBody,
    session: sqlite3.Row = Depends(require_auth),
    conn: sqlite3.Connection = Depends(get_db),
):
    """設一條稽核發現的處置狀態（待處理/已確認/核准例外/已修復）。跨盤點持久。

    decided_by 記登入者——稽核要能查「這條例外是誰、何時核准的」。
    """
    import account_inventory

    try:
        return account_inventory.set_finding_disposition(
            conn, body.ip, body.username, body.rule_id, body.status,
            note=body.note, exempt_until=body.exempt_until,
            decided_by=session["username"])
    except ValueError as exc:
        raise HTTPException(400, str(exc)) from exc


class AccountNoteBody(BaseModel):
    ip: str
    username: str
    note: str = ""


@app.put("/api/accounts/note")
def set_account_note_endpoint(
    body: AccountNoteBody,
    session: sqlite3.Row = Depends(require_auth),
    conn: sqlite3.Connection = Depends(get_db),
):
    """設某帳號的「手動備註」——稽核人員自己輸入的註記（跟 gecos 自動備註是兩回事）。"""
    import account_inventory

    try:
        return account_inventory.set_account_note(conn, body.ip, body.username, body.note)
    except ValueError as exc:
        raise HTTPException(404, str(exc)) from exc


class ExcludeBody(BaseModel):
    asset_serial: str
    exclude: bool


@app.put("/api/accounts/exclude")
def set_account_exclude(
    body: ExcludeBody,
    session: sqlite3.Row = Depends(require_auth),
    conn: sqlite3.Connection = Depends(get_db),
):
    """把一台主機排除／納回帳號稽核。排除時清掉它的舊帳號與稽核資料。"""
    import account_inventory

    if conn.execute("SELECT 1 FROM hardware WHERE asset_serial = ?",
                    (body.asset_serial,)).fetchone() is None:
        raise HTTPException(404, "查無此資產")
    return account_inventory.set_host_excluded(conn, body.asset_serial, body.exclude)


@app.get("/api/accounts/diagnose")
def diagnose_account_host(
    asset_serial: str,
    desensitize: bool = True,
    session: sqlite3.Row = Depends(require_auth),
    conn: sqlite3.Connection = Depends(get_db),
):
    """逐主機收集健檢：即時跑每條收集指令，回報為什麼收不到。

    給公司多 OS 上線 debug 用。安全：不含任何原始輸出內容（見 collect_probe），
    再走 diagnostics 的去識別化＋殘留掃描閘門，可安全匯出貼給開發者。
    desensitize 預設 True——真實資料不出這台機器是硬規則。
    """
    import collect_probe
    import diagnostics as dg
    import manage_state

    row = conn.execute(
        "SELECT ip FROM hardware WHERE asset_serial = ?", (asset_serial,)
    ).fetchone()
    if not row or not row["ip"]:
        raise HTTPException(404, "查無此資產或它沒有 IP")

    account = manage_state.get_collect_account(conn)
    try:
        report = collect_probe.probe(
            row["ip"], account, manage_state.COLLECTOR_KEY_DEFAULT,
            local_ips=manage_state.local_ips())
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(500, f"健檢執行失敗：{exc}") from exc

    if desensitize:
        d = dg.Desensitizer(enabled=True)
        report = d.walk(report)
        residual = dg.residual_scan(json.dumps(report, ensure_ascii=False))
        report["_desensitized"] = True
        report["_residual_scan"] = residual or "通過（未發現未遮蔽的位址）"
        # 殘留掃描沒過就不給——真實位址不出這台機器
        if residual:
            raise HTTPException(
                500, f"健檢輸出殘留掃描未通過，已擋下不外送：{residual[:3]}")
    return report


@app.get("/api/accounts/sudo-rules")
def account_sudo_rules(session: sqlite3.Row = Depends(require_auth)):
    """收集帳號要拿到密碼/sudo/金鑰資訊所需的 sudo 白名單。

    刻意做成「給你看、你自己決定要不要佈」而不是自動佈署：
    這是擴權動作，且會改動所有已納管主機的 sudoers。
    白名單本身不含 `cat /etc/shadow`——chage -l／passwd -S 拿得到同樣的稽核結論
    卻不吐密碼雜湊，能用小權限達成就不該要大的。
    """
    import account_collector

    return {
        "rules": account_collector.SUDO_RULES,
        "note": "存成 /etc/sudoers.d/webit3scan-audit（chmod 440），"
                "並用 visudo -cf 驗證後才生效。不含 /etc/shadow，不會洩漏密碼雜湊。",
    }


@app.on_event("startup")
def _start_scheduler():
    # 只在正式服務啟動時開排程器（webit3-api.service 設 ASSET_SCHEDULER=1）；
    # 測試/開發匯入 app 時不啟動，避免背景亂掃。
    if os.environ.get("ASSET_SCHEDULER") == "1":
        scan_service.start_scheduler()
